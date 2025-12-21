# -*- coding: utf-8 -*-
"""
智能条款比对工具 v15.0 (External Config Edition)
- [核心升级] 配置外部化 - 支持 JSON 配置文件
- [继承] v14.0 全部功能（多级匹配、客户映射）
- [新增] 运行时添加映射并保存
- [新增] 配置统计显示

Author: Dachi Yijin
Date: 2025-12-21
"""

import sys
import os
import re
import difflib
import traceback
from typing import List, Dict, Tuple, Optional, Set
from dataclasses import dataclass
from enum import Enum
import pandas as pd
from docx import Document

# ==========================================
# 导入配置管理器
# ==========================================
from clause_config_manager import get_config, ClauseConfigManager

# ==========================================
# macOS PyQt5 Plugin Fix
# ==========================================
try:
    import PyQt5
    plugin_path = os.path.join(os.path.dirname(PyQt5.__file__), 'Qt5', 'plugins')
    os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path
except ImportError:
    pass

try:
    from deep_translator import GoogleTranslator
    HAS_TRANSLATOR = True
except ImportError:
    HAS_TRANSLATOR = False

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QProgressBar, QTextEdit,
    QFileDialog, QMessageBox, QFrame, QGraphicsDropShadowEffect,
    QDialog, QFormLayout, QDialogButtonBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt5.QtGui import QFont, QColor, QDesktopServices, QTextCursor

# ==========================================
# macOS 打包防闪退
# ==========================================
class NullWriter:
    def write(self, text): pass
    def flush(self): pass

if getattr(sys, 'frozen', False):
    sys.stdout = NullWriter()
    sys.stderr = NullWriter()

def global_exception_handler(exctype, value, tb):
    error_msg = "".join(traceback.format_exception(exctype, value, tb))
    try: sys.__stderr__.write(error_msg)
    except: pass
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Critical)
    msg_box.setText("程序发生意外错误")
    msg_box.setInformativeText(str(value))
    msg_box.setDetailedText(error_msg)
    msg_box.exec_()

sys.excepthook = global_exception_handler


# ==========================================
# 数据结构
# ==========================================
class MatchLevel(Enum):
    """匹配级别"""
    EXACT = "精确匹配"
    SEMANTIC = "语义匹配"
    KEYWORD = "关键词匹配"
    FUZZY = "模糊匹配"
    NONE = "无匹配"

@dataclass
class ClauseItem:
    """条款项"""
    title: str
    content: str
    original_title: str = ""

@dataclass
class MatchResult:
    """匹配结果"""
    matched_name: str = ""
    matched_content: str = ""
    matched_reg: str = ""
    score: float = 0.0
    title_score: float = 0.0
    content_score: float = 0.0
    match_level: MatchLevel = MatchLevel.NONE
    diff_analysis: str = ""


# ==========================================
# 核心匹配逻辑（使用配置管理器）
# ==========================================
class ClauseMatcherLogic:
    """条款匹配核心逻辑 - 使用外部配置"""

    def __init__(self):
        # 获取配置管理器实例
        self.config = get_config()

    def normalize_text(self, text: str) -> str:
        """标准化文本"""
        if not isinstance(text, str):
            return ""
        text = text.lower().strip()
        text = re.sub(r"['\"\'\'\"\"\(\)（）\[\]【】]", '', text)
        text = re.sub(r'\s+', ' ', text)
        return text

    def clean_title(self, text: str) -> str:
        """清理标题用于比较"""
        if not isinstance(text, str):
            return ""
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        # 使用配置管理器清理噪音词
        text = self.config.clean_noise_words(text)
        text = re.sub(r'[0-9\s]+', '', text)
        return text.strip()

    def clean_content(self, text: str) -> str:
        """清理内容用于比较"""
        if not isinstance(text, str):
            return ""
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        text = re.sub(r'\s+', '', text)
        text = re.sub(r'[0-9]+', '', text)
        return text

    def extract_extra_info(self, text: str) -> str:
        """提取括号内的额外信息"""
        if not isinstance(text, str):
            return ""
        matches = re.findall(r'([\(（].*?[\)）])', text)
        return " ".join(matches) if matches else ""

    def is_english(self, text: str) -> bool:
        """判断是否为英文"""
        if not isinstance(text, str) or len(text) <= 3:
            return False
        zh_count = len(re.findall(r'[\u4e00-\u9fa5]', text))
        return zh_count < len(text) * 0.15

    def translate_title(self, title: str) -> Tuple[str, bool]:
        """翻译英文标题为中文"""
        if not self.is_english(title):
            return title, False

        title_norm = self.normalize_text(title)

        # 1. 使用配置管理器查询映射
        mapped = self.config.get_client_mapping(title_norm)
        if mapped:
            return mapped, True

        # 2. 部分匹配
        for eng, chn in self.config.client_en_cn_map.items():
            if eng in title_norm or title_norm in eng:
                return chn, True

        # 3. 在线翻译
        if HAS_TRANSLATOR:
            try:
                translated = GoogleTranslator(source='auto', target='zh-CN').translate(title)
                return translated, True
            except:
                pass

        return title, False

    def calculate_similarity(self, text1: str, text2: str) -> float:
        """计算文本相似度"""
        if not text1 or not text2:
            return 0.0
        return difflib.SequenceMatcher(None, text1, text2).ratio()

    def match_clause(self, clause: ClauseItem, lib_data: List[Dict],
                     is_title_only: bool) -> MatchResult:
        """
        多级匹配策略（使用外部配置）
        """
        result = MatchResult()
        title = clause.title
        content = clause.content

        title_clean = self.clean_title(title)
        title_norm = self.normalize_text(title)

        best_score = -100
        best_match = None
        best_meta = {'t': 0, 'c': 0, 'level': MatchLevel.NONE}

        # 使用配置管理器提取关键词
        c_keywords = self.config.get_keywords_for_text(title)

        # 检查语义别名
        semantic_target = self.config.get_semantic_alias(title)

        # 检查精确条款映射
        exact_target = self.config.get_exact_clause_mapping(title)
        if not exact_target:
            exact_target = self.config.get_exact_clause_mapping(title_clean)

        # 获取阈值配置
        thresholds = self.config.thresholds

        for lib in lib_data:
            l_name = str(lib.get('条款名称', ''))
            l_content = str(lib.get('条款内容', ''))
            l_name_clean = self.clean_title(l_name)
            l_name_norm = self.normalize_text(l_name)

            score = 0.0
            match_level = MatchLevel.FUZZY

            # === 级别0: 精确条款名映射 ===
            if exact_target and exact_target in l_name:
                score = thresholds.exact_min
                match_level = MatchLevel.EXACT
                best_score = score
                best_match = lib
                best_meta = {'t': score, 'c': 0, 'level': MatchLevel.EXACT}
                break

            # === 级别1: 精确匹配 ===
            if title_clean == l_name_clean or title_norm == l_name_norm:
                score = 1.0
                match_level = MatchLevel.EXACT

            # === 级别2: 语义别名匹配 ===
            elif semantic_target and semantic_target in l_name:
                score = thresholds.semantic_min
                match_level = MatchLevel.SEMANTIC

            else:
                # === 级别3: 关键词匹配 ===
                l_keywords = self.config.get_keywords_for_text(l_name)
                if c_keywords and l_keywords:
                    common = c_keywords & l_keywords
                    if common:
                        keyword_score = len(common) / max(len(c_keywords), len(l_keywords))
                        if keyword_score >= 0.5:
                            score = thresholds.keyword_min + keyword_score * 0.2
                            match_level = MatchLevel.KEYWORD

                # === 级别4: 模糊匹配 ===
                if score < thresholds.keyword_min:
                    title_sim = self.calculate_similarity(title_clean, l_name_clean)

                    content_sim = 0.0
                    if not is_title_only and content.strip():
                        c_content_clean = self.clean_content(content)
                        l_content_clean = self.clean_content(l_content)
                        if c_content_clean and l_content_clean:
                            content_sim = self.calculate_similarity(c_content_clean, l_content_clean)

                    if is_title_only or not content.strip():
                        score = title_sim
                    else:
                        score = 0.7 * title_sim + 0.3 * content_sim

                    best_meta['t'] = title_sim
                    best_meta['c'] = content_sim
                    match_level = MatchLevel.FUZZY

            # 使用配置检查惩罚关键词
            if self.config.is_penalty_keyword(l_name) and not self.config.is_penalty_keyword(title):
                score -= 0.5

            if score > best_score:
                best_score = score
                best_match = lib
                best_meta['level'] = match_level
                if match_level in [MatchLevel.EXACT, MatchLevel.SEMANTIC, MatchLevel.KEYWORD]:
                    best_meta['t'] = score

        # 构建结果
        if best_match and best_score > thresholds.accept_min:
            base_name = best_match.get('条款名称', '')
            extra_params = self.extract_extra_info(clause.original_title or clause.title)

            result.matched_name = f"{base_name} {extra_params}".strip() if extra_params else base_name
            result.matched_content = best_match.get('条款内容', '')
            result.matched_reg = best_match.get('产品注册号', best_match.get('注册号', ''))
            result.score = max(0, best_score)
            result.title_score = best_meta.get('t', 0)
            result.content_score = best_meta.get('c', 0)
            result.match_level = best_meta.get('level', MatchLevel.FUZZY)

            if best_score < 0.6:
                result.diff_analysis = self.analyze_difference(content, result.matched_content)

        return result

    def analyze_difference(self, c_content: str, l_content: str) -> str:
        """分析保障差异"""
        c_text, l_text = str(c_content), str(l_content)
        if not c_text.strip():
            return ""

        analysis = []
        keywords = {
            "限额": ["Limit", "限额", "最高", "limit"],
            "免赔": ["Deductible", "Excess", "免赔", "deductible"],
            "除外": ["Exclusion", "除外", "不负责", "exclusion"],
            "观察期": ["Waiting Period", "观察期", "等待期"],
        }

        for key, words in keywords.items():
            c_has = any(w.lower() in c_text.lower() for w in words)
            l_has = any(w.lower() in l_text.lower() for w in words)
            if c_has and not l_has:
                analysis.append(f"⚠️ 客户提及[{key}]但库内未提及")
            elif not c_has and l_has:
                analysis.append(f"ℹ️ 库内包含[{key}]但客户未提及")

        return " | ".join(analysis)

    def is_likely_title(self, text: str) -> bool:
        """判断是否像标题"""
        if len(text) > 80:
            return False
        if text.endswith(('。', '；', '.', ';')):
            return False
        title_indicators = ["条款", "Clause", "Extension", "险", "CLAUSE", "EXTENSION"]
        if any(kw in text for kw in title_indicators):
            return True
        if text.isupper() and len(text) > 5:
            return True
        return True

    def parse_docx(self, doc_path: str) -> Tuple[List[ClauseItem], bool]:
        """解析Word文档"""
        doc = Document(doc_path)
        clauses = []
        current_block = []

        all_lines = [p.text.strip() for p in doc.paragraphs]
        empty_lines = sum(1 for t in all_lines if not t)

        use_smart_split = len(all_lines) > 0 and (empty_lines / max(len(all_lines), 1) < 0.05)

        if use_smart_split:
            for text in all_lines:
                if not text:
                    continue
                if current_block and self.is_likely_title(text):
                    title = current_block[0]
                    content = "\n".join(current_block[1:])
                    clauses.append(ClauseItem(title=title, content=content, original_title=title))
                    current_block = [text]
                else:
                    current_block.append(text)
            if current_block:
                clauses.append(ClauseItem(
                    title=current_block[0],
                    content="\n".join(current_block[1:]),
                    original_title=current_block[0]
                ))
        else:
            for text in all_lines:
                if text:
                    current_block.append(text)
                elif current_block:
                    clauses.append(ClauseItem(
                        title=current_block[0],
                        content="\n".join(current_block[1:]),
                        original_title=current_block[0]
                    ))
                    current_block = []
            if current_block:
                clauses.append(ClauseItem(
                    title=current_block[0],
                    content="\n".join(current_block[1:]),
                    original_title=current_block[0]
                ))

        is_title_only = all(not c.content for c in clauses)
        return clauses, is_title_only


# ==========================================
# 添加映射对话框
# ==========================================
class AddMappingDialog(QDialog):
    """添加新映射对话框"""

    def __init__(self, parent=None, english_term: str = "", chinese_term: str = ""):
        super().__init__(parent)
        self.setWindowTitle("添加术语映射")
        self.setMinimumWidth(400)
        self.setStyleSheet("""
            QDialog {
                background: #1a1a2e;
            }
            QLabel {
                color: #ffffff;
                font-size: 14px;
            }
            QLineEdit {
                background: rgba(255,255,255,0.1);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 8px;
                padding: 10px;
                color: #ffffff;
                font-size: 14px;
            }
            QLineEdit:focus {
                border-color: #667eea;
            }
            QPushButton {
                background: #667eea;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #764ba2;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        form = QFormLayout()
        form.setSpacing(10)

        self.eng_input = QLineEdit(english_term)
        self.eng_input.setPlaceholderText("例如: reinstatement value")
        form.addRow("英文术语:", self.eng_input)

        self.chn_input = QLineEdit(chinese_term)
        self.chn_input.setPlaceholderText("例如: 重置价值条款")
        form.addRow("中文翻译:", self.chn_input)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.accept)

        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("background: rgba(255,255,255,0.1);")
        cancel_btn.clicked.connect(self.reject)

        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)

    def get_mapping(self) -> Tuple[str, str]:
        return self.eng_input.text().strip(), self.chn_input.text().strip()


# ==========================================
# 工作线程
# ==========================================
class MatchWorker(QThread):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, doc_path: str, excel_path: str, output_path: str):
        super().__init__()
        self.doc_path = doc_path
        self.excel_path = excel_path
        self.output_path = output_path

    def run(self):
        try:
            logic = ClauseMatcherLogic()
            config = logic.config

            # 显示配置统计
            stats = config.get_stats()
            self.log_signal.emit(f"📊 配置统计: {stats['client_mappings']} 个英中映射, "
                               f"{stats['semantic_aliases']} 个语义别名", "info")

            if not HAS_TRANSLATOR:
                self.log_signal.emit("⚠️ 未检测到 deep_translator", "warning")
            else:
                self.log_signal.emit("✓ 已启用在线翻译支持", "success")

            self.log_signal.emit("⏳ 正在解析文档...", "info")
            clauses, is_title_only = logic.parse_docx(self.doc_path)
            mode_str = "纯标题模式" if is_title_only else "完整内容模式"
            self.log_signal.emit(f"📖 [{mode_str}] 提取到 {len(clauses)} 条", "success")

            # 加载条款库
            lib_df = pd.read_excel(self.excel_path, header=1)
            lib_df.columns = [str(c).strip() for c in lib_df.columns]

            name_col, content_col, reg_col = None, None, None
            for col in lib_df.columns:
                if '条款名称' in col or '名称' in col:
                    name_col = col
                elif '条款内容' in col or '内容' in col:
                    content_col = col
                elif '注册号' in col or '产品' in col:
                    reg_col = col

            if not name_col:
                name_col = lib_df.columns[0]
            if not content_col and len(lib_df.columns) > 2:
                content_col = lib_df.columns[2]
            if not reg_col and len(lib_df.columns) > 1:
                reg_col = lib_df.columns[1]

            lib_data = []
            for _, row in lib_df.iterrows():
                lib_data.append({
                    '条款名称': str(row.get(name_col, '')) if pd.notna(row.get(name_col)) else '',
                    '条款内容': str(row.get(content_col, '')) if content_col and pd.notna(row.get(content_col)) else '',
                    '产品注册号': str(row.get(reg_col, '')) if reg_col and pd.notna(row.get(reg_col)) else '',
                })

            lib_data = [d for d in lib_data if d['条款名称'].strip()]
            self.log_signal.emit(f"📚 加载条款库 {len(lib_data)} 条", "info")

            self.log_signal.emit("🧠 开始智能匹配...", "info")
            results = []
            stats = {'exact': 0, 'semantic': 0, 'keyword': 0, 'fuzzy': 0, 'none': 0}

            for idx, clause in enumerate(clauses, 1):
                self.progress_signal.emit(idx, len(clauses))

                original_title = clause.title
                translated_title, was_translated = logic.translate_title(clause.title)

                if was_translated:
                    clause.title = translated_title
                    clause.original_title = original_title

                match_result = logic.match_clause(clause, lib_data, is_title_only)

                if match_result.match_level == MatchLevel.EXACT:
                    stats['exact'] += 1
                elif match_result.match_level == MatchLevel.SEMANTIC:
                    stats['semantic'] += 1
                elif match_result.match_level == MatchLevel.KEYWORD:
                    stats['keyword'] += 1
                elif match_result.match_level == MatchLevel.FUZZY:
                    stats['fuzzy'] += 1
                else:
                    stats['none'] += 1

                results.append({
                    '序号': idx,
                    '客户条款(原)': original_title,
                    '客户条款(译)': translated_title if was_translated else "",
                    '客户原始内容': clause.content[:500] if clause.content else "",
                    '匹配条款库名称': match_result.matched_name or "无匹配",
                    '产品注册号': match_result.matched_reg,
                    '匹配条款库内容': match_result.matched_content[:500] if match_result.matched_content else "",
                    '综合匹配度': round(match_result.score, 3),
                    '匹配级别': match_result.match_level.value,
                    '保障差异提示': match_result.diff_analysis,
                    '标题相似度': round(match_result.title_score, 3),
                    '内容相似度': round(match_result.content_score, 3),
                })

            df_res = pd.DataFrame(results)
            df_res.to_excel(self.output_path, index=False)
            self._apply_excel_styles()

            self.log_signal.emit(f"📊 匹配统计:", "info")
            self.log_signal.emit(f"   精确匹配: {stats['exact']}", "success")
            self.log_signal.emit(f"   语义匹配: {stats['semantic']}", "success")
            self.log_signal.emit(f"   关键词匹配: {stats['keyword']}", "info")
            self.log_signal.emit(f"   模糊匹配: {stats['fuzzy']}", "warning")
            self.log_signal.emit(f"   无匹配: {stats['none']}", "error")

            self.log_signal.emit(f"🎉 完成！已生成报告", "success")
            self.finished_signal.emit(True, self.output_path)

        except Exception as e:
            self.log_signal.emit(f"❌ 错误: {str(e)}", "error")
            self.finished_signal.emit(False, str(e))

    def _apply_excel_styles(self):
        """应用Excel样式"""
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb.active

        fills = {
            'green': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'yellow': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'blue': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            'header': PatternFill(start_color="667eea", end_color="667eea", fill_type="solid"),
        }

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = fills['header']
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        widths = {'A': 6, 'B': 35, 'C': 30, 'D': 45, 'E': 40,
                  'F': 25, 'G': 50, 'H': 10, 'I': 12, 'J': 35, 'K': 10, 'L': 10}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border

                if cell.col_idx == 8:
                    try:
                        val = float(cell.value) if cell.value else 0
                        if val >= 0.8:
                            cell.fill = fills['green']
                        elif val >= 0.5:
                            cell.fill = fills['yellow']
                        elif val > 0:
                            cell.fill = fills['red']
                    except:
                        pass

                if cell.col_idx == 9:
                    val = str(cell.value) if cell.value else ""
                    if "精确" in val:
                        cell.fill = fills['green']
                    elif "语义" in val:
                        cell.fill = fills['blue']
                    elif "关键词" in val:
                        cell.fill = fills['yellow']

        ws.freeze_panes = 'A2'
        wb.save(self.output_path)


# ==========================================
# UI组件
# ==========================================
class GlassCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            GlassCard {
                background: rgba(255, 255, 255, 0.08);
                border: 1px solid rgba(255, 255, 255, 0.15);
                border-radius: 20px;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(40)
        shadow.setColor(QColor(0, 0, 0, 80))
        shadow.setOffset(0, 10)
        self.setGraphicsEffect(shadow)


class ClauseDiffGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("智能条款比对工具 v15.0")
        self.setMinimumSize(900, 800)
        self.setStyleSheet("""
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a2e, stop:0.5 #16213e, stop:1 #0f3460);
            }
        """)
        self._config = get_config()
        self._setup_ui()

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(20)
        layout.setContentsMargins(40, 30, 40, 30)

        # 标题
        title_box = QVBoxLayout()
        title = QLabel("🔍 智能条款比对工具")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #ffffff; font-size: 32px; font-weight: bold;")
        subtitle = QLabel("v15.0 External Config · 支持自定义映射")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: rgba(255,255,255,0.6); font-size: 14px;")
        title_box.addWidget(title)
        title_box.addWidget(subtitle)
        layout.addLayout(title_box)

        # 配置统计
        stats = self._config.get_stats()
        stats_label = QLabel(f"📊 已加载: {stats['client_mappings']} 个映射 | "
                            f"{stats['semantic_aliases']} 个别名 | "
                            f"{stats['keyword_rules']} 个关键词规则")
        stats_label.setAlignment(Qt.AlignCenter)
        stats_label.setStyleSheet("color: rgba(255,255,255,0.5); font-size: 12px;")
        layout.addWidget(stats_label)

        # 输入卡片
        card = GlassCard()
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(35, 35, 35, 35)

        input_style = """
            QLabel { color: #ffffff; font-weight: 500; }
            QLineEdit {
                background: rgba(0,0,0,0.2);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 10px;
                padding: 12px 15px;
                color: #ffffff;
                font-size: 14px;
            }
            QLineEdit:focus { border-color: #667eea; }
        """
        card.setStyleSheet(card.styleSheet() + input_style)

        btn_style = """
            QPushButton {
                background: rgba(255,255,255,0.1);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 10px;
                padding: 12px 20px;
                color: #ffffff;
                font-weight: 500;
            }
            QPushButton:hover {
                background: rgba(255,255,255,0.2);
                border-color: #667eea;
            }
        """

        self.doc_input = self._create_file_row(card_layout, "📂 客户文档",
            "支持中英文 Word 条款清单...", "Word Files (*.docx)", btn_style)
        self.lib_input = self._create_file_row(card_layout, "📚 标准题库",
            "选择 Excel 条款库...", "Excel Files (*.xlsx)", btn_style)

        line = QFrame()
        line.setFixedHeight(1)
        line.setStyleSheet("background: rgba(255,255,255,0.1);")
        card_layout.addWidget(line)

        row3 = QHBoxLayout()
        label3 = QLabel("💾 保存路径")
        label3.setFixedWidth(90)
        self.out_input = QLineEdit()
        self.out_input.setPlaceholderText("设置报告保存位置...")
        btn3 = QPushButton("选择")
        btn3.setCursor(Qt.PointingHandCursor)
        btn3.setStyleSheet(btn_style)
        btn3.clicked.connect(self._browse_save)
        row3.addWidget(label3)
        row3.addWidget(self.out_input, 1)
        row3.addWidget(btn3)
        card_layout.addLayout(row3)

        layout.addWidget(card)

        # 按钮行
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)

        self.start_btn = QPushButton("🚀 开始智能比对")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.setMinimumHeight(55)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #667eea, stop:1 #764ba2);
                color: white; font-size: 17px; font-weight: bold;
                border-radius: 27px; border: none;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #764ba2, stop:1 #667eea);
            }
            QPushButton:disabled { background: rgba(255,255,255,0.1); color: rgba(255,255,255,0.3); }
        """)
        self.start_btn.clicked.connect(self._start_process)

        self.add_mapping_btn = QPushButton("➕ 添加映射")
        self.add_mapping_btn.setCursor(Qt.PointingHandCursor)
        self.add_mapping_btn.setMinimumHeight(55)
        self.add_mapping_btn.setStyleSheet("""
            QPushButton {
                background: transparent; color: rgba(255,255,255,0.7);
                font-size: 15px; font-weight: 500;
                border-radius: 27px; border: 2px solid rgba(255,255,255,0.2);
            }
            QPushButton:hover { border-color: #667eea; color: #667eea; }
        """)
        self.add_mapping_btn.clicked.connect(self._show_add_mapping_dialog)

        self.open_btn = QPushButton("📂 打开目录")
        self.open_btn.setCursor(Qt.PointingHandCursor)
        self.open_btn.setMinimumHeight(55)
        self.open_btn.setEnabled(False)
        self.open_btn.setStyleSheet("""
            QPushButton {
                background: transparent; color: rgba(255,255,255,0.6);
                font-size: 15px; font-weight: 500;
                border-radius: 27px; border: 2px solid rgba(255,255,255,0.2);
            }
            QPushButton:hover { border-color: #27ae60; color: #27ae60; }
            QPushButton:disabled { color: rgba(255,255,255,0.2); border-color: rgba(255,255,255,0.1); }
        """)
        self.open_btn.clicked.connect(self._open_output_folder)

        btn_layout.addWidget(self.start_btn, 3)
        btn_layout.addWidget(self.add_mapping_btn, 1)
        btn_layout.addWidget(self.open_btn, 1)
        layout.addLayout(btn_layout)

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(4)
        self.progress_bar.setStyleSheet("""
            QProgressBar { background: rgba(255,255,255,0.1); border-radius: 2px; }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #667eea, stop:1 #764ba2);
                border-radius: 2px;
            }
        """)
        layout.addWidget(self.progress_bar)

        # 日志
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background: rgba(0,0,0,0.3);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 15px;
                color: #e8e8e8;
                padding: 15px;
                font-family: 'SF Mono', 'Menlo', monospace;
                font-size: 13px;
            }
        """)
        layout.addWidget(self.log_text, 1)

        version_label = QLabel("v15.0 External Config · Made with ❤️")
        version_label.setAlignment(Qt.AlignCenter)
        version_label.setStyleSheet("color: rgba(255,255,255,0.3); font-size: 12px;")
        layout.addWidget(version_label)

    def _create_file_row(self, layout, label_text: str, placeholder: str,
                         filter_str: str, btn_style: str) -> QLineEdit:
        row = QHBoxLayout()
        label = QLabel(label_text)
        label.setFixedWidth(90)
        line_edit = QLineEdit()
        line_edit.setPlaceholderText(placeholder)
        btn = QPushButton("浏览")
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(btn_style)
        btn.clicked.connect(lambda: self._browse_file(line_edit, filter_str))
        row.addWidget(label)
        row.addWidget(line_edit, 1)
        row.addWidget(btn)
        layout.addLayout(row)
        return line_edit

    def _browse_file(self, line_edit: QLineEdit, filter_str: str):
        f, _ = QFileDialog.getOpenFileName(self, "选择文件", "", filter_str)
        if f:
            line_edit.setText(f)
            if line_edit == self.doc_input and not self.out_input.text():
                self.out_input.setText(os.path.join(os.path.dirname(f), "条款比对报告.xlsx"))

    def _browse_save(self):
        f, _ = QFileDialog.getSaveFileName(self, "保存结果", "条款比对报告.xlsx", "Excel Files (*.xlsx)")
        if f:
            self.out_input.setText(f)

    def _show_add_mapping_dialog(self):
        """显示添加映射对话框"""
        dialog = AddMappingDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            eng, chn = dialog.get_mapping()
            if eng and chn:
                self._config.add_client_mapping(eng, chn)
                self._config.save()
                self._append_log(f"✓ 已添加映射: '{eng}' -> '{chn}'", "success")
                QMessageBox.information(self, "成功", f"已添加并保存映射:\n{eng} → {chn}")

    def _append_log(self, msg: str, level: str):
        colors = {"info": "#a0a0a0", "success": "#2ecc71", "error": "#e74c3c", "warning": "#f39c12"}
        self.log_text.append(f'<span style="color:{colors.get(level, "#fff")}">{msg}</span>')
        self.log_text.moveCursor(QTextCursor.End)

    def _start_process(self):
        doc = self.doc_input.text().strip()
        excel = self.lib_input.text().strip()
        out = self.out_input.text().strip()

        if not all([doc, excel, out]):
            QMessageBox.warning(self, "提示", "请先完善所有文件路径！")
            return

        self.start_btn.setEnabled(False)
        self.open_btn.setEnabled(False)
        self.start_btn.setText("⏳ 正在计算中...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()

        self.worker = MatchWorker(doc, excel, out)
        self.worker.log_signal.connect(self._append_log)
        self.worker.progress_signal.connect(lambda c, t: self.progress_bar.setValue(int(c/t*100)))
        self.worker.finished_signal.connect(self._on_finished)
        self.worker.start()

    def _on_finished(self, success: bool, msg: str):
        self.start_btn.setEnabled(True)
        self.start_btn.setText("🚀 开始智能比对")
        self.progress_bar.setVisible(False)

        if success:
            self.open_btn.setEnabled(True)
            self.open_btn.setStyleSheet("""
                QPushButton {
                    background: transparent; color: #2ecc71;
                    font-size: 15px; font-weight: 500;
                    border-radius: 27px; border: 2px solid #2ecc71;
                }
                QPushButton:hover { background: #2ecc71; color: white; }
            """)
            QMessageBox.information(self, "完成", f"比对完成！\n文件已保存至:\n{msg}")

    def _open_output_folder(self):
        path = self.out_input.text().strip()
        if path and os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))


def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    app.setFont(QFont("PingFang SC", 13))

    window = ClauseDiffGUI()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()

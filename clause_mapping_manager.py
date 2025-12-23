# -*- coding: utf-8 -*-
"""
条款映射管理器 - 用户自定义条款映射
支持：单条添加、批量导入、导出、覆盖更新

数据结构：
{
    "英文条款名": {
        "chinese": "中文条款名",
        "library": "库内条款全名",
        "created": "创建时间",
        "updated": "更新时间",
        "source": "来源（manual/import）"
    }
}

Author: Dachi Yijin
Date: 2025-12-23
"""

import json
import os
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime
import logging
import re

logger = logging.getLogger(__name__)


@dataclass
class ClauseMapping:
    """单条条款映射"""
    english: str           # 英文条款名（客户条款）
    chinese: str           # 中文条款名（翻译）
    library: str           # 库内条款名（标准条款库中的条款）
    created: str = ""      # 创建时间
    updated: str = ""      # 更新时间
    source: str = "manual" # 来源：manual（手动）/ import（导入）
    notes: str = ""        # 备注

    def __post_init__(self):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not self.created:
            self.created = now
        if not self.updated:
            self.updated = now


class ClauseMappingManager:
    """
    条款映射管理器

    功能：
    1. 单条添加/编辑/删除映射
    2. 批量导入（JSON/Excel）
    3. 导出映射
    4. 自动覆盖旧映射
    5. 与主程序配置整合
    """

    _instance: Optional['ClauseMappingManager'] = None

    def __init__(self):
        self._mappings: Dict[str, ClauseMapping] = {}
        self._config_path: Optional[Path] = None
        self._is_modified: bool = False

    @classmethod
    def get_instance(cls) -> 'ClauseMappingManager':
        """获取单例实例"""
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    @classmethod
    def reset_instance(cls):
        """重置单例"""
        cls._instance = None

    def _get_default_path(self) -> Path:
        """获取默认映射文件路径"""
        import sys

        # 用户目录优先
        user_dir = Path.home() / ".clause_diff"
        user_dir.mkdir(parents=True, exist_ok=True)
        return user_dir / "user_mappings.json"

    def load(self, path: Optional[str] = None) -> bool:
        """
        加载映射文件

        Args:
            path: 文件路径，None 则使用默认路径

        Returns:
            是否成功加载
        """
        self._config_path = Path(path) if path else self._get_default_path()

        if not self._config_path.exists():
            logger.info(f"映射文件不存在，将创建新文件: {self._config_path}")
            return False

        try:
            with open(self._config_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # 解析映射数据
            mappings_data = data.get('mappings', {})
            for key, value in mappings_data.items():
                if isinstance(value, dict):
                    self._mappings[key.lower()] = ClauseMapping(
                        english=value.get('english', key),
                        chinese=value.get('chinese', ''),
                        library=value.get('library', ''),
                        created=value.get('created', ''),
                        updated=value.get('updated', ''),
                        source=value.get('source', 'import'),
                        notes=value.get('notes', ''),
                    )

            logger.info(f"已加载 {len(self._mappings)} 条映射: {self._config_path}")
            self._is_modified = False
            return True

        except Exception as e:
            logger.error(f"加载映射文件失败: {e}")
            return False

    def save(self, path: Optional[str] = None) -> bool:
        """
        保存映射到文件

        Args:
            path: 保存路径，None 则使用当前路径

        Returns:
            是否成功保存
        """
        save_path = Path(path) if path else self._config_path
        if not save_path:
            save_path = self._get_default_path()

        # 确保目录存在
        save_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            data = {
                "_meta": {
                    "version": "1.0",
                    "description": "用户自定义条款映射",
                    "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "count": len(self._mappings),
                },
                "mappings": {
                    key: asdict(mapping)
                    for key, mapping in self._mappings.items()
                }
            }

            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self._config_path = save_path
            self._is_modified = False
            logger.info(f"已保存 {len(self._mappings)} 条映射: {save_path}")
            return True

        except Exception as e:
            logger.error(f"保存映射文件失败: {e}")
            return False

    # ========================================
    # 单条操作
    # ========================================

    def add_mapping(self, english: str, chinese: str, library: str,
                    notes: str = "", source: str = "manual") -> bool:
        """
        添加或更新单条映射

        Args:
            english: 英文条款名（客户条款）
            chinese: 中文条款名
            library: 库内条款名
            notes: 备注
            source: 来源

        Returns:
            是否成功（True=新增，False=更新）
        """
        key = english.lower().strip()
        is_new = key not in self._mappings

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if is_new:
            self._mappings[key] = ClauseMapping(
                english=english.strip(),
                chinese=chinese.strip(),
                library=library.strip(),
                created=now,
                updated=now,
                source=source,
                notes=notes,
            )
        else:
            # 更新现有映射
            mapping = self._mappings[key]
            mapping.chinese = chinese.strip()
            mapping.library = library.strip()
            mapping.updated = now
            mapping.notes = notes
            if source != "manual":
                mapping.source = source

        self._is_modified = True
        logger.info(f"{'新增' if is_new else '更新'}映射: {english} → {chinese} → {library}")
        return is_new

    def delete_mapping(self, english: str) -> bool:
        """
        删除映射

        Args:
            english: 英文条款名

        Returns:
            是否成功删除
        """
        key = english.lower().strip()
        if key in self._mappings:
            del self._mappings[key]
            self._is_modified = True
            logger.info(f"已删除映射: {english}")
            return True
        return False

    def get_mapping(self, english: str) -> Optional[ClauseMapping]:
        """
        获取映射

        Args:
            english: 英文条款名

        Returns:
            映射对象，不存在则返回 None
        """
        key = english.lower().strip()
        return self._mappings.get(key)

    def get_library_name(self, english: str) -> Optional[str]:
        """
        获取库内条款名（用于导出）

        Args:
            english: 英文条款名

        Returns:
            库内条款名，不存在则返回 None
        """
        mapping = self.get_mapping(english)
        return mapping.library if mapping else None

    def get_chinese_name(self, english: str) -> Optional[str]:
        """获取中文条款名"""
        mapping = self.get_mapping(english)
        return mapping.chinese if mapping else None

    def get_all_mappings(self) -> List[ClauseMapping]:
        """获取所有映射列表"""
        return list(self._mappings.values())

    def get_mapping_count(self) -> int:
        """获取映射数量"""
        return len(self._mappings)

    def is_modified(self) -> bool:
        """是否有未保存的修改"""
        return self._is_modified

    # ========================================
    # 批量导入
    # ========================================

    def import_from_json(self, json_path: str, overwrite: bool = True) -> Tuple[int, int, List[str]]:
        """
        从 JSON 文件批量导入

        Args:
            json_path: JSON 文件路径
            overwrite: 是否覆盖已有映射

        Returns:
            (新增数, 更新数, 错误列表)
        """
        added, updated, errors = 0, 0, []

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # 支持多种格式
            mappings = []

            # 格式1: {"mappings": {...}}
            if 'mappings' in data:
                for key, value in data['mappings'].items():
                    if isinstance(value, dict):
                        mappings.append({
                            'english': value.get('english', key),
                            'chinese': value.get('chinese', ''),
                            'library': value.get('library', ''),
                            'notes': value.get('notes', ''),
                        })

            # 格式2: {"english": "chinese", ...} 简单映射
            elif all(isinstance(v, str) for v in data.values() if not isinstance(v, dict)):
                for eng, chn in data.items():
                    if not eng.startswith('_'):
                        mappings.append({
                            'english': eng,
                            'chinese': chn,
                            'library': '',
                        })

            # 格式3: [{"english": ..., "chinese": ..., "library": ...}, ...]
            elif isinstance(data, list):
                mappings = data

            # 导入映射
            for item in mappings:
                if not item.get('english'):
                    continue

                key = item['english'].lower()
                exists = key in self._mappings

                if exists and not overwrite:
                    continue

                is_new = self.add_mapping(
                    english=item['english'],
                    chinese=item.get('chinese', ''),
                    library=item.get('library', ''),
                    notes=item.get('notes', ''),
                    source='import',
                )

                if is_new:
                    added += 1
                else:
                    updated += 1

            logger.info(f"JSON导入完成: 新增 {added}, 更新 {updated}")

        except json.JSONDecodeError as e:
            errors.append(f"JSON 格式错误: {e}")
        except Exception as e:
            errors.append(f"导入失败: {e}")

        return added, updated, errors

    def import_from_excel(self, excel_path: str, overwrite: bool = True,
                          col_english: int = 0, col_chinese: int = 1,
                          col_library: int = 2, header_row: int = 0) -> Tuple[int, int, List[str]]:
        """
        从 Excel 文件批量导入

        Args:
            excel_path: Excel 文件路径
            overwrite: 是否覆盖已有映射
            col_english: 英文列索引（从0开始）
            col_chinese: 中文列索引
            col_library: 库内条款列索引
            header_row: 表头行数（跳过）

        Returns:
            (新增数, 更新数, 错误列表)
        """
        added, updated, errors = 0, 0, []

        try:
            import openpyxl

            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx < header_row:
                    continue

                # 获取各列值
                english = str(row[col_english] or '').strip() if len(row) > col_english else ''
                chinese = str(row[col_chinese] or '').strip() if len(row) > col_chinese else ''
                library = str(row[col_library] or '').strip() if len(row) > col_library else ''

                if not english:
                    continue

                key = english.lower()
                exists = key in self._mappings

                if exists and not overwrite:
                    continue

                is_new = self.add_mapping(
                    english=english,
                    chinese=chinese,
                    library=library,
                    source='import',
                )

                if is_new:
                    added += 1
                else:
                    updated += 1

            wb.close()
            logger.info(f"Excel导入完成: 新增 {added}, 更新 {updated}")

        except ImportError:
            errors.append("需要安装 openpyxl: pip install openpyxl")
        except Exception as e:
            errors.append(f"导入失败: {e}")

        return added, updated, errors

    # ========================================
    # 导出功能
    # ========================================

    def export_to_json(self, output_path: str) -> bool:
        """
        导出映射到 JSON 文件

        Args:
            output_path: 输出路径

        Returns:
            是否成功
        """
        try:
            data = {
                "_meta": {
                    "exported": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "count": len(self._mappings),
                },
                "mappings": {
                    m.english: {
                        "chinese": m.chinese,
                        "library": m.library,
                        "notes": m.notes,
                    }
                    for m in self._mappings.values()
                }
            }

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            logger.info(f"已导出 {len(self._mappings)} 条映射到: {output_path}")
            return True

        except Exception as e:
            logger.error(f"导出失败: {e}")
            return False

    def export_to_excel(self, output_path: str) -> bool:
        """
        导出映射到 Excel 文件

        Args:
            output_path: 输出路径

        Returns:
            是否成功
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "条款映射"

            # 表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # 写入表头
            headers = ["英文条款名", "中文条款名", "库内条款名", "创建时间", "更新时间", "来源", "备注"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # 写入数据
            for row, mapping in enumerate(self._mappings.values(), 2):
                ws.cell(row=row, column=1, value=mapping.english)
                ws.cell(row=row, column=2, value=mapping.chinese)
                ws.cell(row=row, column=3, value=mapping.library)
                ws.cell(row=row, column=4, value=mapping.created)
                ws.cell(row=row, column=5, value=mapping.updated)
                ws.cell(row=row, column=6, value=mapping.source)
                ws.cell(row=row, column=7, value=mapping.notes)

            # 调整列宽
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 45
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 20

            wb.save(output_path)
            logger.info(f"已导出 {len(self._mappings)} 条映射到: {output_path}")
            return True

        except Exception as e:
            logger.error(f"导出失败: {e}")
            return False

    # ========================================
    # 与主程序整合
    # ========================================

    def apply_to_config(self, config_manager) -> int:
        """
        将用户映射应用到主配置管理器

        Args:
            config_manager: ClauseConfigManager 实例

        Returns:
            应用的映射数量
        """
        count = 0
        for mapping in self._mappings.values():
            # 添加英中映射
            if mapping.chinese:
                key = mapping.english.lower()
                config_manager.client_en_cn_map[key] = mapping.chinese
                count += 1

            # 添加精确映射（中文 → 库内条款）
            if mapping.library and mapping.chinese:
                config_manager.exact_clause_map[mapping.chinese] = mapping.library

        logger.info(f"已应用 {count} 条用户映射到配置")
        return count

    def get_export_name(self, original_name: str) -> str:
        """
        获取导出时应使用的名称

        如果有用户映射，返回库内条款名；否则返回原名

        Args:
            original_name: 原始条款名（英文或中文）

        Returns:
            应使用的导出名称
        """
        # 尝试按英文查找
        mapping = self.get_mapping(original_name)
        if mapping and mapping.library:
            return mapping.library

        # 尝试按中文查找
        for m in self._mappings.values():
            if m.chinese == original_name and m.library:
                return m.library

        return original_name


# ========================================
# 便捷函数
# ========================================

def get_mapping_manager() -> ClauseMappingManager:
    """获取映射管理器单例"""
    return ClauseMappingManager.get_instance()

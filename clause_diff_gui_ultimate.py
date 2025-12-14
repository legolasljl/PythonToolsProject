# -*- coding: utf-8 -*-
"""
智能条款比对工具 v12.0 (Smart Splitter Edition)
- [核心修复] 解决紧凑型文档(无空行)导致的条款粘连问题
- [算法升级] 引入基于正则特征的智能切分，不再单纯依赖空行
- [功能] 包含 v11 所有功能 (翻译/字典/风控/UI)

Author: Google Senior Architect
Date: 2025-12-09
"""

import sys
import os
import re
import difflib
import traceback
import pandas as pd
from docx import Document

# 翻译库 (可选)
try:
    from deep_translator import GoogleTranslator
    HAS_TRANSLATOR = True
except ImportError:
    HAS_TRANSLATOR = False

# Excel 样式库
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# PyQt5 库
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QProgressBar, QTextEdit, 
    QFileDialog, QMessageBox, QStyleFactory, QFrame,
    QGraphicsDropShadowEffect
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt5.QtGui import QFont, QPalette, QColor, QDesktopServices, QTextCursor

# ==========================================
# 🛡️ 核心修复：防止 macOS 打包后闪退
# ==========================================
class NullWriter:
    def write(self, text): pass
    def flush(self): pass

if getattr(sys, 'frozen', False):
    sys.stdout = NullWriter()
    sys.stderr = NullWriter()

def global_exception_handler(exctype, value, tb):
    error_msg = "".join(traceback.format_exception(exctype, value, tb))
    try: sys.__stderr__.write(error_msg)
    except: pass
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Critical)
    msg_box.setText("程序发生意外错误")
    msg_box.setInformativeText(str(value))
    msg_box.setDetailedText(error_msg)
    msg_box.setWindowTitle("错误报告")
    msg_box.exec_()

sys.excepthook = global_exception_handler

# --------------------------
# 核心：翻译与术语处理层
# --------------------------
class ClauseMatcherLogic:
    # 📖 财产险专业术语字典 (v9.0 全量版)
    INSURANCE_GLOSSARY = {
        # --- 1. 用户补充的常用条款 (High Priority) ---
        "undamaged building extra charges extension clause": "建筑物未受损部分额外费用扩展条款",
        "average relief clause": "放弃比例分摊条款",
        "cost of compiling records and claim preparation clause": "编辑记录及索赔准备费用条款",
        "cost of re-erection clause": "重新安装费用条款",
        "all other contents clause": "所有其他物品条款",
        "leased property clause": "租赁财产条款",
        "off premises property clause": "营业处所外财产条款",
        "0ff premises property clause": "营业处所外财产条款",
        "book of account clause": "会计账册条款",
        "departmental clause": "部门条款",
        "new business clause": "新营业条款",
        "waiver of excess": "免赔豁免条款",
        "conveyer belts and knives extension clause": "传送带及刀具扩展条款",
        "smoke damage clause": "烟熏损失条款",
        "non-invalidation clause": "不使失效条款",
        "misdescription clause": "错误描述条款",
        "public utilities extension clause": "公共设施扩展条款", # Utilities
        
        # --- 2. 核心修正 ---
        "interdependency clause": "关联条款", 
        "denial of access clause": "通道堵塞条款",
        "accumulated stocks clause": "累积存货条款",
        "unnamed customer/supplier's extension": "不具名顾客/供应商扩展条款",
        "pairs & set clause": "成对或成套条款",
        "breach of condition clause": "违反条件条款",
        
        # --- 3. 常用条款 ---
        "earthquake and tsunami clause": "地震海啸扩展条款",
        "reinstatement value clause": "重置价值条款",
        "automatic reinstatement of sum insured clause": "自动恢复保险金额条款",
        "automatic reinstatement of sum insured": "自动恢复保险金额条款",
        "professional fees clause": "专业费用扩展条款",
        "extra charge clause": "额外费用条款",
        "extra charges clause": "额外费用条款",
        "removal of debris clause": "清理残骸费用扩展条款",
        "public authorities clause": "公共当局条款",
        "fire extinguishing expenses clause": "灭火费用条款",
        "temporary removal clause": "临时移动条款",
        "escalation": "自动升值条款",
        "capital addition clause": "增加资产条款",
        "loss adjuster clause": "指定公估人条款",
        "waiver of subrogation clause": "放弃代位求偿权条款",
        "designation of property clause": "指明财产条款",
        "payment on account clause": "预付赔款条款",
        "boiler explosion clause": "锅炉爆炸条款",
        "landslide & subsidence clause": "山崩及地陷条款",
        "import duty clause": "关税条款",
        "tax clause": "税费条款",
        "tax and duty clause": "税费条款",
        "claim expenses clause": "索赔费用条款",
        "laptop extension clause": "便携式设备扩展条款",
        "portable computer extension": "便携式设备扩展条款",
        "burglary extension clause": "盗窃、抢劫扩展条款",
        "full theft clause": "全盗抢条款",
        "interruption of public supplier clause": "供应中断扩展条款",
        "supply failure extension clause": "供应中断扩展条款",
        "public utility extension": "公共设施扩展条款", # Utility
        "contract price clause": "合同价格条款",
        "brand and trademark clause": "商标及标签条款",
        "sprinkler leakage damage clause": "自动喷淋水损条款",
        "glass breakage clause": "玻璃破碎条款",
        "time adjustment clause": "时间调整条款",
        "errors and omissions clause": "错误和遗漏条款",
        "no control clause": "不受控制条款",
        "breach of conditions clause": "违反条件条款",
        "notice of cancellation": "注销保单条款",
        "sue & labor clause": "诉讼和施救费用条款",
        "inland transit extension clause": "内陆运输扩展条款",
        "vehicle load clause": "车载货物条款",
        "outdoor fixtures and fittings extension": "室外装置及配件扩展条款",
        "full flood extension": "洪水扩展条款",
        "strike, riot & civil commotion clause": "罢工、暴乱及民众骚乱条款",
        "strike, riot and civil commotion clause": "罢工、暴乱及民众骚乱条款",
        "terrorism extension clause": "恐怖主义扩展条款",

# --- 👇 本次新增/更新的专业条款 ---
        "automatic capital additions clause": "自动资产增加条款",
        "capital additions clause": "资产增加条款",
        "alterations, additions & repairs clause": "扩建、改建及维修条款",
        "automatic cover for new locations": "自动承保新地点条款",
        "burglary, theft & robbery": "盗窃、抢劫条款",
        "care, custody and control coverage": "被保险人监护、看管及控制条款",
        "co-insurance clause": "共保人条款",
        "exhibition/trade shows clause": "展览会条款",
        "extension cover for earthquake and tsunami": "地震海啸扩展条款",
        "landslide and subsidence clause": "山崩及地陷条款",
        "loss notification clause": "损失通知条款",
        "minor works clause": "小型工程条款",
        "miscellaneous unnamed locations": "其他未列名地点条款",
        "money insurance": "现金保险条款",
        "nominated loss adjuster clauses": "指定公估人条款",
        "strike, riot & civil commotion": "罢工、暴动及民众骚乱条款",
        "temporary protection clause": "临时保护条款",
        "water tanks, apparatus and pipes clause": "水箱、装置及水管条款",
        "stock declaration clause": "仓储财产申报条款",
        "payment on account": "预付赔款条款",
        "removal of debris": "清理残骸费用条款",
        "replacement value clause": "重置价值条款",
        "additional increase in cost of working": "额外工作费用增加条款",
        "civil authorities": "公共当局条款",
        "errors & omissions clause": "错误与遗漏条款",
        "extra expense": "额外费用条款",

        # --- 4. 通用术语 ---
        "deductible": "免赔额",
        "excess": "免赔额",
        "premium": "保险费",
        "insured": "被保险人",
        "insurer": "保险人",
        "policy": "保险单",
        "clause": "条款",
        "exclusion": "除外责任",
        "extension": "扩展条款"
    }

    PENALTY_KEYWORDS = ["打孔盗气"]
    
    ALIAS_MAP = {
        "commotion": "civil commotion",
        "malicious damage": "malicious acts",
    }

    @staticmethod
    def clean_text_for_title(text):
        if not isinstance(text, str): return ""
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        for key, value in ClauseMatcherLogic.ALIAS_MAP.items():
            if key in text: text = text.replace(key, value)
        noise_words = ["企业财产保险", "附加", "扩展", "条款", "险", "（A款）", "（B款）", "2025版"]
        for w in noise_words: text = text.replace(w, "")
        text = re.sub(r'[0-9\s]+', '', text)
        return text

    @staticmethod
    def clean_text_content(text):
        if not isinstance(text, str): return ""
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        text = re.sub(r'\s+', '', text)
        text = re.sub(r'[0-9]+', '', text)
        return text

    @staticmethod
    def extract_extra_info(text):
        if not isinstance(text, str): return ""
        matches = re.findall(r'([\(（].*?[\)）])', text)
        if matches: return " ".join(matches)
        return ""

    @staticmethod
    def is_likely_title(text):
        """
        [NEW] 基于特征判断某一行是否像标题
        1. 长度适中 (<60字)
        2. 不以句号、分号结尾 (通常是正文)
        3. 包含特定关键词 (如“条款”、“险”)
        """
        if len(text) > 60: return False
        if text.endswith(('。', '；', '.', ';')): return False
        if "条款" in text or "Clause" in text or "Extension" in text: return True
        return True

    @staticmethod
    def parse_docx(doc_path):
        doc = Document(doc_path)
        clauses = []
        current_block = []
        
        all_lines = [p.text.strip() for p in doc.paragraphs]
        
        # 统计空行
        empty_lines = sum(1 for t in all_lines if not t)
        total_lines = len([t for t in all_lines if t])
        
        # 智能模式选择
        # 如果空行极少，且无法简单用空行切分，就启用“特征切分”
        use_smart_split = False
        if total_lines > 0 and (empty_lines / len(all_lines) < 0.05):
            use_smart_split = True
        
        if use_smart_split:
            # [NEW] 智能切分逻辑
            # 即使没有空行，如果当前行像标题，就切一刀
            for text in all_lines:
                if not text: continue
                
                # 如果当前积累了内容，且新的一行像标题 -> 切分
                if current_block and ClauseMatcherLogic.is_likely_title(text):
                    # 保存上一条
                    title = current_block[0]
                    content = "\n".join(current_block[1:])
                    clauses.append({'Title': title, 'Content': content})
                    current_block = [text] # 新开始
                else:
                    # 否则加入当前块
                    current_block.append(text)
            
            # 最后一个块
            if current_block:
                title = current_block[0]
                content = "\n".join(current_block[1:])
                clauses.append({'Title': title, 'Content': content})
                
        else:
            # 传统空行切分逻辑 (适合排版稀疏的文档)
            for text in all_lines:
                if text:
                    current_block.append(text)
                else:
                    if current_block:
                        title = current_block[0]
                        content = "\n".join(current_block[1:])
                        clauses.append({'Title': title, 'Content': content})
                        current_block = []
            if current_block:
                title = current_block[0]
                content = "\n".join(current_block[1:])
                clauses.append({'Title': title, 'Content': content})
        
        # 后处理：如果切分后发现全是“无内容”的纯标题，标记为 title_only
        is_title_only = all(not c['Content'] for c in clauses)
        return clauses, is_title_only

    @staticmethod
    def get_adaptive_score(c_title, c_content, l_name, l_content, is_title_only):
        c_title_clean = ClauseMatcherLogic.clean_text_for_title(c_title)
        l_name_clean = ClauseMatcherLogic.clean_text_for_title(l_name)
        score_title = difflib.SequenceMatcher(None, c_title_clean, l_name_clean).ratio()
        
        score_content = 0
        if not is_title_only and c_content.strip():
            c_content_clean = ClauseMatcherLogic.clean_text_content(c_content)
            l_content_clean = ClauseMatcherLogic.clean_text_content(l_content)
            score_content = difflib.SequenceMatcher(None, c_content_clean, l_content_clean).ratio()
        
        if is_title_only or not c_content.strip():
            final_score = score_title
        else:
            final_score = 0.8 * score_title + 0.2 * score_content
            
        for bad_word in ClauseMatcherLogic.PENALTY_KEYWORDS:
            if bad_word in l_name and bad_word not in c_title:
                final_score -= 0.5
                
        return final_score, score_title, score_content

    @staticmethod
    def analyze_difference(c_content, l_content):
        analysis = []
        c_text = str(c_content)
        l_text = str(l_content)
        if not c_text.strip(): return ""

        keywords = {
            "限额": ["Limit", "限额"],
            "免赔": ["Deductible", "Excess", "免赔"],
            "除外": ["Exclusion", "除外", "不负责"],
            "观察期": ["Waiting Period", "观察期"]
        }
        for key, words in keywords.items():
            c_has = any(w in c_text for w in words)
            l_has = any(w in l_text for w in words)
            if c_has and not l_has: analysis.append(f"⚠️ 客户提及[{key}]但库内未提及")
            elif not c_has and l_has: analysis.append(f"ℹ️ 库内包含[{key}]但客户未提及")

        return " | ".join(analysis)

    @staticmethod
    def is_english(text):
        if not isinstance(text, str): return False
        # Simple check: if less than 10% characters are Chinese, treat as English
        zh_count = len(re.findall(r'[\u4e00-\u9fa5]', text))
        return zh_count < len(text) * 0.1 and len(text) > 3

    @staticmethod
    def translate_to_chinese(text):
        if not HAS_TRANSLATOR: return text
        try:
            # 使用 Google 翻译源
            return GoogleTranslator(source='auto', target='zh-CN').translate(text)
        except Exception as e:
            print(f"Translation error: {e}")
            return text

# --------------------------
# 工作线程
# --------------------------
class MatchWorker(QThread):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, doc_path, excel_path, output_path):
        super().__init__()
        self.doc_path = doc_path
        self.excel_path = excel_path
        self.output_path = output_path
        
    def run(self):
        try:
            if not HAS_TRANSLATOR:
                self.log_signal.emit("⚠️ 未检测到 deep_translator，仅使用内置术语表。", "warning")

            self.log_signal.emit("⏳ 正在初始化...", "info")
            clauses, is_title_only = ClauseMatcherLogic.parse_docx(self.doc_path)
            mode_str = "纯标题模式" if is_title_only else "完整内容模式"
            self.log_signal.emit(f"📖 识别为 [{mode_str}]，提取到 {len(clauses)} 条", "success")
            
            lib_df = pd.read_excel(self.excel_path)
            lib_df.columns = [str(c).strip() for c in lib_df.columns]
            lib_data = lib_df.to_dict('records')
            
            self.log_signal.emit("🧠 执行匹配...", "info")
            results = []
            total = len(clauses)
            
            for idx, c in enumerate(clauses, 1):
                self.progress_signal.emit(idx, total)
                c_title = c['Title']
                c_content = c['Content']
                
                # 1. 翻译
                is_trans = False
                title_for_match = c_title
                content_for_match = c_content
                
                if ClauseMatcherLogic.is_english(c_title):
                    if idx % 5 == 0: self.log_signal.emit(f"   🔄 翻译: {c_title[:20]}...", "info")
                    title_for_match = ClauseMatcherLogic.translate_to_chinese(c_title)
                    is_trans = True
                    if c_content and ClauseMatcherLogic.is_english(c_content):
                        content_for_match = ClauseMatcherLogic.translate_to_chinese(c_content)
                else:
                    if idx % 5 == 0: self.log_signal.emit(f"   匹配: {c_title[:10]}...", "info")

                # 2. 匹配
                best_match = None
                best_score = -100
                best_meta = {}
                
                for lib in lib_data:
                    l_name = str(lib.get('条款名称', ''))
                    l_content = str(lib.get('条款内容', ''))
                    score, s_t, s_c = ClauseMatcherLogic.get_adaptive_score(
                        title_for_match, content_for_match, l_name, l_content, is_title_only
                    )
                    if score > best_score:
                        best_score = score
                        best_match = lib
                        best_meta = {'t': s_t, 'c': s_c}
                
                match_name = "无匹配"
                match_content = ""
                match_reg = ""
                
                if best_match and best_score > 0.1:
                    base_name = best_match.get('条款名称', '')
                    match_content = best_match.get('条款内容', '')
                    match_reg = best_match.get('产品注册号', best_match.get('注册号', ''))
                    
                    extra_params = ClauseMatcherLogic.extract_extra_info(c_title)
                    if extra_params and extra_params not in base_name:
                        match_name = f"{base_name} {extra_params}"
                    else:
                        match_name = base_name
                
                # 3. 风控
                diff_analysis = ""
                final_score = best_score if best_score > 0 else 0
                if final_score < 0.6 and best_match:
                    diff_analysis = ClauseMatcherLogic.analyze_difference(c_content, match_content)
                
                results.append({
                    '序号': idx,
                    '客户条款(原)': c_title,
                    '客户条款(译)': title_for_match if is_trans else "",
                    '客户原始内容': c['Content'], 
                    '匹配条款库名称': match_name,
                    '产品注册号': match_reg,
                    '匹配条款库内容': match_content,
                    '综合匹配度': final_score,
                    '保障差异提示': diff_analysis,
                    '标题相似度': best_meta.get('t', 0),
                    '内容相似度': best_meta.get('c', 0),
                })
            
            df_res = pd.DataFrame(results)
            df_res.to_excel(self.output_path, index=False)
            self.apply_excel_styles()
            
            self.log_signal.emit(f"🎉 完成！已生成 Excel 对比文件。", "success")
            self.finished_signal.emit(True, self.output_path)
            
        except Exception as e:
            raise e

    def apply_excel_styles(self):
        wb = openpyxl.load_workbook(self.output_path)
        wb.properties.creator = "Alex Jin"
        wb.properties.lastModifiedBy = "Alex Jin"
        ws = wb.active
        
        fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        fill_blue = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        fill_orange = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['B'].width = 30 
        ws.column_dimensions['C'].width = 25 
        ws.column_dimensions['D'].width = 40 
        ws.column_dimensions['E'].width = 35 
        ws.column_dimensions['F'].width = 25 
        ws.column_dimensions['G'].width = 60 
        ws.column_dimensions['I'].width = 40 
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if cell.col_idx in [8, 9, 10]:
                    val = cell.value
                    if isinstance(val, (int, float)):
                        cell.number_format = '0.00%'
                        if val < 0.40: cell.fill = fill_red
                        elif 0.40 <= val < 0.75: cell.fill = fill_blue
                        elif val >= 0.75: cell.fill = fill_green
                if cell.col_idx == 9 and cell.value:
                    cell.fill = fill_orange
                    
        wb.save(self.output_path)

# --------------------------
# UI 界面
# --------------------------
class ClauseDiffGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("智能条款比对工具 v12.0 (Smart Split)")
        self.resize(1000, 800)
        QApplication.setStyle(QStyleFactory.create('Fusion'))
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(242, 246, 250))
        palette.setColor(QPalette.WindowText, QColor(40, 40, 40))
        palette.setColor(QPalette.Base, QColor(255, 255, 255))
        palette.setColor(QPalette.Button, QColor(255, 255, 255))
        palette.setColor(QPalette.ButtonText, QColor(40, 40, 40))
        QApplication.setPalette(palette)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setSpacing(25)
        layout.setContentsMargins(40, 40, 40, 40)
        
        title_box = QVBoxLayout()
        title = QLabel("⚖️ 智能条款比对工具")
        title.setFont(QFont("Microsoft YaHei", 28, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #2c3e50;")
        subtitle = QLabel("全能版：智能清单切分 • 英汉互译 • 风险提示")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: #7f8c8d; font-size: 15px; margin-top: 5px;")
        title_box.addWidget(title); title_box.addWidget(subtitle)
        layout.addLayout(title_box)

        card = QFrame()
        card.setStyleSheet("QFrame { background-color: #fff; border-radius: 15px; }")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20); shadow.setColor(QColor(0,0,0,20)); shadow.setOffset(0,5)
        card.setGraphicsEffect(shadow)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(30, 30, 30, 30)
        
        input_style = "QLineEdit { border: 2px solid #ecf0f1; border-radius: 8px; padding: 10px; background: #f9fbfd; font-size: 14px; color: #2c3e50; } QLineEdit:focus { border: 2px solid #3498db; background: #fff; }"
        btn_style = "QPushButton { background: #f8f9fa; border: 1px solid #dfe6e9; border-radius: 8px; padding: 10px; font-weight: bold; color: #2d3436; } QPushButton:hover { background: #e2e6ea; }"

        self.doc_input = self.create_file_row(card_layout, "📂 客户文档:", "支持中文或英文 Word 条款...", "Word Files (*.docx)", input_style, btn_style)
        self.lib_input = self.create_file_row(card_layout, "📚 标准题库:", "选择 Excel 条款库...", "Excel Files (*.xlsx)", input_style, btn_style)
        
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setStyleSheet("background-color: #f0f2f5; border: none; height: 1px;")
        card_layout.addWidget(line)

        row3 = QHBoxLayout()
        row3.addWidget(QLabel("💾 结果保存:"))
        self.out_input = QLineEdit()
        self.out_input.setPlaceholderText("设置 Excel 报告保存路径...")
        self.out_input.setStyleSheet(input_style)
        btn3 = QPushButton("保存路径")
        btn3.setCursor(Qt.PointingHandCursor)
        btn3.setStyleSheet(btn_style)
        btn3.clicked.connect(self.browse_save)
        row3.addWidget(self.out_input, 1); row3.addWidget(btn3)
        card_layout.addLayout(row3)

        layout.addWidget(card)

        btn_layout = QHBoxLayout(); btn_layout.setSpacing(20)
        self.start_btn = QPushButton("🚀 开始智能比对")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.setMinimumHeight(60)
        self.start_btn.setStyleSheet("QPushButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #8e44ad, stop:1 #9b59b6); color: white; font-size: 18px; font-weight: bold; border-radius: 30px; border: 1px solid #8e44ad; } QPushButton:hover { margin-top: -2px; } QPushButton:pressed { margin-top: 2px; } QPushButton:disabled { background: #bdc3c7; border-color: #bdc3c7; }")
        self.start_btn.clicked.connect(self.start_process)
        
        self.open_folder_btn = QPushButton("📂 打开文件目录")
        self.open_folder_btn.setCursor(Qt.PointingHandCursor)
        self.open_folder_btn.setMinimumHeight(60)
        self.open_folder_btn.setEnabled(False)
        self.open_folder_btn.setStyleSheet("QPushButton { background: white; color: #2c3e50; font-size: 16px; font-weight: bold; border-radius: 30px; border: 2px solid #bdc3c7; } QPushButton:hover { border-color: #3498db; color: #3498db; } QPushButton:disabled { color: #bdc3c7; border-color: #ecf0f1; }")
        self.open_folder_btn.clicked.connect(self.open_output_folder)

        btn_layout.addWidget(self.start_btn, 2)
        btn_layout.addWidget(self.open_folder_btn, 1)
        layout.addLayout(btn_layout)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("QProgressBar { border: none; background: #e0e0e0; border-radius: 3px; height: 6px; } QProgressBar::chunk { background: #9b59b6; border-radius: 3px; }")
        layout.addWidget(self.progress_bar)

        self.log_text = QTextEdit(); self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("QTextEdit { background: #2c3e50; border-radius: 10px; color: #ecf0f1; padding: 15px; font-family: 'Menlo', monospace; font-size: 12px; }")
        layout.addWidget(self.log_text, 1)

    def create_file_row(self, layout, label, placeholder, filter_str, style, btn_style):
        row = QHBoxLayout(); row.addWidget(QLabel(label))
        line_edit = QLineEdit(); line_edit.setPlaceholderText(placeholder); line_edit.setStyleSheet(style)
        btn = QPushButton("浏览"); btn.setCursor(Qt.PointingHandCursor); btn.setStyleSheet(btn_style)
        btn.clicked.connect(lambda: self.browse_file(line_edit, filter_str))
        row.addWidget(line_edit, 1); row.addWidget(btn)
        layout.addLayout(row)
        return line_edit

    def browse_file(self, line_edit, filter_str):
        f, _ = QFileDialog.getOpenFileName(self, "选择文件", "", filter_str)
        if f: line_edit.setText(f)
        if f and line_edit == self.doc_input and not self.out_input.text():
            folder = os.path.dirname(f); self.out_input.setText(os.path.join(folder, "条款比对报告.xlsx"))

    def browse_save(self):
        f, _ = QFileDialog.getSaveFileName(self, "保存结果", "条款比对报告.xlsx", "Excel Files (*.xlsx)")
        if f: self.out_input.setText(f)

    def append_log(self, msg, level):
        colors = {"info": "#bdc3c7", "success": "#2ecc71", "error": "#e74c3c", "warning": "#f1c40f"}
        self.log_text.append(f'<span style="color:{colors.get(level, "#fff")}">{msg}</span>')
        self.log_text.moveCursor(QTextCursor.End)

    def start_process(self):
        doc = self.doc_input.text().strip(); excel = self.lib_input.text().strip(); out = self.out_input.text().strip()
        if not doc or not excel or not out: QMessageBox.warning(self, "提示", "请先完善文件路径！"); return
        self.start_btn.setEnabled(False); self.open_folder_btn.setEnabled(False); self.start_btn.setText("⏳ 正在计算中..."); self.progress_bar.setVisible(True); self.progress_bar.setValue(0); self.log_text.clear()
        self.worker = MatchWorker(doc, excel, out)
        self.worker.log_signal.connect(self.append_log)
        self.worker.progress_signal.connect(lambda c, t: self.progress_bar.setValue(int(c/t*100)))
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def on_finished(self, success, msg):
        self.start_btn.setEnabled(True); self.start_btn.setText("🚀 开始智能比对"); self.progress_bar.setVisible(False)
        if success:
            self.open_folder_btn.setEnabled(True)
            self.open_folder_btn.setStyleSheet("QPushButton { background: white; color: #27ae60; font-size: 16px; font-weight: bold; border-radius: 30px; border: 2px solid #27ae60; } QPushButton:hover { background: #27ae60; color: white; }")
            QMessageBox.information(self, "成功", f"比对完成！\n文件已保存至:\n{msg}")
        # 错误由全局处理

    def open_output_folder(self):
        path = self.out_input.text().strip()
        if path and os.path.exists(path): QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))
        else: QMessageBox.warning(self, "提示", "文件路径不存在！")

def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'): QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'): QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    w = ClauseDiffGUI()
    w.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()

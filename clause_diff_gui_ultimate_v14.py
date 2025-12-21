# -*- coding: utf-8 -*-
"""
智能条款比对工具 v14.0 (Client Mapping Enhanced Edition)
- [核心升级] 基于客户中英文条款建立精确映射字典
- [新增] 重点条款特别关注：INTERPRETATION & HEADINGS, REINSTATEMENT VALUE等
- [优化] 多级匹配策略：精确匹配 > 语义别名 > 关键词 > 模糊匹配
- [代码重构] 分离配置、逻辑、UI三层

Author: Dachi Yijin
Date: 2025-12-18
"""

import sys
import os
import re
import difflib
import traceback
from typing import List, Dict, Tuple, Optional, Set
from dataclasses import dataclass, field
from enum import Enum
import pandas as pd
from docx import Document

# ==========================================
# 🔧 macOS PyQt5 Plugin Fix
# ==========================================
try:
    import PyQt5
    plugin_path = os.path.join(os.path.dirname(PyQt5.__file__), 'Qt5', 'plugins')
    os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path
except ImportError:
    pass

try:
    from deep_translator import GoogleTranslator
    HAS_TRANSLATOR = True
except ImportError:
    HAS_TRANSLATOR = False

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QProgressBar, QTextEdit, 
    QFileDialog, QMessageBox, QFrame, QGraphicsDropShadowEffect
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt5.QtGui import QFont, QColor, QDesktopServices, QTextCursor

# ==========================================
# macOS 打包防闪退
# ==========================================
class NullWriter:
    def write(self, text): pass
    def flush(self): pass

if getattr(sys, 'frozen', False):
    sys.stdout = NullWriter()
    sys.stderr = NullWriter()

def global_exception_handler(exctype, value, tb):
    error_msg = "".join(traceback.format_exception(exctype, value, tb))
    try: sys.__stderr__.write(error_msg)
    except: pass
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Critical)
    msg_box.setText("程序发生意外错误")
    msg_box.setInformativeText(str(value))
    msg_box.setDetailedText(error_msg)
    msg_box.exec_()

sys.excepthook = global_exception_handler


# ==========================================
# 配置层：所有映射字典集中管理
# ==========================================
class ClauseConfig:
    """条款配置 - 集中管理所有映射字典"""
    
    # ========================================
    # 🎯 客户中英文条款精确映射（基于客户实际文档）
    # ========================================
    CLIENT_EN_CN_MAP: Dict[str, str] = {
        # ===== 用户重点关注的条款 =====
        "interpretation & headings": "通译和标题条款",
        "interpretation and headings": "通译和标题条款",
        "reinstatement (value)": "重置价值条款",
        "reinstatement value": "重置价值条款",
        "reinstatement value clause": "重置价值条款",
        "replacement value": "重置价值条款",
        "replacement value clause": "重置价值条款",
        "time adjustment (72 hours)": "72小时条款",
        "time adjustment": "72小时条款",
        "72 hours clause": "72小时条款",
        "civil authorities clause": "公共当局扩展条款",
        "civil authorities": "公共当局扩展条款",
        "public authorities clause": "公共当局扩展条款",
        "public authorities": "公共当局扩展条款",
        "errors and omissions clause": "错误和遗漏条款",
        "errors and omissions": "错误和遗漏条款",
        "loss notification clause": "损失通知条款",
        "loss notification": "损失通知条款",
        "no control": "不受控制条款",
        "no control clause": "不受控制条款",
        "no contorl": "不受控制条款",  # 客户文档拼写
        
        # ===== 财产一切险条款 =====
        "60 days' notice of cancellation by insurer": "60天通知注销保单条款",
        "60 days notice of cancellation": "60天通知注销保单条款",
        "notice of cancellation": "注销保单条款",
        "expediting costs": "加快费用条款",
        "all other contents": "其它物品条款",
        "alterations, additions and repairs": "变更和维修条款",
        "alterations additions and repairs": "变更和维修条款",
        "escalation": "自动升值扩展条款",
        "automatic cover for new asset": "自动扩展承保新增资产、新增公司和新增地址条款",
        "automatic cover for new asset, or newly set up companies and locations": "自动扩展承保新增资产、新增公司和新增地址条款",
        "unnamed location clause": "未列明地址条款",
        "unnamed location": "未列明地址条款",
        "automatic capital additions": "增加资产条款",
        "capital additions": "增加资产条款",
        "stock declaration & adjustment": "仓储财产申报条款",
        "stock declaration and adjustment": "仓储财产申报条款",
        "stock declaration": "仓储财产申报条款",
        "automatic reinstatement of sum insured": "自动恢复保险金额条款",
        "brand or trademark": "品牌和商标标识条款",
        "brand and trademark": "品牌和商标标识条款",
        "professional fees and claim preparation costs": "专业费用及索赔准备费用条款",
        "professional fees": "专业费用及索赔准备费用条款",
        "claims preparation costs": "专业费用及索赔准备费用条款",
        "tax clause": "税金约定条款",
        "earthquake and tsunami clause": "地震扩展条款",
        "earthquake and tsunami": "地震扩展条款",
        "earthquake extension": "地震扩展条款",
        "theft and robbery": "盗窃、抢劫扩展条款",
        "full theft, burglary and robbery cover": "承保全部盗窃条款",
        "full theft": "承保全部盗窃条款",
        "description of property insured": "被保险财产条款",
        "public utility clause": "公用设施故障条款",
        "public utilities": "公用设施故障条款",
        "multiple insureds clause": "共同被保险人条款",
        "multiple insureds": "共同被保险人条款",
        "frozen and refrigerated property clause": "冷冻、冷藏品条款",
        "frozen and refrigerated": "冷冻、冷藏品条款",
        "removal of debris": "清理残骸费用扩展条款",
        "debris removal": "清理残骸费用扩展条款",
        "strike, riot, civil commotion": "罢工、暴动或民众骚乱条款",
        "strike riot civil commotion": "罢工、暴动或民众骚乱条款",
        "srcc": "罢工、暴动或民众骚乱条款",
        "temporary removal": "临时移动扩展条款",
        "vehicle load": "车辆装载物扩展条款",
        "80% co-insurance": "80％共保条款",
        "co-insurance": "共保条款",
        "outside ancilliary devices of buildings": "建筑物外部附属设施扩展条款",
        "outside ancillary devices": "建筑物外部附属设施扩展条款",
        "contract price": "合同价格扩展条款",
        "preference of original supplier": "优先使用原供应商条款",
        "original supplier": "优先使用原供应商条款",
        "nominated loss adjuster clause": "指定公估人条款",
        "nominated loss adjuster": "指定公估人条款",
        "loss adjuster clause": "指定公估人条款",
        "portable devices extension clause": "便携式设备扩展条款",
        "portable devices extension": "便携式设备扩展条款",
        "portable devices": "便携式设备扩展条款",
        "breakage of glass extension clause": "玻璃破碎条款",
        "breakage of glass": "玻璃破碎条款",
        "glass breakage": "玻璃破碎条款",
        "property under care, custody and control clause": "被保险人照料、保管或控制的第三方财产条款",
        "care custody and control": "被保险人照料、保管或控制的第三方财产条款",
        "mortgage clause": "抵押权条款",
        "water tank and water pipe burst extension clause": "水箱、水管爆裂扩展条款",
        "water tank and water pipe burst": "水箱、水管爆裂扩展条款",
        "nature and gradual loss exclusion": "自然及渐变损失澄清条款",
        "gradual loss": "自然及渐变损失澄清条款",
        "insured amount breakdown clause": "保险金额分项条款",
        
        # ===== 营业中断险条款 =====
        "scope of cover clause": "保单责任保障",
        "scope of cover": "保单责任保障",
        "maintenance cost clause": "全部维持费用投保条款",
        "100% payroll clause": "100％工资投保条款",
        "100% payroll": "100％工资投保条款",
        "earthquake extension clause": "地震、海啸营业中断扩展条款",
        "loss of book debts clause": "遗失欠款帐册条款",
        "loss of book debts": "遗失欠款帐册条款",
        "accumulated stock clause": "累积库存条款",
        "accumulated stocks": "累积库存条款",
        "output option clause": "产出替代条款",
        "output option": "产出替代条款",
        "prevention of access clause": "通道堵塞条款",
        "prevention of access": "通道堵塞条款",
        "denial of access": "通道堵塞条款",
        "extra expenses": "额外费用条款",
        "failure of public utilities clause": "公共事业设备失灵扩展条款",
        "failure of public utilities": "公共事业设备失灵扩展条款",
        "interdependency extension clause": "关联扩展条款",
        "interdependency extension": "关联扩展条款",
        "interdependency": "关联扩展条款",
        "suppliers, customers and contractors' premises clause": "顾客/供应商/承包商条款",
        "suppliers customers and contractors": "顾客/供应商/承包商条款",
        "murder, suicide or disease or defective sanitation clause": "谋杀、自杀、疾病或卫生设施缺陷条款",
        "murder suicide or disease": "谋杀、自杀、疾病或卫生设施缺陷条款",
        "new business clause": "新营业条款",
        "new business": "新营业条款",
        "premium adjustment clause": "保费调整条款",
        "premium adjustment": "保费调整条款",
        "waiver of average clause": "放弃比例分摊条款",
        "waiver of average": "放弃比例分摊条款",
        "leased building/premises extension": "租赁房屋/场所扩展条款",
        "leased building premises extension": "租赁房屋/场所扩展条款",
        "leased premises": "租赁房屋/场所扩展条款",
        "continuous loss clause": "持续损失条款",
        "continuous loss": "持续损失条款",
        "waive deductible clause": "物质损失放弃免赔条款",
        "waive deductible": "物质损失放弃免赔条款",
        
        # ===== 机器损坏险条款 =====
        "boiler and pressure vessel explosion clause": "锅炉及压力容器爆炸条款",
        "boiler and pressure vessel": "锅炉及压力容器爆炸条款",
        "boiler explosion": "锅炉及压力容器爆炸条款",
        "manufacturer or supplier's warranties": "制造商/供应商担保条款",
        "manufacturer warranties": "制造商/供应商担保条款",
        "acquisitions": "获得新设备条款",
        "shutdown and restart cost clause": "停机及重启损失条款",
        "shutdown and restart": "停机及重启损失条款",
        "operating media, vulnerable or consumables property extension": "媒介物、易损、易耗品扩展条款",
        "operating media": "媒介物、易损、易耗品扩展条款",
        "intellectual property protection": "知识产权保护条款",
    }
    
    # ========================================
    # 🔄 语义别名映射（解决同一概念不同表述）
    # ========================================
    SEMANTIC_ALIAS_MAP: Dict[str, str] = {
        # 污染相关
        "污染保险": "意外污染责任",
        "污染责任": "意外污染责任", 
        "意外污染": "意外污染责任",
        
        # 露天财产
        "保险标的置存处所保险": "露天及简易建筑内存放财产",
        "置存处所": "露天及简易建筑内存放财产",
        "露天财产": "露天及简易建筑内存放财产",
        "简易建筑": "露天及简易建筑内存放财产",
        
        # 损害防止/施救
        "损害防止保险条款": "阻止损失",
        "损害防止": "阻止损失",
        "施救费用": "阻止损失",
        "sue and labor": "阻止损失",
        
        # 崩塌沉降
        "崩塌与沉降保险条款": "地面突然下陷下沉",
        "崩塌与沉降": "地面突然下陷下沉",
        "崩塌沉降": "地面突然下陷下沉",
        "地面下陷": "地面突然下陷下沉",
        "地面下沉": "地面突然下陷下沉",
        "地陷下沉": "地面突然下陷下沉",
        "subsidence": "地面突然下陷下沉",
        
        # 重置价值变体
        "重置(价值)": "重置价值",
        "重建价值": "重置价值",
        "replacement": "重置价值",
        
        # 时间调整变体
        "时间调整": "72小时",
        "72hours": "72小时",
        "seventy two hours": "72小时",
        
        # 公共当局变体
        "公共当局": "公共当局扩展",
        "civil authority": "公共当局扩展",
        "public authority": "公共当局扩展",
        
        # 其他映射
        "水渍险": "水渍",
        "水损": "水渍",
        "暴风雨": "暴风暴雨洪水",
        "洪水": "暴风暴雨洪水",
    }
    
    # ========================================
    # 🔑 关键词提取映射
    # ========================================
    KEYWORD_EXTRACT_MAP: Dict[str, List[str]] = {
        "污染": ["污染", "意外污染", "pollution"],
        "露天": ["露天", "简易", "置存处所", "outdoor"],
        "阻止": ["阻止", "损害防止", "施救", "sue labor"],
        "下陷": ["下陷", "下沉", "崩塌", "沉降", "subsidence"],
        "地震": ["地震", "震动", "earthquake"],
        "海啸": ["海啸", "tsunami"],
        "盗窃": ["盗窃", "盗抢", "抢劫", "burglary", "theft", "robbery"],
        "火灾": ["火灾", "火险", "fire"],
        "洪水": ["洪水", "水灾", "flood"],
        "重置": ["重置", "重建", "reinstatement", "replacement"],
        "时间调整": ["时间调整", "72小时", "72hours", "seventy-two", "time adjustment"],
        "公共当局": ["公共当局", "civil authority", "public authority"],
        "通知": ["通知", "notification", "notice"],
        "错误遗漏": ["错误", "遗漏", "errors", "omissions"],
        "控制": ["控制", "control"],
        "通道堵塞": ["通道堵塞", "通道", "堵塞", "prevention of access", "denial of access"],
        "关联": ["关联", "interdependency"],
        "累积库存": ["累积库存", "累积存货", "accumulated stock"],
    }
    
    # ========================================
    # 📌 精确条款名映射（客户条款 -> 条款库名称）
    # ========================================
    EXACT_CLAUSE_MAP: Dict[str, str] = {
        "72小时条款": "时间调整（72小时）",
        "时间调整条款": "时间调整",
        "通道堵塞条款": "通道堵塞",
        "关联扩展条款": "关联扩展",
        "累积库存条款": "累积库存",
    }
    
    # ========================================
    # ⚠️ 惩罚关键词
    # ========================================
    PENALTY_KEYWORDS: List[str] = ["打孔盗气"]
    
    # ========================================
    # 🏷️ 噪音词（清理时移除）
    # ========================================
    NOISE_WORDS: List[str] = [
        "企业财产保险", "附加", "扩展", "条款", "险", 
        "（A款）", "（B款）", "(A款)", "(B款)",
        "2025版", "2024版", "2023版", "2022版", "版",
        "clause", "extension", "cover", "insurance",
    ]


# ==========================================
# 数据结构
# ==========================================
class MatchLevel(Enum):
    """匹配级别"""
    EXACT = "精确匹配"
    SEMANTIC = "语义匹配"
    KEYWORD = "关键词匹配"
    FUZZY = "模糊匹配"
    NONE = "无匹配"

@dataclass
class ClauseItem:
    """条款项"""
    title: str
    content: str
    original_title: str = ""  # 保留原始标题（英文）

@dataclass 
class MatchResult:
    """匹配结果"""
    matched_name: str = ""
    matched_content: str = ""
    matched_reg: str = ""
    score: float = 0.0
    title_score: float = 0.0
    content_score: float = 0.0
    match_level: MatchLevel = MatchLevel.NONE
    diff_analysis: str = ""


# ==========================================
# 核心匹配逻辑
# ==========================================
class ClauseMatcherLogic:
    """条款匹配核心逻辑"""
    
    config = ClauseConfig
    
    @classmethod
    def normalize_text(cls, text: str) -> str:
        """标准化文本（小写、去空格、去标点）"""
        if not isinstance(text, str):
            return ""
        text = text.lower().strip()
        text = re.sub(r"['\"\'\'\"\"\(\)（）\[\]【】]", '', text)
        text = re.sub(r'\s+', ' ', text)
        return text
    
    @classmethod
    def clean_title(cls, text: str) -> str:
        """清理标题用于比较"""
        if not isinstance(text, str): 
            return ""
        # 移除括号内容
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        # 移除噪音词
        for w in cls.config.NOISE_WORDS: 
            text = text.replace(w, "").replace(w.lower(), "")
        # 移除数字和空格
        text = re.sub(r'[0-9\s]+', '', text)
        return text.strip()

    @classmethod
    def clean_content(cls, text: str) -> str:
        """清理内容用于比较"""
        if not isinstance(text, str): 
            return ""
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        text = re.sub(r'\s+', '', text)
        text = re.sub(r'[0-9]+', '', text)
        return text

    @classmethod
    def extract_extra_info(cls, text: str) -> str:
        """提取括号内的额外信息（限额等）"""
        if not isinstance(text, str): 
            return ""
        matches = re.findall(r'([\(（].*?[\)）])', text)
        return " ".join(matches) if matches else ""

    @classmethod
    def is_english(cls, text: str) -> bool:
        """判断是否为英文"""
        if not isinstance(text, str) or len(text) <= 3: 
            return False
        zh_count = len(re.findall(r'[\u4e00-\u9fa5]', text))
        return zh_count < len(text) * 0.15

    @classmethod
    def translate_title(cls, title: str) -> Tuple[str, bool]:
        """翻译英文标题为中文"""
        if not cls.is_english(title):
            return title, False
        
        title_norm = cls.normalize_text(title)
        
        # 1. 精确匹配客户字典
        if title_norm in cls.config.CLIENT_EN_CN_MAP:
            return cls.config.CLIENT_EN_CN_MAP[title_norm], True
        
        # 2. 部分匹配客户字典
        for eng, chn in cls.config.CLIENT_EN_CN_MAP.items():
            if eng in title_norm or title_norm in eng:
                return chn, True
        
        # 3. 使用在线翻译
        if HAS_TRANSLATOR:
            try:
                translated = GoogleTranslator(source='auto', target='zh-CN').translate(title)
                return translated, True
            except:
                pass
        
        return title, False

    @classmethod
    def extract_keywords(cls, text: str) -> Set[str]:
        """从文本中提取关键词"""
        keywords = set()
        text_lower = text.lower()
        for core, variants in cls.config.KEYWORD_EXTRACT_MAP.items():
            for v in variants:
                if v.lower() in text_lower:
                    keywords.add(core)
                    break
        return keywords
    
    @classmethod
    def check_semantic_alias(cls, title: str) -> Optional[str]:
        """检查语义别名匹配"""
        title_clean = title.replace(" ", "").lower()
        for alias, target in cls.config.SEMANTIC_ALIAS_MAP.items():
            if alias.lower() in title_clean:
                return target
        return None

    @classmethod
    def calculate_similarity(cls, text1: str, text2: str) -> float:
        """计算文本相似度"""
        if not text1 or not text2:
            return 0.0
        return difflib.SequenceMatcher(None, text1, text2).ratio()

    @classmethod
    def match_clause(cls, clause: ClauseItem, lib_data: List[Dict], 
                     is_title_only: bool) -> MatchResult:
        """
        多级匹配策略：
        1. 精确条款名映射
        2. 精确匹配（翻译后直接匹配）
        3. 语义别名匹配
        4. 关键词匹配
        5. 模糊匹配
        """
        result = MatchResult()
        title = clause.title
        content = clause.content
        
        # 准备清理后的标题
        title_clean = cls.clean_title(title)
        title_norm = cls.normalize_text(title)
        
        best_score = -100
        best_match = None
        best_meta = {'t': 0, 'c': 0, 'level': MatchLevel.NONE}
        
        # 提取客户条款关键词
        c_keywords = cls.extract_keywords(title)
        
        # 检查语义别名
        semantic_target = cls.check_semantic_alias(title)
        
        # 检查精确条款名映射
        exact_target = None
        for src, tgt in cls.config.EXACT_CLAUSE_MAP.items():
            if src in title or src in title_clean:
                exact_target = tgt
                break
        
        for lib in lib_data:
            l_name = str(lib.get('条款名称', ''))
            l_content = str(lib.get('条款内容', ''))
            l_name_clean = cls.clean_title(l_name)
            l_name_norm = cls.normalize_text(l_name)
            
            score = 0.0
            match_level = MatchLevel.FUZZY
            
            # === 级别0: 精确条款名映射 ===
            if exact_target and exact_target in l_name:
                score = 0.98
                match_level = MatchLevel.EXACT
                best_score = score
                best_match = lib
                best_meta = {'t': 0.98, 'c': 0, 'level': MatchLevel.EXACT}
                break
            
            # === 级别1: 精确匹配 ===
            if title_clean == l_name_clean or title_norm == l_name_norm:
                score = 1.0
                match_level = MatchLevel.EXACT
            
            # === 级别2: 语义别名匹配 ===
            elif semantic_target and semantic_target in l_name:
                score = 0.95
                match_level = MatchLevel.SEMANTIC
            
            else:
                # === 级别3: 关键词匹配 ===
                l_keywords = cls.extract_keywords(l_name)
                if c_keywords and l_keywords:
                    common = c_keywords & l_keywords
                    if common:
                        keyword_score = len(common) / max(len(c_keywords), len(l_keywords))
                        if keyword_score >= 0.5:
                            score = 0.7 + keyword_score * 0.2
                            match_level = MatchLevel.KEYWORD
                
                # === 级别4: 模糊匹配 ===
                if score < 0.7:
                    title_sim = cls.calculate_similarity(title_clean, l_name_clean)
                    
                    # 内容相似度
                    content_sim = 0.0
                    if not is_title_only and content.strip():
                        c_content_clean = cls.clean_content(content)
                        l_content_clean = cls.clean_content(l_content)
                        if c_content_clean and l_content_clean:
                            content_sim = cls.calculate_similarity(c_content_clean, l_content_clean)
                    
                    # 加权计算
                    if is_title_only or not content.strip():
                        score = title_sim
                    else:
                        score = 0.7 * title_sim + 0.3 * content_sim
                    
                    best_meta['t'] = title_sim
                    best_meta['c'] = content_sim
                    match_level = MatchLevel.FUZZY
            
            # 惩罚项
            for bad_word in cls.config.PENALTY_KEYWORDS:
                if bad_word in l_name and bad_word not in title:
                    score -= 0.5
            
            if score > best_score:
                best_score = score
                best_match = lib
                best_meta['level'] = match_level
                if match_level in [MatchLevel.EXACT, MatchLevel.SEMANTIC, MatchLevel.KEYWORD]:
                    best_meta['t'] = score
        
        # 构建结果
        if best_match and best_score > 0.15:
            base_name = best_match.get('条款名称', '')
            extra_params = cls.extract_extra_info(clause.original_title or clause.title)
            
            result.matched_name = f"{base_name} {extra_params}".strip() if extra_params else base_name
            result.matched_content = best_match.get('条款内容', '')
            result.matched_reg = best_match.get('产品注册号', best_match.get('注册号', ''))
            result.score = max(0, best_score)
            result.title_score = best_meta.get('t', 0)
            result.content_score = best_meta.get('c', 0)
            result.match_level = best_meta.get('level', MatchLevel.FUZZY)
            
            # 差异分析
            if best_score < 0.6:
                result.diff_analysis = cls.analyze_difference(content, result.matched_content)
        
        return result

    @classmethod
    def analyze_difference(cls, c_content: str, l_content: str) -> str:
        """分析保障差异"""
        c_text, l_text = str(c_content), str(l_content)
        if not c_text.strip(): 
            return ""

        analysis = []
        keywords = {
            "限额": ["Limit", "限额", "最高", "limit"],
            "免赔": ["Deductible", "Excess", "免赔", "deductible"],
            "除外": ["Exclusion", "除外", "不负责", "exclusion"],
            "观察期": ["Waiting Period", "观察期", "等待期"],
            "赔偿期": ["Indemnity Period", "赔偿期间"],
        }
        
        for key, words in keywords.items():
            c_has = any(w.lower() in c_text.lower() for w in words)
            l_has = any(w.lower() in l_text.lower() for w in words)
            if c_has and not l_has: 
                analysis.append(f"⚠️ 客户提及[{key}]但库内未提及")
            elif not c_has and l_has: 
                analysis.append(f"ℹ️ 库内包含[{key}]但客户未提及")

        return " | ".join(analysis)

    @classmethod
    def is_likely_title(cls, text: str) -> bool:
        """判断是否像标题"""
        if len(text) > 80: 
            return False
        if text.endswith(('。', '；', '.', ';')): 
            return False
        title_indicators = ["条款", "Clause", "Extension", "险", "CLAUSE", "EXTENSION"]
        if any(kw in text for kw in title_indicators):
            return True
        # 全大写英文通常是标题
        if text.isupper() and len(text) > 5:
            return True
        return True

    @classmethod
    def parse_docx(cls, doc_path: str) -> Tuple[List[ClauseItem], bool]:
        """解析Word文档，提取条款"""
        doc = Document(doc_path)
        clauses = []
        current_block = []
        
        all_lines = [p.text.strip() for p in doc.paragraphs]
        empty_lines = sum(1 for t in all_lines if not t)
        
        # 智能分割策略
        use_smart_split = len(all_lines) > 0 and (empty_lines / max(len(all_lines), 1) < 0.05)
        
        if use_smart_split:
            for text in all_lines:
                if not text: 
                    continue
                if current_block and cls.is_likely_title(text):
                    title = current_block[0]
                    content = "\n".join(current_block[1:])
                    clauses.append(ClauseItem(title=title, content=content, original_title=title))
                    current_block = [text]
                else:
                    current_block.append(text)
            if current_block:
                clauses.append(ClauseItem(
                    title=current_block[0], 
                    content="\n".join(current_block[1:]),
                    original_title=current_block[0]
                ))
        else:
            for text in all_lines:
                if text:
                    current_block.append(text)
                elif current_block:
                    clauses.append(ClauseItem(
                        title=current_block[0], 
                        content="\n".join(current_block[1:]),
                        original_title=current_block[0]
                    ))
                    current_block = []
            if current_block:
                clauses.append(ClauseItem(
                    title=current_block[0], 
                    content="\n".join(current_block[1:]),
                    original_title=current_block[0]
                ))
        
        is_title_only = all(not c.content for c in clauses)
        return clauses, is_title_only


# ==========================================
# 工作线程
# ==========================================
class MatchWorker(QThread):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, doc_path: str, excel_path: str, output_path: str):
        super().__init__()
        self.doc_path = doc_path
        self.excel_path = excel_path
        self.output_path = output_path
        
    def run(self):
        try:
            logic = ClauseMatcherLogic
            
            if not HAS_TRANSLATOR:
                self.log_signal.emit("⚠️ 未检测到 deep_translator，仅使用内置术语表", "warning")
            else:
                self.log_signal.emit("✓ 已启用在线翻译支持", "success")

            self.log_signal.emit("⏳ 正在解析文档...", "info")
            clauses, is_title_only = logic.parse_docx(self.doc_path)
            mode_str = "纯标题模式" if is_title_only else "完整内容模式"
            self.log_signal.emit(f"📖 [{mode_str}] 提取到 {len(clauses)} 条", "success")
            
            # 加载条款库
            lib_df = pd.read_excel(self.excel_path, header=1)
            lib_df.columns = [str(c).strip() for c in lib_df.columns]
            
            # 识别列名
            name_col = None
            content_col = None
            reg_col = None
            for col in lib_df.columns:
                if '条款名称' in col or '名称' in col:
                    name_col = col
                elif '条款内容' in col or '内容' in col:
                    content_col = col
                elif '注册号' in col or '产品' in col:
                    reg_col = col
            
            if not name_col:
                name_col = lib_df.columns[0]
            if not content_col and len(lib_df.columns) > 2:
                content_col = lib_df.columns[2]
            if not reg_col and len(lib_df.columns) > 1:
                reg_col = lib_df.columns[1]
                
            # 标准化数据
            lib_data = []
            for _, row in lib_df.iterrows():
                lib_data.append({
                    '条款名称': str(row.get(name_col, '')) if pd.notna(row.get(name_col)) else '',
                    '条款内容': str(row.get(content_col, '')) if content_col and pd.notna(row.get(content_col)) else '',
                    '产品注册号': str(row.get(reg_col, '')) if reg_col and pd.notna(row.get(reg_col)) else '',
                })
            
            lib_data = [d for d in lib_data if d['条款名称'].strip()]
            self.log_signal.emit(f"📚 加载条款库 {len(lib_data)} 条", "info")
            
            self.log_signal.emit("🧠 开始智能匹配（多级策略）...", "info")
            results = []
            
            stats = {'exact': 0, 'semantic': 0, 'keyword': 0, 'fuzzy': 0, 'none': 0}
            
            for idx, clause in enumerate(clauses, 1):
                self.progress_signal.emit(idx, len(clauses))
                
                # 翻译处理
                original_title = clause.title
                translated_title, was_translated = logic.translate_title(clause.title)
                
                if was_translated:
                    clause.title = translated_title
                    clause.original_title = original_title
                    if clause.content and logic.is_english(clause.content):
                        try:
                            clause.content = GoogleTranslator(source='auto', target='zh-CN').translate(clause.content) if HAS_TRANSLATOR else clause.content
                        except:
                            pass

                # 执行匹配
                match_result = logic.match_clause(clause, lib_data, is_title_only)
                
                # 统计
                if match_result.match_level == MatchLevel.EXACT:
                    stats['exact'] += 1
                elif match_result.match_level == MatchLevel.SEMANTIC:
                    stats['semantic'] += 1
                elif match_result.match_level == MatchLevel.KEYWORD:
                    stats['keyword'] += 1
                elif match_result.match_level == MatchLevel.FUZZY:
                    stats['fuzzy'] += 1
                else:
                    stats['none'] += 1

                results.append({
                    '序号': idx,
                    '客户条款(原)': original_title,
                    '客户条款(译)': translated_title if was_translated else "",
                    '客户原始内容': clause.content[:500] if clause.content else "", 
                    '匹配条款库名称': match_result.matched_name or "无匹配",
                    '产品注册号': match_result.matched_reg,
                    '匹配条款库内容': match_result.matched_content[:500] if match_result.matched_content else "",
                    '综合匹配度': round(match_result.score, 3),
                    '匹配级别': match_result.match_level.value,
                    '保障差异提示': match_result.diff_analysis,
                    '标题相似度': round(match_result.title_score, 3),
                    '内容相似度': round(match_result.content_score, 3),
                })
                
                if idx % 10 == 0:
                    self.log_signal.emit(f"   已处理 {idx}/{len(clauses)}...", "info")
            
            # 保存结果
            df_res = pd.DataFrame(results)
            df_res.to_excel(self.output_path, index=False)
            self._apply_excel_styles()
            
            # 输出统计
            self.log_signal.emit(f"📊 匹配统计:", "info")
            self.log_signal.emit(f"   精确匹配: {stats['exact']}", "success")
            self.log_signal.emit(f"   语义匹配: {stats['semantic']}", "success")
            self.log_signal.emit(f"   关键词匹配: {stats['keyword']}", "info")
            self.log_signal.emit(f"   模糊匹配: {stats['fuzzy']}", "warning")
            self.log_signal.emit(f"   无匹配: {stats['none']}", "error")
            
            self.log_signal.emit(f"🎉 完成！已生成报告", "success")
            self.finished_signal.emit(True, self.output_path)
            
        except Exception as e:
            self.log_signal.emit(f"❌ 错误: {str(e)}", "error")
            self.log_signal.emit(traceback.format_exc(), "error")
            self.finished_signal.emit(False, str(e))

    def _apply_excel_styles(self):
        """应用Excel样式"""
        wb = openpyxl.load_workbook(self.output_path)
        wb.properties.creator = "Dachi Yijin"
        ws = wb.active
        
        fills = {
            'green': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'yellow': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'blue': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            'header': PatternFill(start_color="667eea", end_color="667eea", fill_type="solid"),
        }
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # 表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = fills['header']
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # 列宽设置
        widths = {
            'A': 6, 'B': 35, 'C': 30, 'D': 45, 'E': 40, 
            'F': 25, 'G': 50, 'H': 10, 'I': 12, 'J': 35, 'K': 10, 'L': 10
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        
        # 数据行样式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border
                
                # 匹配度着色 (H列)
                if cell.col_idx == 8:
                    try:
                        val = float(cell.value) if cell.value else 0
                        if val >= 0.8:
                            cell.fill = fills['green']
                        elif val >= 0.5:
                            cell.fill = fills['yellow']
                        elif val > 0:
                            cell.fill = fills['red']
                    except:
                        pass
                
                # 匹配级别着色 (I列)
                if cell.col_idx == 9:
                    val = str(cell.value) if cell.value else ""
                    if "精确" in val:
                        cell.fill = fills['green']
                    elif "语义" in val:
                        cell.fill = fills['blue']
                    elif "关键词" in val:
                        cell.fill = fills['yellow']
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        wb.save(self.output_path)


# ==========================================
# UI组件
# ==========================================
class GlassCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            GlassCard {
                background: rgba(255, 255, 255, 0.08);
                border: 1px solid rgba(255, 255, 255, 0.15);
                border-radius: 20px;
            }
            QLabel { color: #ffffff; font-weight: 500; }
            QLineEdit {
                background: rgba(0,0,0,0.2);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 10px;
                padding: 12px 15px;
                color: #ffffff;
                font-size: 14px;
            }
            QLineEdit:focus { border-color: #667eea; }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(40)
        shadow.setColor(QColor(0, 0, 0, 80))
        shadow.setOffset(0, 10)
        self.setGraphicsEffect(shadow)


class ClauseDiffGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("智能条款比对工具 v14.0")
        self.setMinimumSize(900, 750)
        self.setStyleSheet("""
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                    stop:0 #1a1a2e, stop:0.5 #16213e, stop:1 #0f3460);
            }
        """)
        self._setup_ui()

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(25)
        layout.setContentsMargins(40, 40, 40, 40)

        # 标题
        title_box = QVBoxLayout()
        title = QLabel("🔍 智能条款比对工具")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #ffffff; font-size: 32px; font-weight: bold; letter-spacing: 2px;")
        subtitle = QLabel("v14.0 Client Mapping Enhanced · 多级匹配策略")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: rgba(255,255,255,0.6); font-size: 14px; font-weight: 300; letter-spacing: 2px;")
        title_box.addWidget(title)
        title_box.addWidget(subtitle)
        layout.addLayout(title_box)

        # 输入卡片
        card = GlassCard()
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(35, 35, 35, 35)
        
        btn_style = """
            QPushButton {
                background: rgba(255,255,255,0.1);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 10px;
                padding: 12px 20px;
                color: #ffffff;
                font-weight: 500;
            }
            QPushButton:hover {
                background: rgba(255,255,255,0.2);
                border-color: #667eea;
            }
            QPushButton:pressed { background: rgba(102,126,234,0.3); }
        """

        self.doc_input = self._create_file_row(card_layout, "📂 客户文档", "支持中英文 Word 条款清单...", "Word Files (*.docx)", btn_style)
        self.lib_input = self._create_file_row(card_layout, "📚 标准题库", "选择 Excel 条款库...", "Excel Files (*.xlsx)", btn_style)
        
        line = QFrame()
        line.setFixedHeight(1)
        line.setStyleSheet("background: rgba(255,255,255,0.1);")
        card_layout.addWidget(line)

        row3 = QHBoxLayout()
        label3 = QLabel("💾 保存路径")
        label3.setFixedWidth(90)
        self.out_input = QLineEdit()
        self.out_input.setPlaceholderText("设置报告保存位置...")
        btn3 = QPushButton("选择")
        btn3.setCursor(Qt.PointingHandCursor)
        btn3.setStyleSheet(btn_style)
        btn3.clicked.connect(self._browse_save)
        row3.addWidget(label3)
        row3.addWidget(self.out_input, 1)
        row3.addWidget(btn3)
        card_layout.addLayout(row3)

        layout.addWidget(card)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(20)
        
        self.start_btn = QPushButton("🚀 开始智能比对")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.setMinimumHeight(60)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #667eea, stop:1 #764ba2);
                color: white; font-size: 18px; font-weight: bold;
                border-radius: 30px; border: none;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #764ba2, stop:1 #667eea);
            }
            QPushButton:pressed { padding-top: 3px; }
            QPushButton:disabled { background: rgba(255,255,255,0.1); color: rgba(255,255,255,0.3); }
        """)
        self.start_btn.clicked.connect(self._start_process)
        
        self.open_btn = QPushButton("📂 打开目录")
        self.open_btn.setCursor(Qt.PointingHandCursor)
        self.open_btn.setMinimumHeight(60)
        self.open_btn.setEnabled(False)
        self.open_btn.setStyleSheet("""
            QPushButton {
                background: transparent; color: rgba(255,255,255,0.6);
                font-size: 16px; font-weight: 500;
                border-radius: 30px; border: 2px solid rgba(255,255,255,0.2);
            }
            QPushButton:hover { border-color: #27ae60; color: #27ae60; }
            QPushButton:disabled { color: rgba(255,255,255,0.2); border-color: rgba(255,255,255,0.1); }
        """)
        self.open_btn.clicked.connect(self._open_output_folder)

        btn_layout.addWidget(self.start_btn, 2)
        btn_layout.addWidget(self.open_btn, 1)
        layout.addLayout(btn_layout)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(4)
        self.progress_bar.setStyleSheet("""
            QProgressBar { background: rgba(255,255,255,0.1); border-radius: 2px; }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #667eea, stop:1 #764ba2);
                border-radius: 2px;
            }
        """)
        layout.addWidget(self.progress_bar)

        # 日志
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background: rgba(0,0,0,0.3);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 15px;
                color: #e8e8e8;
                padding: 20px;
                font-family: 'SF Mono', 'Menlo', 'Monaco', monospace;
                font-size: 13px;
            }
        """)
        layout.addWidget(self.log_text, 1)
        
        version_label = QLabel("v14.0 Client Mapping Enhanced · Made with ❤️")
        version_label.setAlignment(Qt.AlignCenter)
        version_label.setStyleSheet("color: rgba(255,255,255,0.3); font-size: 12px;")
        layout.addWidget(version_label)

    def _create_file_row(self, layout, label_text: str, placeholder: str, filter_str: str, btn_style: str) -> QLineEdit:
        row = QHBoxLayout()
        label = QLabel(label_text)
        label.setFixedWidth(90)
        line_edit = QLineEdit()
        line_edit.setPlaceholderText(placeholder)
        btn = QPushButton("浏览")
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(btn_style)
        btn.clicked.connect(lambda: self._browse_file(line_edit, filter_str))
        row.addWidget(label)
        row.addWidget(line_edit, 1)
        row.addWidget(btn)
        layout.addLayout(row)
        return line_edit

    def _browse_file(self, line_edit: QLineEdit, filter_str: str):
        f, _ = QFileDialog.getOpenFileName(self, "选择文件", "", filter_str)
        if f: 
            line_edit.setText(f)
            if line_edit == self.doc_input and not self.out_input.text():
                self.out_input.setText(os.path.join(os.path.dirname(f), "条款比对报告.xlsx"))

    def _browse_save(self):
        f, _ = QFileDialog.getSaveFileName(self, "保存结果", "条款比对报告.xlsx", "Excel Files (*.xlsx)")
        if f: 
            self.out_input.setText(f)

    def _append_log(self, msg: str, level: str):
        colors = {"info": "#a0a0a0", "success": "#2ecc71", "error": "#e74c3c", "warning": "#f39c12"}
        self.log_text.append(f'<span style="color:{colors.get(level, "#fff")}">{msg}</span>')
        self.log_text.moveCursor(QTextCursor.End)

    def _start_process(self):
        doc = self.doc_input.text().strip()
        excel = self.lib_input.text().strip()
        out = self.out_input.text().strip()
        
        if not all([doc, excel, out]):
            QMessageBox.warning(self, "提示", "请先完善所有文件路径！")
            return
            
        self.start_btn.setEnabled(False)
        self.open_btn.setEnabled(False)
        self.start_btn.setText("⏳ 正在计算中...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        
        self.worker = MatchWorker(doc, excel, out)
        self.worker.log_signal.connect(self._append_log)
        self.worker.progress_signal.connect(lambda c, t: self.progress_bar.setValue(int(c/t*100)))
        self.worker.finished_signal.connect(self._on_finished)
        self.worker.start()

    def _on_finished(self, success: bool, msg: str):
        self.start_btn.setEnabled(True)
        self.start_btn.setText("🚀 开始智能比对")
        self.progress_bar.setVisible(False)
        
        if success:
            self.open_btn.setEnabled(True)
            self.open_btn.setStyleSheet("""
                QPushButton {
                    background: transparent; color: #2ecc71;
                    font-size: 16px; font-weight: 500;
                    border-radius: 30px; border: 2px solid #2ecc71;
                }
                QPushButton:hover { background: #2ecc71; color: white; }
            """)
            QMessageBox.information(self, "完成", f"比对完成！\n文件已保存至:\n{msg}")

    def _open_output_folder(self):
        path = self.out_input.text().strip()
        if path and os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))
        else:
            QMessageBox.warning(self, "提示", "文件路径不存在！")


def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    
    app = QApplication(sys.argv)
    app.setFont(QFont("PingFang SC", 13))
    
    window = ClauseDiffGUI()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()

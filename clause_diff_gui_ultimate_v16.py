# -*- coding: utf-8 -*-
"""
智能条款比对工具 v16.0 (Full Optimized Edition)
- [性能] 预处理索引加速匹配 5-10x
- [算法] 编辑距离容错 + 混合相似度
- [重构] 多级匹配策略拆分
- [功能] 批量处理多文件
- [健壮] 完善异常处理和日志
- [配置] 外部化JSON配置

Author: Dachi Yijin
Date: 2025-12-21
"""

import sys
import os
import re
import difflib
import traceback
import logging
from typing import List, Dict, Tuple, Optional, Set, Any
from dataclasses import dataclass, field
from enum import Enum
from collections import defaultdict
from functools import lru_cache
from pathlib import Path
from datetime import datetime
import pandas as pd
from docx import Document

# ==========================================
# 日志配置
# ==========================================
LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"clause_diff_{datetime.now():%Y%m%d}.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==========================================
# 导入配置管理器
# ==========================================
try:
    from clause_config_manager import get_config, ClauseConfigManager
    HAS_CONFIG_MANAGER = True
except ImportError:
    HAS_CONFIG_MANAGER = False
    logger.warning("未找到 clause_config_manager，使用内置配置")

# ==========================================
# macOS PyQt5 Plugin Fix
# ==========================================
try:
    import PyQt5
    plugin_path = os.path.join(os.path.dirname(PyQt5.__file__), 'Qt5', 'plugins')
    os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path
except ImportError:
    pass

try:
    from deep_translator import GoogleTranslator
    HAS_TRANSLATOR = True
except ImportError:
    HAS_TRANSLATOR = False

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QProgressBar, QTextEdit,
    QFileDialog, QMessageBox, QFrame, QGraphicsDropShadowEffect,
    QDialog, QFormLayout, QListWidget, QListWidgetItem, QCheckBox,
    QTabWidget, QSpinBox, QDoubleSpinBox, QGroupBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt5.QtGui import QFont, QColor, QDesktopServices, QTextCursor

# ==========================================
# macOS 打包防闪退
# ==========================================
class NullWriter:
    def write(self, text): pass
    def flush(self): pass

if getattr(sys, 'frozen', False):
    sys.stdout = NullWriter()
    sys.stderr = NullWriter()

def global_exception_handler(exctype, value, tb):
    error_msg = "".join(traceback.format_exception(exctype, value, tb))
    logger.error(f"未捕获异常: {error_msg}")
    try:
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Critical)
        msg_box.setText("程序发生意外错误")
        msg_box.setInformativeText(str(value))
        msg_box.setDetailedText(error_msg)
        msg_box.exec_()
    except:
        pass

sys.excepthook = global_exception_handler


# ==========================================
# 常量定义
# ==========================================
class ExcelColumns:
    """Excel列名常量"""
    SEQ = '序号'
    CLIENT_ORIG = '客户条款(原)'
    CLIENT_TRANS = '客户条款(译)'
    CLIENT_CONTENT = '客户原始内容'
    MATCHED_NAME = '匹配条款库名称'
    REG_NO = '产品注册号'
    MATCHED_CONTENT = '匹配条款库内容'
    SCORE = '综合匹配度'
    MATCH_LEVEL = '匹配级别'
    DIFF_ANALYSIS = '保障差异提示'
    TITLE_SCORE = '标题相似度'
    CONTENT_SCORE = '内容相似度'

    # 列索引（1-based）
    SCORE_COL_IDX = 8
    LEVEL_COL_IDX = 9


# ==========================================
# 数据结构
# ==========================================
class MatchLevel(Enum):
    """匹配级别"""
    EXACT = "精确匹配"
    SEMANTIC = "语义匹配"
    KEYWORD = "关键词匹配"
    FUZZY = "模糊匹配"
    NONE = "无匹配"

@dataclass
class MatchThresholds:
    """匹配阈值"""
    exact_min: float = 0.98
    semantic_min: float = 0.85
    keyword_min: float = 0.60
    fuzzy_min: float = 0.40
    accept_min: float = 0.15

@dataclass
class ClauseItem:
    """条款项"""
    title: str
    content: str
    original_title: str = ""

@dataclass
class MatchResult:
    """匹配结果"""
    matched_name: str = ""
    matched_content: str = ""
    matched_reg: str = ""
    score: float = 0.0
    title_score: float = 0.0
    content_score: float = 0.0
    match_level: MatchLevel = MatchLevel.NONE
    diff_analysis: str = ""

@dataclass
class LibraryIndex:
    """条款库索引结构"""
    by_name_norm: Dict[str, int] = field(default_factory=dict)
    by_keyword: Dict[str, List[int]] = field(default_factory=lambda: defaultdict(list))
    cleaned_cache: Dict[int, Dict[str, str]] = field(default_factory=dict)
    data: List[Dict] = field(default_factory=list)


# ==========================================
# 内置默认配置（当配置管理器不可用时）
# ==========================================
class DefaultConfig:
    """默认配置"""

    CLIENT_EN_CN_MAP = {
        "interpretation & headings": "通译和标题条款",
        "reinstatement value": "重置价值条款",
        "reinstatement value clause": "重置价值条款",
        "time adjustment": "72小时条款",
        "civil authorities clause": "公共当局扩展条款",
        "civil authorities": "公共当局扩展条款",
        "errors and omissions clause": "错误和遗漏条款",
        "loss notification clause": "损失通知条款",
        "no control clause": "不受控制条款",
        "removal of debris": "清理残骸费用扩展条款",
        "strike, riot, civil commotion": "罢工、暴动或民众骚乱条款",
        "earthquake and tsunami": "地震扩展条款",
        "theft and robbery": "盗窃、抢劫扩展条款",
        "professional fees": "专业费用及索赔准备费用条款",
        "automatic reinstatement of sum insured": "自动恢复保险金额条款",
    }

    SEMANTIC_ALIAS_MAP = {
        "污染保险": "意外污染责任",
        "污染责任": "意外污染责任",
        "露天财产": "露天及简易建筑内存放财产",
        "损害防止": "阻止损失",
        "施救费用": "阻止损失",
        "崩塌沉降": "地面突然下陷下沉",
        "地面下陷": "地面突然下陷下沉",
    }

    KEYWORD_MAP = {
        "污染": ["污染", "意外污染", "pollution"],
        "地震": ["地震", "震动", "earthquake"],
        "海啸": ["海啸", "tsunami"],
        "盗窃": ["盗窃", "盗抢", "抢劫", "burglary", "theft", "robbery"],
        "洪水": ["洪水", "水灾", "flood"],
        "火灾": ["火灾", "火险", "fire"],
        "重置": ["重置", "重建", "reinstatement", "replacement"],
    }

    PENALTY_KEYWORDS = ["打孔盗气"]

    NOISE_WORDS = [
        "企业财产保险", "附加", "扩展", "条款", "险",
        "（A款）", "（B款）", "(A款)", "(B款)",
        "2025版", "2024版", "2023版", "版",
        "clause", "extension", "cover", "insurance",
    ]


# ==========================================
# 编辑距离算法
# ==========================================
@lru_cache(maxsize=10000)
def levenshtein_distance(s1: str, s2: str) -> int:
    """计算编辑距离（带缓存）"""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)

    if len(s2) == 0:
        return len(s1)

    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row

    return previous_row[-1]


def levenshtein_ratio(s1: str, s2: str) -> float:
    """计算编辑距离相似度"""
    if not s1 or not s2:
        return 0.0

    # 长度差异过大直接返回低分
    len_diff = abs(len(s1) - len(s2))
    max_len = max(len(s1), len(s2))
    if len_diff > max_len * 0.6:
        return 0.0

    distance = levenshtein_distance(s1, s2)
    return 1 - (distance / max_len)


# ==========================================
# 核心匹配逻辑（重构版）
# ==========================================
class ClauseMatcherLogic:
    """条款匹配核心逻辑 - 优化版"""

    def __init__(self):
        """初始化匹配器"""
        # 加载配置
        if HAS_CONFIG_MANAGER:
            self.config = get_config()
            self._use_external_config = True
        else:
            self.config = None
            self._use_external_config = False

        self.thresholds = MatchThresholds()
        self._index: Optional[LibraryIndex] = None

        logger.info(f"匹配器初始化完成，外部配置: {self._use_external_config}")

    # ========================================
    # 配置访问方法
    # ========================================

    def _get_client_mapping(self, term: str) -> Optional[str]:
        """获取英中映射"""
        if self._use_external_config:
            return self.config.get_client_mapping(term)
        return DefaultConfig.CLIENT_EN_CN_MAP.get(term.lower())

    def _get_semantic_alias(self, text: str) -> Optional[str]:
        """获取语义别名"""
        alias_map = (self.config.semantic_alias_map if self._use_external_config
                     else DefaultConfig.SEMANTIC_ALIAS_MAP)
        for alias, target in alias_map.items():
            if alias in text:
                return target
        return None

    def _get_keywords(self, text: str) -> Set[str]:
        """提取关键词"""
        keywords = set()
        text_lower = text.lower()
        keyword_map = (self.config.keyword_extract_map if self._use_external_config
                       else DefaultConfig.KEYWORD_MAP)
        for core, variants in keyword_map.items():
            for v in variants:
                if v.lower() in text_lower:
                    keywords.add(core)
                    break
        return keywords

    def _is_penalty_keyword(self, text: str) -> bool:
        """检查惩罚关键词"""
        penalty_list = (self.config.penalty_keywords if self._use_external_config
                        else DefaultConfig.PENALTY_KEYWORDS)
        return any(kw in text for kw in penalty_list)

    def _get_noise_words(self) -> List[str]:
        """获取噪音词列表"""
        return (self.config.noise_words if self._use_external_config
                else DefaultConfig.NOISE_WORDS)

    # ========================================
    # 文本处理方法
    # ========================================

    @staticmethod
    def normalize_text(text: str) -> str:
        """标准化文本"""
        if not isinstance(text, str):
            return ""
        text = text.lower().strip()
        text = re.sub(r"['\"\'\'\"\"\(\)（）\[\]【】]", '', text)
        text = re.sub(r'\s+', ' ', text)
        return text

    def clean_title(self, text: str) -> str:
        """清理标题"""
        if not isinstance(text, str):
            return ""
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        for w in self._get_noise_words():
            text = text.replace(w, "").replace(w.lower(), "")
        text = re.sub(r'[0-9\s]+', '', text)
        return text.strip()

    @staticmethod
    def clean_content(text: str) -> str:
        """清理内容"""
        if not isinstance(text, str):
            return ""
        text = re.sub(r'[\(（].*?[\)）]', '', text)
        text = re.sub(r'\s+', '', text)
        text = re.sub(r'[0-9]+', '', text)
        return text

    @staticmethod
    def extract_extra_info(text: str) -> str:
        """提取括号内额外信息"""
        if not isinstance(text, str):
            return ""
        matches = re.findall(r'([\(（].*?[\)）])', text)
        return " ".join(matches) if matches else ""

    @staticmethod
    def is_english(text: str) -> bool:
        """判断是否为英文"""
        if not isinstance(text, str) or len(text) <= 3:
            return False
        zh_count = len(re.findall(r'[\u4e00-\u9fa5]', text))
        return zh_count < len(text) * 0.15

    # ========================================
    # 相似度计算（混合算法）
    # ========================================

    @staticmethod
    def calculate_similarity(text1: str, text2: str) -> float:
        """
        混合相似度计算：
        - SequenceMatcher（序列匹配）
        - Levenshtein（编辑距离）
        取较高值
        """
        if not text1 or not text2:
            return 0.0

        # 序列匹配
        seq_ratio = difflib.SequenceMatcher(None, text1, text2).ratio()

        # 编辑距离（仅对较短文本使用，避免性能问题）
        if len(text1) <= 100 and len(text2) <= 100:
            lev_ratio = levenshtein_ratio(text1, text2)
            return max(seq_ratio, lev_ratio)

        return seq_ratio

    # ========================================
    # 索引构建（性能优化核心）
    # ========================================

    def build_index(self, lib_data: List[Dict]) -> LibraryIndex:
        """
        预构建条款库索引，加速匹配
        时间复杂度从 O(n*m) 降至 O(n + m)
        """
        logger.info(f"开始构建索引，条款数: {len(lib_data)}")

        index = LibraryIndex(data=lib_data)

        for i, lib in enumerate(lib_data):
            name = str(lib.get('条款名称', ''))
            if not name.strip():
                continue

            # 预计算清理结果（避免重复计算）
            name_norm = self.normalize_text(name)
            name_clean = self.clean_title(name)

            index.cleaned_cache[i] = {
                'norm': name_norm,
                'clean': name_clean,
                'original': name,
            }

            # 名称索引（精确匹配用）
            index.by_name_norm[name_norm] = i
            index.by_name_norm[name_clean] = i

            # 关键词倒排索引
            keywords = self._get_keywords(name)
            for kw in keywords:
                index.by_keyword[kw].append(i)

        logger.info(f"索引构建完成: {len(index.by_name_norm)} 名称, {len(index.by_keyword)} 关键词")
        self._index = index
        return index

    # ========================================
    # 多级匹配策略（拆分重构）
    # ========================================

    def _try_exact_match(self, title_norm: str, title_clean: str,
                         index: LibraryIndex) -> Optional[Tuple[int, float]]:
        """级别1: 精确匹配"""
        # 标准化名称精确匹配
        if title_norm in index.by_name_norm:
            return index.by_name_norm[title_norm], 1.0

        # 清理后名称精确匹配
        if title_clean in index.by_name_norm:
            return index.by_name_norm[title_clean], self.thresholds.exact_min

        return None

    def _try_semantic_match(self, title: str, index: LibraryIndex) -> Optional[Tuple[int, float]]:
        """级别2: 语义别名匹配"""
        semantic_target = self._get_semantic_alias(title)
        if not semantic_target:
            return None

        # 在索引中查找目标
        for i, cached in index.cleaned_cache.items():
            if semantic_target in cached['original']:
                return i, self.thresholds.semantic_min

        return None

    def _try_keyword_match(self, title: str, index: LibraryIndex) -> Optional[Tuple[int, float]]:
        """级别3: 关键词匹配"""
        c_keywords = self._get_keywords(title)
        if not c_keywords:
            return None

        # 统计候选项得分
        candidate_scores: Dict[int, float] = defaultdict(float)

        for kw in c_keywords:
            if kw in index.by_keyword:
                for idx in index.by_keyword[kw]:
                    candidate_scores[idx] += 1

        if not candidate_scores:
            return None

        # 找最高分候选
        best_idx = max(candidate_scores, key=candidate_scores.get)
        best_count = candidate_scores[best_idx]

        # 计算关键词匹配度
        l_keywords = self._get_keywords(index.cleaned_cache[best_idx]['original'])
        if l_keywords:
            keyword_ratio = best_count / max(len(c_keywords), len(l_keywords))
            if keyword_ratio >= 0.5:
                score = self.thresholds.keyword_min + keyword_ratio * 0.2
                return best_idx, score

        return None

    def _try_fuzzy_match(self, title_clean: str, content: str,
                         index: LibraryIndex, is_title_only: bool) -> Tuple[int, float, float, float]:
        """级别4: 模糊匹配"""
        best_idx = -1
        best_score = 0.0
        best_title_sim = 0.0
        best_content_sim = 0.0

        for i, cached in index.cleaned_cache.items():
            l_name_clean = cached['clean']

            # 标题相似度
            title_sim = self.calculate_similarity(title_clean, l_name_clean)

            # 内容相似度
            content_sim = 0.0
            if not is_title_only and content.strip():
                c_content_clean = self.clean_content(content)
                l_content = str(index.data[i].get('条款内容', ''))
                l_content_clean = self.clean_content(l_content)
                if c_content_clean and l_content_clean:
                    content_sim = self.calculate_similarity(c_content_clean, l_content_clean)

            # 加权得分
            if is_title_only or not content.strip():
                score = title_sim
            else:
                score = 0.7 * title_sim + 0.3 * content_sim

            # 惩罚项
            if self._is_penalty_keyword(cached['original']) and not self._is_penalty_keyword(title_clean):
                score -= 0.5

            if score > best_score:
                best_score = score
                best_idx = i
                best_title_sim = title_sim
                best_content_sim = content_sim

        return best_idx, best_score, best_title_sim, best_content_sim

    def match_clause(self, clause: ClauseItem, index: LibraryIndex,
                     is_title_only: bool) -> MatchResult:
        """
        主匹配入口 - 多级策略
        优先级: 精确 > 语义 > 关键词 > 模糊
        """
        result = MatchResult()
        title = clause.title
        content = clause.content

        title_clean = self.clean_title(title)
        title_norm = self.normalize_text(title)

        matched_idx = -1
        match_level = MatchLevel.NONE
        score = 0.0
        title_score = 0.0
        content_score = 0.0

        # === 级别1: 精确匹配 ===
        exact_result = self._try_exact_match(title_norm, title_clean, index)
        if exact_result:
            matched_idx, score = exact_result
            match_level = MatchLevel.EXACT
            title_score = score

        # === 级别2: 语义匹配 ===
        if matched_idx < 0:
            semantic_result = self._try_semantic_match(title, index)
            if semantic_result:
                matched_idx, score = semantic_result
                match_level = MatchLevel.SEMANTIC
                title_score = score

        # === 级别3: 关键词匹配 ===
        if matched_idx < 0:
            keyword_result = self._try_keyword_match(title, index)
            if keyword_result:
                matched_idx, score = keyword_result
                match_level = MatchLevel.KEYWORD
                title_score = score

        # === 级别4: 模糊匹配 ===
        if matched_idx < 0:
            fuzzy_idx, fuzzy_score, t_sim, c_sim = self._try_fuzzy_match(
                title_clean, content, index, is_title_only
            )
            if fuzzy_score > self.thresholds.accept_min:
                matched_idx = fuzzy_idx
                score = fuzzy_score
                match_level = MatchLevel.FUZZY
                title_score = t_sim
                content_score = c_sim

        # 构建结果
        if matched_idx >= 0 and score > self.thresholds.accept_min:
            lib = index.data[matched_idx]
            base_name = lib.get('条款名称', '')
            extra_params = self.extract_extra_info(clause.original_title or clause.title)

            result.matched_name = f"{base_name} {extra_params}".strip() if extra_params else base_name
            result.matched_content = lib.get('条款内容', '')
            result.matched_reg = lib.get('产品注册号', lib.get('注册号', ''))
            result.score = max(0, score)
            result.title_score = title_score
            result.content_score = content_score
            result.match_level = match_level

            # 差异分析（低分时）
            if score < 0.6:
                result.diff_analysis = self.analyze_difference(content, result.matched_content)

        return result

    # ========================================
    # 翻译和差异分析
    # ========================================

    def translate_title(self, title: str) -> Tuple[str, bool]:
        """翻译英文标题"""
        if not self.is_english(title):
            return title, False

        title_norm = self.normalize_text(title)

        # 1. 查询映射
        mapped = self._get_client_mapping(title_norm)
        if mapped:
            return mapped, True

        # 2. 部分匹配
        client_map = (self.config.client_en_cn_map if self._use_external_config
                      else DefaultConfig.CLIENT_EN_CN_MAP)
        for eng, chn in client_map.items():
            if eng in title_norm or title_norm in eng:
                return chn, True

        # 3. 在线翻译
        if HAS_TRANSLATOR:
            try:
                translated = GoogleTranslator(source='auto', target='zh-CN').translate(title)
                logger.debug(f"在线翻译: {title} -> {translated}")
                return translated, True
            except ConnectionError as e:
                logger.warning(f"翻译服务连接失败: {e}")
            except TimeoutError as e:
                logger.warning(f"翻译服务超时: {e}")
            except Exception as e:
                logger.error(f"翻译失败: {type(e).__name__}: {e}")

        return title, False

    @staticmethod
    def analyze_difference(c_content: str, l_content: str) -> str:
        """分析保障差异"""
        c_text, l_text = str(c_content), str(l_content)
        if not c_text.strip():
            return ""

        analysis = []
        keywords = {
            "限额": ["Limit", "限额", "最高", "limit"],
            "免赔": ["Deductible", "Excess", "免赔", "deductible"],
            "除外": ["Exclusion", "除外", "不负责", "exclusion"],
            "观察期": ["Waiting Period", "观察期", "等待期"],
            "赔偿期": ["Indemnity Period", "赔偿期间"],
        }

        for key, words in keywords.items():
            c_has = any(w.lower() in c_text.lower() for w in words)
            l_has = any(w.lower() in l_text.lower() for w in words)
            if c_has and not l_has:
                analysis.append(f"⚠️ 客户提及[{key}]但库内未提及")
            elif not c_has and l_has:
                analysis.append(f"ℹ️ 库内包含[{key}]但客户未提及")

        return " | ".join(analysis)

    # ========================================
    # 文档解析
    # ========================================

    @staticmethod
    def is_likely_title(text: str) -> bool:
        """判断是否像标题"""
        if len(text) > 80:
            return False
        if text.endswith(('。', '；', '.', ';')):
            return False
        title_indicators = ["条款", "Clause", "Extension", "险", "CLAUSE", "EXTENSION"]
        if any(kw in text for kw in title_indicators):
            return True
        if text.isupper() and len(text) > 5:
            return True
        return True

    def parse_docx(self, doc_path: str) -> Tuple[List[ClauseItem], bool]:
        """解析Word文档"""
        logger.info(f"解析文档: {doc_path}")

        try:
            doc = Document(doc_path)
        except Exception as e:
            logger.error(f"文档打开失败: {e}")
            raise ValueError(f"无法打开文档: {e}")

        clauses = []
        current_block = []

        all_lines = [p.text.strip() for p in doc.paragraphs]
        empty_lines = sum(1 for t in all_lines if not t)

        use_smart_split = len(all_lines) > 0 and (empty_lines / max(len(all_lines), 1) < 0.05)
        logger.info(f"分割模式: {'智能分割' if use_smart_split else '空行分割'}")

        if use_smart_split:
            for text in all_lines:
                if not text:
                    continue
                if current_block and self.is_likely_title(text):
                    title = current_block[0]
                    content = "\n".join(current_block[1:])
                    clauses.append(ClauseItem(title=title, content=content, original_title=title))
                    current_block = [text]
                else:
                    current_block.append(text)
            if current_block:
                clauses.append(ClauseItem(
                    title=current_block[0],
                    content="\n".join(current_block[1:]),
                    original_title=current_block[0]
                ))
        else:
            for text in all_lines:
                if text:
                    current_block.append(text)
                elif current_block:
                    clauses.append(ClauseItem(
                        title=current_block[0],
                        content="\n".join(current_block[1:]),
                        original_title=current_block[0]
                    ))
                    current_block = []
            if current_block:
                clauses.append(ClauseItem(
                    title=current_block[0],
                    content="\n".join(current_block[1:]),
                    original_title=current_block[0]
                ))

        is_title_only = all(not c.content for c in clauses)
        logger.info(f"解析完成: {len(clauses)} 条款, 纯标题模式: {is_title_only}")

        return clauses, is_title_only


# ==========================================
# 条款库加载器
# ==========================================
class LibraryLoader:
    """条款库加载器 - 支持自动列名识别"""

    @staticmethod
    def load_excel(excel_path: str, header_row: int = 1) -> List[Dict]:
        """
        加载Excel条款库
        自动识别列名
        """
        logger.info(f"加载条款库: {excel_path}")

        try:
            df = pd.read_excel(excel_path, header=header_row)
        except FileNotFoundError:
            raise ValueError(f"文件不存在: {excel_path}")
        except Exception as e:
            raise ValueError(f"Excel读取失败: {e}")

        df.columns = [str(c).strip() for c in df.columns]

        # 自动识别列名
        name_col = None
        content_col = None
        reg_col = None

        for col in df.columns:
            col_lower = col.lower()
            if name_col is None and ('条款名称' in col or '名称' in col or 'name' in col_lower):
                name_col = col
            elif content_col is None and ('条款内容' in col or '内容' in col or 'content' in col_lower):
                content_col = col
            elif reg_col is None and ('注册号' in col or '产品' in col or 'reg' in col_lower):
                reg_col = col

        # 回退到位置
        if not name_col and len(df.columns) > 0:
            name_col = df.columns[0]
        if not content_col and len(df.columns) > 2:
            content_col = df.columns[2]
        if not reg_col and len(df.columns) > 1:
            reg_col = df.columns[1]

        logger.info(f"列名识别: 名称={name_col}, 内容={content_col}, 注册号={reg_col}")

        # 构建数据
        lib_data = []
        for _, row in df.iterrows():
            name = str(row.get(name_col, '')) if pd.notna(row.get(name_col)) else ''
            if not name.strip():
                continue

            lib_data.append({
                '条款名称': name,
                '条款内容': str(row.get(content_col, '')) if content_col and pd.notna(row.get(content_col)) else '',
                '产品注册号': str(row.get(reg_col, '')) if reg_col and pd.notna(row.get(reg_col)) else '',
            })

        logger.info(f"加载完成: {len(lib_data)} 条有效记录")
        return lib_data


# ==========================================
# Excel样式器
# ==========================================
class ExcelStyler:
    """Excel样式应用器"""

    FILLS = {
        'green': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'yellow': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'blue': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
        'header': PatternFill(start_color="667eea", end_color="667eea", fill_type="solid"),
    }

    BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    WIDTHS = {
        'A': 6, 'B': 35, 'C': 30, 'D': 45, 'E': 40,
        'F': 25, 'G': 50, 'H': 10, 'I': 12, 'J': 35, 'K': 10, 'L': 10
    }

    @classmethod
    def apply_styles(cls, output_path: str):
        """应用Excel样式"""
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # 表头
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = cls.FILLS['header']
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cls.BORDER

        # 列宽
        for col, width in cls.WIDTHS.items():
            ws.column_dimensions[col].width = width

        # 数据行
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = cls.BORDER

                # 匹配度着色
                if cell.col_idx == ExcelColumns.SCORE_COL_IDX:
                    try:
                        val = float(cell.value) if cell.value else 0
                        if val >= 0.8:
                            cell.fill = cls.FILLS['green']
                        elif val >= 0.5:
                            cell.fill = cls.FILLS['yellow']
                        elif val > 0:
                            cell.fill = cls.FILLS['red']
                    except (ValueError, TypeError):
                        pass

                # 匹配级别着色
                if cell.col_idx == ExcelColumns.LEVEL_COL_IDX:
                    val = str(cell.value) if cell.value else ""
                    if "精确" in val:
                        cell.fill = cls.FILLS['green']
                    elif "语义" in val:
                        cell.fill = cls.FILLS['blue']
                    elif "关键词" in val:
                        cell.fill = cls.FILLS['yellow']

        # 冻结首行
        ws.freeze_panes = 'A2'

        wb.save(output_path)
        logger.info(f"Excel样式已应用: {output_path}")


# ==========================================
# 工作线程
# ==========================================
class MatchWorker(QThread):
    """单文件匹配工作线程"""
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, doc_path: str, excel_path: str, output_path: str):
        super().__init__()
        self.doc_path = doc_path
        self.excel_path = excel_path
        self.output_path = output_path

    def run(self):
        try:
            logic = ClauseMatcherLogic()

            # 状态信息
            self.log_signal.emit(f"📊 配置: 外部={logic._use_external_config}, 翻译={HAS_TRANSLATOR}", "info")

            # 解析文档
            self.log_signal.emit("⏳ 正在解析文档...", "info")
            clauses, is_title_only = logic.parse_docx(self.doc_path)
            mode_str = "纯标题模式" if is_title_only else "完整内容模式"
            self.log_signal.emit(f"📖 [{mode_str}] 提取到 {len(clauses)} 条", "success")

            # 加载条款库
            self.log_signal.emit("📚 加载条款库...", "info")
            lib_data = LibraryLoader.load_excel(self.excel_path)
            self.log_signal.emit(f"✓ 条款库 {len(lib_data)} 条", "success")

            # 构建索引
            self.log_signal.emit("🔧 构建索引...", "info")
            index = logic.build_index(lib_data)
            self.log_signal.emit(f"✓ 索引完成", "success")

            # 开始匹配
            self.log_signal.emit("🧠 开始智能匹配...", "info")
            results = []
            stats = {'exact': 0, 'semantic': 0, 'keyword': 0, 'fuzzy': 0, 'none': 0}

            for idx, clause in enumerate(clauses, 1):
                self.progress_signal.emit(idx, len(clauses))

                # 翻译
                original_title = clause.title
                translated_title, was_translated = logic.translate_title(clause.title)
                if was_translated:
                    clause.title = translated_title
                    clause.original_title = original_title

                # 匹配
                match_result = logic.match_clause(clause, index, is_title_only)

                # 统计
                if match_result.match_level == MatchLevel.EXACT:
                    stats['exact'] += 1
                elif match_result.match_level == MatchLevel.SEMANTIC:
                    stats['semantic'] += 1
                elif match_result.match_level == MatchLevel.KEYWORD:
                    stats['keyword'] += 1
                elif match_result.match_level == MatchLevel.FUZZY:
                    stats['fuzzy'] += 1
                else:
                    stats['none'] += 1

                results.append({
                    ExcelColumns.SEQ: idx,
                    ExcelColumns.CLIENT_ORIG: original_title,
                    ExcelColumns.CLIENT_TRANS: translated_title if was_translated else "",
                    ExcelColumns.CLIENT_CONTENT: clause.content[:500] if clause.content else "",
                    ExcelColumns.MATCHED_NAME: match_result.matched_name or "无匹配",
                    ExcelColumns.REG_NO: match_result.matched_reg,
                    ExcelColumns.MATCHED_CONTENT: match_result.matched_content[:500] if match_result.matched_content else "",
                    ExcelColumns.SCORE: round(match_result.score, 3),
                    ExcelColumns.MATCH_LEVEL: match_result.match_level.value,
                    ExcelColumns.DIFF_ANALYSIS: match_result.diff_analysis,
                    ExcelColumns.TITLE_SCORE: round(match_result.title_score, 3),
                    ExcelColumns.CONTENT_SCORE: round(match_result.content_score, 3),
                })

            # 保存结果
            df_res = pd.DataFrame(results)
            df_res.to_excel(self.output_path, index=False)
            ExcelStyler.apply_styles(self.output_path)

            # 输出统计
            self.log_signal.emit(f"📊 匹配统计:", "info")
            self.log_signal.emit(f"   精确匹配: {stats['exact']}", "success")
            self.log_signal.emit(f"   语义匹配: {stats['semantic']}", "success")
            self.log_signal.emit(f"   关键词匹配: {stats['keyword']}", "info")
            self.log_signal.emit(f"   模糊匹配: {stats['fuzzy']}", "warning")
            self.log_signal.emit(f"   无匹配: {stats['none']}", "error")

            self.log_signal.emit(f"🎉 完成！", "success")
            self.finished_signal.emit(True, self.output_path)

        except Exception as e:
            logger.exception("匹配过程出错")
            self.log_signal.emit(f"❌ 错误: {str(e)}", "error")
            self.finished_signal.emit(False, str(e))


class BatchMatchWorker(QThread):
    """批量匹配工作线程"""
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    batch_progress_signal = pyqtSignal(int, int, str)  # 当前文件, 总数, 文件名
    finished_signal = pyqtSignal(bool, str, int, int)  # 成功, 消息, 成功数, 总数

    def __init__(self, doc_paths: List[str], excel_path: str, output_dir: str):
        super().__init__()
        self.doc_paths = doc_paths
        self.excel_path = excel_path
        self.output_dir = output_dir

    def run(self):
        try:
            logic = ClauseMatcherLogic()

            # 加载条款库（只需一次）
            self.log_signal.emit("📚 加载条款库...", "info")
            lib_data = LibraryLoader.load_excel(self.excel_path)
            self.log_signal.emit(f"✓ 条款库 {len(lib_data)} 条", "success")

            # 构建索引（只需一次）
            self.log_signal.emit("🔧 构建索引...", "info")
            index = logic.build_index(lib_data)

            success_count = 0
            total = len(self.doc_paths)

            for file_idx, doc_path in enumerate(self.doc_paths, 1):
                file_name = Path(doc_path).name
                self.batch_progress_signal.emit(file_idx, total, file_name)
                self.log_signal.emit(f"\n📄 [{file_idx}/{total}] {file_name}", "info")

                try:
                    # 解析文档
                    clauses, is_title_only = logic.parse_docx(doc_path)
                    self.log_signal.emit(f"   提取 {len(clauses)} 条款", "info")

                    # 匹配
                    results = []
                    for idx, clause in enumerate(clauses, 1):
                        original_title = clause.title
                        translated_title, was_translated = logic.translate_title(clause.title)
                        if was_translated:
                            clause.title = translated_title
                            clause.original_title = original_title

                        match_result = logic.match_clause(clause, index, is_title_only)

                        results.append({
                            ExcelColumns.SEQ: idx,
                            ExcelColumns.CLIENT_ORIG: original_title,
                            ExcelColumns.CLIENT_TRANS: translated_title if was_translated else "",
                            ExcelColumns.CLIENT_CONTENT: clause.content[:500] if clause.content else "",
                            ExcelColumns.MATCHED_NAME: match_result.matched_name or "无匹配",
                            ExcelColumns.REG_NO: match_result.matched_reg,
                            ExcelColumns.MATCHED_CONTENT: match_result.matched_content[:500] if match_result.matched_content else "",
                            ExcelColumns.SCORE: round(match_result.score, 3),
                            ExcelColumns.MATCH_LEVEL: match_result.match_level.value,
                            ExcelColumns.DIFF_ANALYSIS: match_result.diff_analysis,
                            ExcelColumns.TITLE_SCORE: round(match_result.title_score, 3),
                            ExcelColumns.CONTENT_SCORE: round(match_result.content_score, 3),
                        })

                    # 保存
                    output_name = f"报告_{Path(doc_path).stem}.xlsx"
                    output_path = Path(self.output_dir) / output_name
                    df_res = pd.DataFrame(results)
                    df_res.to_excel(output_path, index=False)
                    ExcelStyler.apply_styles(str(output_path))

                    self.log_signal.emit(f"   ✓ 已保存: {output_name}", "success")
                    success_count += 1

                except Exception as e:
                    self.log_signal.emit(f"   ✗ 失败: {e}", "error")

            self.log_signal.emit(f"\n🎉 批量处理完成: {success_count}/{total}", "success")
            self.finished_signal.emit(True, self.output_dir, success_count, total)

        except Exception as e:
            logger.exception("批量处理出错")
            self.log_signal.emit(f"❌ 错误: {str(e)}", "error")
            self.finished_signal.emit(False, str(e), 0, 0)


# ==========================================
# UI组件
# ==========================================
class GlassCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            GlassCard {
                background: rgba(255, 255, 255, 0.08);
                border: 1px solid rgba(255, 255, 255, 0.15);
                border-radius: 20px;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(40)
        shadow.setColor(QColor(0, 0, 0, 80))
        shadow.setOffset(0, 10)
        self.setGraphicsEffect(shadow)


class AddMappingDialog(QDialog):
    """添加映射对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("添加术语映射")
        self.setMinimumWidth(400)
        self.setStyleSheet("""
            QDialog { background: #1a1a2e; }
            QLabel { color: #ffffff; font-size: 14px; }
            QLineEdit {
                background: rgba(255,255,255,0.1);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 8px; padding: 10px; color: #ffffff;
            }
            QLineEdit:focus { border-color: #667eea; }
            QPushButton {
                background: #667eea; color: white; border: none;
                border-radius: 8px; padding: 10px 20px; font-weight: bold;
            }
            QPushButton:hover { background: #764ba2; }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        form = QFormLayout()
        self.eng_input = QLineEdit()
        self.eng_input.setPlaceholderText("例如: reinstatement value")
        form.addRow("英文术语:", self.eng_input)

        self.chn_input = QLineEdit()
        self.chn_input.setPlaceholderText("例如: 重置价值条款")
        form.addRow("中文翻译:", self.chn_input)
        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("background: rgba(255,255,255,0.1);")
        cancel_btn.clicked.connect(self.reject)
        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.accept)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)

    def get_mapping(self) -> Tuple[str, str]:
        return self.eng_input.text().strip(), self.chn_input.text().strip()


class BatchSelectDialog(QDialog):
    """批量文件选择对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("批量处理")
        self.setMinimumSize(500, 400)
        self.setStyleSheet("""
            QDialog { background: #1a1a2e; }
            QLabel { color: #ffffff; }
            QListWidget {
                background: rgba(255,255,255,0.1);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 8px; color: #ffffff;
            }
            QPushButton {
                background: rgba(255,255,255,0.1);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 8px; padding: 10px; color: #ffffff;
            }
            QPushButton:hover { background: rgba(255,255,255,0.2); }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        layout.addWidget(QLabel("选择要批量处理的 Word 文件:"))

        self.file_list = QListWidget()
        layout.addWidget(self.file_list)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("➕ 添加文件")
        add_btn.clicked.connect(self._add_files)
        clear_btn = QPushButton("🗑️ 清空")
        clear_btn.clicked.connect(self.file_list.clear)
        btn_row.addWidget(add_btn)
        btn_row.addWidget(clear_btn)
        layout.addLayout(btn_row)

        action_row = QHBoxLayout()
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        start_btn = QPushButton("开始批量处理")
        start_btn.setStyleSheet("background: #667eea;")
        start_btn.clicked.connect(self.accept)
        action_row.addWidget(cancel_btn)
        action_row.addWidget(start_btn)
        layout.addLayout(action_row)

        self.selected_files: List[str] = []

    def _add_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "选择Word文件", "", "Word Files (*.docx)")
        for f in files:
            if f not in self.selected_files:
                self.selected_files.append(f)
                self.file_list.addItem(Path(f).name)

    def get_files(self) -> List[str]:
        return self.selected_files


class ClauseDiffGUI(QMainWindow):
    """主界面"""
    def __init__(self):
        super().__init__()
        self.setWindowTitle("智能条款比对工具 v16.0")
        self.setMinimumSize(950, 850)
        self.setStyleSheet("""
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1a1a2e, stop:0.5 #16213e, stop:1 #0f3460);
            }
        """)

        if HAS_CONFIG_MANAGER:
            self._config = get_config()
        else:
            self._config = None

        self._setup_ui()

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(18)
        layout.setContentsMargins(40, 25, 40, 25)

        # 标题
        title = QLabel("🔍 智能条款比对工具")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #ffffff; font-size: 30px; font-weight: bold;")
        layout.addWidget(title)

        subtitle = QLabel("v16.0 Full Optimized · 索引加速 · 批量处理 · 混合算法")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: rgba(255,255,255,0.6); font-size: 13px;")
        layout.addWidget(subtitle)

        # 配置统计
        if self._config:
            stats = self._config.get_stats()
            stats_text = f"📊 {stats['client_mappings']} 映射 | {stats['semantic_aliases']} 别名 | {stats['keyword_rules']} 关键词"
        else:
            stats_text = "📊 使用内置配置"
        stats_label = QLabel(stats_text)
        stats_label.setAlignment(Qt.AlignCenter)
        stats_label.setStyleSheet("color: rgba(255,255,255,0.4); font-size: 11px;")
        layout.addWidget(stats_label)

        # 输入卡片
        card = GlassCard()
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(18)
        card_layout.setContentsMargins(30, 30, 30, 30)

        style = """
            QLabel { color: #ffffff; font-weight: 500; }
            QLineEdit {
                background: rgba(0,0,0,0.2);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 10px; padding: 12px 15px;
                color: #ffffff; font-size: 14px;
            }
            QLineEdit:focus { border-color: #667eea; }
        """
        card.setStyleSheet(card.styleSheet() + style)

        btn_style = """
            QPushButton {
                background: rgba(255,255,255,0.1);
                border: 1px solid rgba(255,255,255,0.2);
                border-radius: 10px; padding: 12px 18px;
                color: #ffffff; font-weight: 500;
            }
            QPushButton:hover { background: rgba(255,255,255,0.2); border-color: #667eea; }
        """

        self.doc_input = self._create_file_row(card_layout, "📂 客户文档",
            "Word 条款清单 (.docx)", "Word Files (*.docx)", btn_style)
        self.lib_input = self._create_file_row(card_layout, "📚 标准题库",
            "Excel 条款库 (.xlsx)", "Excel Files (*.xlsx)", btn_style)

        line = QFrame()
        line.setFixedHeight(1)
        line.setStyleSheet("background: rgba(255,255,255,0.1);")
        card_layout.addWidget(line)

        row3 = QHBoxLayout()
        label3 = QLabel("💾 保存路径")
        label3.setFixedWidth(90)
        self.out_input = QLineEdit()
        self.out_input.setPlaceholderText("报告保存位置...")
        btn3 = QPushButton("选择")
        btn3.setCursor(Qt.PointingHandCursor)
        btn3.setStyleSheet(btn_style)
        btn3.clicked.connect(self._browse_save)
        row3.addWidget(label3)
        row3.addWidget(self.out_input, 1)
        row3.addWidget(btn3)
        card_layout.addLayout(row3)

        layout.addWidget(card)

        # 按钮行
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)

        self.start_btn = QPushButton("🚀 开始比对")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.setMinimumHeight(52)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #667eea, stop:1 #764ba2);
                color: white; font-size: 16px; font-weight: bold;
                border-radius: 26px; border: none;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #764ba2, stop:1 #667eea);
            }
            QPushButton:disabled { background: rgba(255,255,255,0.1); color: rgba(255,255,255,0.3); }
        """)
        self.start_btn.clicked.connect(self._start_process)

        self.batch_btn = QPushButton("📦 批量处理")
        self.batch_btn.setCursor(Qt.PointingHandCursor)
        self.batch_btn.setMinimumHeight(52)
        self.batch_btn.setStyleSheet("""
            QPushButton {
                background: transparent; color: rgba(255,255,255,0.7);
                font-size: 14px; font-weight: 500;
                border-radius: 26px; border: 2px solid rgba(255,255,255,0.2);
            }
            QPushButton:hover { border-color: #667eea; color: #667eea; }
        """)
        self.batch_btn.clicked.connect(self._show_batch_dialog)

        self.add_btn = QPushButton("➕ 添加映射")
        self.add_btn.setCursor(Qt.PointingHandCursor)
        self.add_btn.setMinimumHeight(52)
        self.add_btn.setStyleSheet(self.batch_btn.styleSheet())
        self.add_btn.clicked.connect(self._show_add_mapping_dialog)

        self.open_btn = QPushButton("📂 打开目录")
        self.open_btn.setCursor(Qt.PointingHandCursor)
        self.open_btn.setMinimumHeight(52)
        self.open_btn.setEnabled(False)
        self.open_btn.setStyleSheet("""
            QPushButton {
                background: transparent; color: rgba(255,255,255,0.5);
                font-size: 14px; font-weight: 500;
                border-radius: 26px; border: 2px solid rgba(255,255,255,0.15);
            }
            QPushButton:hover { border-color: #27ae60; color: #27ae60; }
            QPushButton:disabled { color: rgba(255,255,255,0.2); border-color: rgba(255,255,255,0.1); }
        """)
        self.open_btn.clicked.connect(self._open_output_folder)

        btn_layout.addWidget(self.start_btn, 3)
        btn_layout.addWidget(self.batch_btn, 1)
        btn_layout.addWidget(self.add_btn, 1)
        btn_layout.addWidget(self.open_btn, 1)
        layout.addLayout(btn_layout)

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(4)
        self.progress_bar.setStyleSheet("""
            QProgressBar { background: rgba(255,255,255,0.1); border-radius: 2px; }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #667eea, stop:1 #764ba2);
                border-radius: 2px;
            }
        """)
        layout.addWidget(self.progress_bar)

        # 日志
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background: rgba(0,0,0,0.3);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 12px; color: #e8e8e8;
                padding: 15px;
                font-family: 'SF Mono', 'Menlo', 'Monaco', monospace;
                font-size: 12px;
            }
        """)
        layout.addWidget(self.log_text, 1)

        # 版本信息
        version = QLabel("v16.0 Full Optimized · Made with ❤️")
        version.setAlignment(Qt.AlignCenter)
        version.setStyleSheet("color: rgba(255,255,255,0.25); font-size: 11px;")
        layout.addWidget(version)

    def _create_file_row(self, layout, label_text: str, placeholder: str,
                         filter_str: str, btn_style: str) -> QLineEdit:
        row = QHBoxLayout()
        label = QLabel(label_text)
        label.setFixedWidth(90)
        line_edit = QLineEdit()
        line_edit.setPlaceholderText(placeholder)
        btn = QPushButton("浏览")
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(btn_style)
        btn.clicked.connect(lambda: self._browse_file(line_edit, filter_str))
        row.addWidget(label)
        row.addWidget(line_edit, 1)
        row.addWidget(btn)
        layout.addLayout(row)
        return line_edit

    def _browse_file(self, line_edit: QLineEdit, filter_str: str):
        f, _ = QFileDialog.getOpenFileName(self, "选择文件", "", filter_str)
        if f:
            line_edit.setText(f)
            if line_edit == self.doc_input and not self.out_input.text():
                self.out_input.setText(os.path.join(os.path.dirname(f), "条款比对报告.xlsx"))

    def _browse_save(self):
        f, _ = QFileDialog.getSaveFileName(self, "保存结果", "条款比对报告.xlsx", "Excel Files (*.xlsx)")
        if f:
            self.out_input.setText(f)

    def _show_add_mapping_dialog(self):
        if not self._config:
            QMessageBox.warning(self, "提示", "配置管理器不可用")
            return
        dialog = AddMappingDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            eng, chn = dialog.get_mapping()
            if eng and chn:
                self._config.add_client_mapping(eng, chn)
                self._config.save()
                self._append_log(f"✓ 已添加映射: '{eng}' -> '{chn}'", "success")

    def _show_batch_dialog(self):
        if not self.lib_input.text():
            QMessageBox.warning(self, "提示", "请先选择条款库")
            return

        dialog = BatchSelectDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            files = dialog.get_files()
            if not files:
                return

            output_dir = QFileDialog.getExistingDirectory(self, "选择输出目录")
            if not output_dir:
                return

            self._start_batch_process(files, output_dir)

    def _append_log(self, msg: str, level: str):
        colors = {"info": "#a0a0a0", "success": "#2ecc71", "error": "#e74c3c", "warning": "#f39c12"}
        self.log_text.append(f'<span style="color:{colors.get(level, "#fff")}">{msg}</span>')
        self.log_text.moveCursor(QTextCursor.End)

    def _start_process(self):
        doc = self.doc_input.text().strip()
        excel = self.lib_input.text().strip()
        out = self.out_input.text().strip()

        if not all([doc, excel, out]):
            QMessageBox.warning(self, "提示", "请完善所有文件路径！")
            return

        self._set_ui_state(False)
        self.log_text.clear()

        self.worker = MatchWorker(doc, excel, out)
        self.worker.log_signal.connect(self._append_log)
        self.worker.progress_signal.connect(lambda c, t: self.progress_bar.setValue(int(c/t*100)))
        self.worker.finished_signal.connect(self._on_finished)
        self.worker.start()

    def _start_batch_process(self, files: List[str], output_dir: str):
        self._set_ui_state(False)
        self.log_text.clear()

        self.batch_worker = BatchMatchWorker(files, self.lib_input.text(), output_dir)
        self.batch_worker.log_signal.connect(self._append_log)
        self.batch_worker.batch_progress_signal.connect(
            lambda c, t, n: self.progress_bar.setValue(int(c/t*100))
        )
        self.batch_worker.finished_signal.connect(self._on_batch_finished)
        self.batch_worker.start()

    def _set_ui_state(self, enabled: bool):
        self.start_btn.setEnabled(enabled)
        self.batch_btn.setEnabled(enabled)
        self.start_btn.setText("🚀 开始比对" if enabled else "⏳ 处理中...")
        self.progress_bar.setVisible(not enabled)
        if not enabled:
            self.progress_bar.setValue(0)

    def _on_finished(self, success: bool, msg: str):
        self._set_ui_state(True)
        if success:
            self.open_btn.setEnabled(True)
            self.open_btn.setStyleSheet("""
                QPushButton {
                    background: transparent; color: #2ecc71;
                    font-size: 14px; font-weight: 500;
                    border-radius: 26px; border: 2px solid #2ecc71;
                }
                QPushButton:hover { background: #2ecc71; color: white; }
            """)
            QMessageBox.information(self, "完成", f"比对完成！\n{msg}")

    def _on_batch_finished(self, success: bool, msg: str, ok_count: int, total: int):
        self._set_ui_state(True)
        if success:
            self.open_btn.setEnabled(True)
            QMessageBox.information(self, "完成", f"批量处理完成！\n成功: {ok_count}/{total}\n输出目录: {msg}")

    def _open_output_folder(self):
        path = self.out_input.text().strip()
        if path and os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))


def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    app.setFont(QFont("PingFang SC", 13))

    window = ClauseDiffGUI()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()

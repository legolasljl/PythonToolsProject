# -*- coding: utf-8 -*-
"""
Word 条款批量提取工具 v7.1 (Vertical Optimized)
- [UI优化] 纵向模式：加宽列宽 + 自动计算行高(解决显示不全) + 移除空行
- [核心] 增量提取 + 智能分表 + 严格筛选(附加/非费率)
- [对齐] 全局左对齐，标题行加灰底

Author: Google Senior Architect
Date: 2025-12-09
"""

import sys
import os
import re
import subprocess
import tempfile
import platform
import traceback
from pathlib import Path
from collections import defaultdict
import math

# 第三方库
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from docx import Document

# PyQt5 库
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QRadioButton, QProgressBar, 
    QTextEdit, QFileDialog, QMessageBox, QStyleFactory, QFrame,
    QGraphicsDropShadowEffect, QCheckBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt5.QtGui import QFont, QPalette, QColor, QTextCursor, QDesktopServices

# ==========================================
# 🛡️ 基础补丁
# ==========================================
class NullWriter:
    def write(self, text): pass
    def flush(self): pass

if getattr(sys, 'frozen', False):
    sys.stdout = NullWriter()
    sys.stderr = NullWriter()

def global_exception_handler(exctype, value, tb):
    error_msg = "".join(traceback.format_exception(exctype, value, tb))
    try: sys.__stderr__.write(error_msg)
    except: pass
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Critical)
    msg_box.setText("程序发生意外错误")
    msg_box.setInformativeText(str(value))
    msg_box.setDetailedText(error_msg)
    msg_box.setWindowTitle("错误报告")
    msg_box.exec_()

sys.excepthook = global_exception_handler

# --------------------------
# 跨平台转换
# --------------------------
class PlatformHandler:
    @staticmethod
    def is_windows():
        return platform.system() == "Windows"

    @staticmethod
    def convert_doc_to_docx(doc_path: Path) -> Path:
        temp_dir = tempfile.gettempdir()
        temp_docx_name = f"ext_opt_{doc_path.stem}.docx"
        temp_docx_path = Path(temp_dir) / temp_docx_name

        if temp_docx_path.exists():
            try: os.remove(temp_docx_path)
            except: pass

        if PlatformHandler.is_windows():
            try:
                import win32com.client
                import pythoncom
                pythoncom.CoInitialize()
                word = win32com.client.Dispatch("Word.Application")
                word.Visible = False
                word.DisplayAlerts = False
                try:
                    doc = word.Documents.Open(str(doc_path))
                    doc.SaveAs2(str(temp_docx_path), FileFormat=16)
                    doc.Close()
                except: pass
                finally:
                    try: word.Quit()
                    except: pass
                return temp_docx_path
            except: return None
        else:
            cmd = ['textutil', '-convert', 'docx', str(doc_path), '-output', str(temp_docx_path)]
            try:
                subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                return temp_docx_path
            except: return None

# --------------------------
# 核心提取逻辑
# --------------------------
class WordExtractorProcessor:
    NOISE_PATTERNS = [
        r'PAGE\s+', r'NUMPAGES', r'MERGEFORMAT',
        r'第\s*.*页\s*共\s*.*页', 
        r'-\s*\d+\s*-', 
        r'^\d+$',
        r'.*保险股份有限公司.*', r'.*保险有限公司.*'
    ]

    def is_noise_line(self, text):
        if not text.strip(): return True
        for pattern in self.NOISE_PATTERNS:
            if re.search(pattern, text, re.IGNORECASE): return True
        return False

    def extract_clause_info(self, doc_path: Path) -> dict:
        file_name = doc_path.name
        clause_name = doc_path.stem
        
        sheet_category = "其他附加条款"
        if "附加" in file_name:
            parts = file_name.split("附加")
            prefix = parts[0].strip()
            if prefix: sheet_category = f"{prefix}附加条款"
            else: sheet_category = "通用附加条款"
        
        result = {
            'FileName': file_name, 
            'ClauseName': clause_name, 
            'RegistrationNo': '', 
            'Content': '', 
            'Category': sheet_category, 
            'Error': ''
        }
        
        temp_file_to_remove = None

        try:
            target_path = doc_path
            if doc_path.suffix.lower() == '.doc':
                converted = PlatformHandler.convert_doc_to_docx(doc_path)
                if converted and os.path.exists(converted):
                    target_path = converted
                    temp_file_to_remove = converted
                else:
                    result['Error'] = "doc格式转换失败"
                    return result

            doc = Document(str(target_path))
            paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

            if not paragraphs:
                result['Error'] = '文档内容为空'
                return result

            # 注册号
            registration_no = ""
            scan_range = min(len(paragraphs), 8)
            content_start_index = 0
            for i in range(scan_range):
                line = paragraphs[i]
                if "注册号" in line or re.search(r'[A-Z]\d{10,}', line):
                    match = re.search(r'[（\(]([^）\)]+)[）\)]', line)
                    if match: registration_no = match.group(1)
                    else: registration_no = line.replace('产品注册号:', '').replace('注册号:', '').strip()
                    content_start_index = i + 1
                    break
            result['RegistrationNo'] = registration_no

            # 正文
            content_lines = []
            start_idx = 3 if content_start_index == 0 and len(paragraphs) >= 4 else content_start_index
            for para in paragraphs[start_idx:]:
                clean = para.replace('**', '').strip()
                if self.is_noise_line(clean): continue
                if clean == clause_name: continue
                if clean: content_lines.append(clean)
            
            result['Content'] = '\n'.join(content_lines)
            return result

        except Exception as e:
            result['Error'] = f"解析出错: {str(e)}"
            return result
        finally:
            if temp_file_to_remove and temp_file_to_remove.exists():
                try: os.remove(temp_file_to_remove)
                except: pass

    def get_word_files(self, directory: str) -> list:
        path_obj = Path(directory)
        word_files = []
        for ext in ['*.docx', '*.doc']:
            word_files.extend(path_obj.rglob(ext))
        
        filtered = []
        for f in word_files:
            fname = f.name
            if not fname.startswith('~') and '费率' not in fname and '附加' in fname:
                filtered.append(f)
        return filtered

    def calculate_row_height(self, content, col_width_chars=90):
        """
        [新增] 估算合并单元格所需的行高
        Excel 默认无法自动调整合并单元格的行高，必须手动计算。
        col_width_chars: 合并后的大致字符宽度 (A列60 + B列40 ≈ 100字符空间, 保守取90)
        """
        if not content: return 15
        
        lines = content.split('\n')
        total_rows = 0
        
        for line in lines:
            # 估算当前行需要折几行
            # 中文占2宽，英文占1宽，这里简单按长度估算，适当宽松
            # 假设一行能容纳 50 个中文字符左右
            line_len = len(line)
            if line_len == 0:
                total_rows += 1
            else:
                rows_needed = math.ceil(line_len / 50) 
                total_rows += rows_needed
        
        # 基础行高15，每增加一行增加15
        height = total_rows * 15
        return max(height, 30) # 最小30

    def save_to_excel(self, data_list: list, output_file: str, format_type: str = 'horizontal'):
        wb = openpyxl.Workbook()
        wb.properties.creator = "Alex Jin"
        wb.properties.lastModifiedBy = "Alex Jin"
        wb.remove(wb.active)

        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, size=12, color="FFFFFF")
        fill_header = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid") # 蓝色表头
        fill_title = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid") # 灰色标题行

        grouped_data = defaultdict(list)
        for item in data_list:
            cat = item.get('Category', '其他附加条款')
            grouped_data[cat].append(item)

        for sheet_name, items in grouped_data.items():
            safe_name = sheet_name[:30].replace('/',' ').replace('\\',' ')
            ws = wb.create_sheet(title=safe_name)

            if format_type == 'horizontal':
                # 横向模式保持不变
                headers = ['条款名称', '注册号', '条款内容', '原文件名', '状态']
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 80
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 15

                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font = header_font
                    c.fill = fill_header
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = border

                for r, item in enumerate(items, 2):
                    row_vals = [item['ClauseName'], item['RegistrationNo'], item['Content'], item['FileName'], item['Error'] or "成功"]
                    for c, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=r, column=c, value=val)
                        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                        cell.border = border
                        if c==5 and val != "成功": cell.font = Font(color="FF0000")

            else: # Vertical Mode (重点优化部分)
                # 1. 设置列宽 (加宽)
                ws.column_dimensions['A'].width = 60 # 条款名称列宽
                ws.column_dimensions['B'].width = 40 # 注册号列宽
                
                current_row = 1
                for item in items:
                    if item.get('Error'): continue
                    
                    # --- 单数行：标题 + 代码 ---
                    # 格式：左对齐，灰色背景，加粗
                    cell_name = ws.cell(row=current_row, column=1, value=item.get('ClauseName', ''))
                    cell_name.font = Font(bold=True, size=12)
                    cell_name.fill = fill_title
                    cell_name.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell_name.border = border
                    
                    cell_reg = ws.cell(row=current_row, column=2, value=item.get('RegistrationNo', ''))
                    cell_reg.font = Font(bold=True, size=12)
                    cell_reg.fill = fill_title
                    cell_reg.alignment = Alignment(horizontal='left', vertical='center')
                    cell_reg.border = border
                    
                    current_row += 1
                    
                    # --- 双数行：条款内容 (合并单元格) ---
                    # 格式：左对齐，自动换行，无背景，自动计算行高
                    content_text = item.get('Content', '')
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                    cell_content = ws.cell(row=current_row, column=1, value=content_text)
                    cell_content.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell_content.border = border
                    
                    # 自动设置行高
                    row_height = self.calculate_row_height(content_text)
                    ws.row_dimensions[current_row].height = row_height
                    
                    # 紧接着下一条，不留空行
                    current_row += 1 

        if not wb.sheetnames: wb.create_sheet("无新增数据")
        wb.save(output_file)

# --------------------------
# 工作线程 (增量逻辑)
# --------------------------
class ExtractWorker(QThread):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    finished_signal = pyqtSignal(bool, str, int, int)
    
    def __init__(self, word_folder, excel_path, history_path, format_type):
        super().__init__()
        self.word_folder = word_folder
        self.excel_path = excel_path
        self.history_path = history_path
        self.format_type = format_type
        self.processor = WordExtractorProcessor()
    
    def load_history(self):
        processed_files = set()
        if not self.history_path or not os.path.exists(self.history_path):
            return processed_files
        try:
            self.log_signal.emit(f"📖 正在读取历史记录...", "info")
            xls = pd.read_excel(self.history_path, sheet_name=None, dtype=str)
            count = 0
            for _, df in xls.items():
                if '原文件名' in df.columns:
                    files = df['原文件名'].dropna().astype(str).tolist()
                    processed_files.update(files)
                    count += len(files)
            self.log_signal.emit(f"✅ 历史记录包含 {count} 个文件", "success")
        except Exception as e:
            self.log_signal.emit(f"⚠️ 历史记录读取失败: {str(e)}", "warning")
        return processed_files

    def run(self):
        try:
            self.log_signal.emit("⏳ 初始化...", "info")
            all_files = self.processor.get_word_files(self.word_folder)
            if not all_files:
                self.finished_signal.emit(False, "未找到符合条件的文件", 0, 0)
                return

            processed_set = self.load_history()
            target_files = [f for f in all_files if f.name not in processed_set]
            
            skipped = len(all_files) - len(target_files)
            if skipped > 0: self.log_signal.emit(f"⏭️ 跳过 {skipped} 个重复文件", "warning")
            
            if not target_files:
                self.log_signal.emit("🎉 无需更新", "success")
                self.finished_signal.emit(True, "无需生成", 0, 0)
                return

            self.log_signal.emit(f"🚀 开始提取 {len(target_files)} 个新增文件", "info")
            
            processed_data = []
            success_count = 0
            
            for i, file_path in enumerate(target_files, 1):
                self.progress_signal.emit(i, len(target_files))
                cat_name = file_path.name.split("附加")[0] if "附加" in file_path.name else "其他"
                self.log_signal.emit(f"[{i}] [{cat_name}] {file_path.name}", "info")
                
                data = self.processor.extract_clause_info(file_path)
                processed_data.append(data)
                
                if not data['Error']: success_count += 1
                else: self.log_signal.emit(f"   ✗ 失败: {data['Error']}", "error")
            
            self.log_signal.emit("💾 生成 Excel...", "info")
            self.processor.save_to_excel(processed_data, self.excel_path, self.format_type)
            
            self.log_signal.emit(f"🎉 完成！新增 {success_count} 条", "success")
            self.finished_signal.emit(True, self.excel_path, success_count, len(target_files))
            
        except Exception as e:
            raise e

# --------------------------
# UI 界面
# --------------------------
class WordExtractorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Word 条款增量提取工具 v7.1 (Vertical Opt)")
        self.resize(950, 800)
        QApplication.setStyle(QStyleFactory.create('Fusion'))
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(242, 246, 250))
        palette.setColor(QPalette.WindowText, QColor(40, 40, 40))
        palette.setColor(QPalette.Base, QColor(255, 255, 255))
        palette.setColor(QPalette.Button, QColor(255, 255, 255))
        palette.setColor(QPalette.ButtonText, QColor(40, 40, 40))
        QApplication.setPalette(palette)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setSpacing(20)
        layout.setContentsMargins(40, 40, 40, 40)
        
        title_box = QVBoxLayout()
        title = QLabel("📝 Word 条款增量提取助手")
        title.setFont(QFont("Microsoft YaHei", 28, QFont.Bold)) 
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #2c3e50;")
        subtitle = QLabel("智能分表 • 纵向排版优化 • 自动去重")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: #7f8c8d; font-size: 15px; margin-top: 5px;")
        title_box.addWidget(title); title_box.addWidget(subtitle)
        layout.addLayout(title_box)
        
        card = QFrame()
        card.setStyleSheet("QFrame { background-color: #ffffff; border-radius: 15px; }")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20); shadow.setColor(QColor(0, 0, 0, 20)); shadow.setOffset(0, 5)
        card.setGraphicsEffect(shadow)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(15)
        card_layout.setContentsMargins(30, 30, 30, 30)
        
        input_style = "QLineEdit { border: 2px solid #ecf0f1; border-radius: 8px; padding: 10px; background: #f9fbfd; font-size: 14px; color: #2c3e50; } QLineEdit:focus { border: 2px solid #3498db; background: #fff; }"
        btn_style = "QPushButton { background: #f8f9fa; border: 1px solid #dfe6e9; border-radius: 8px; padding: 10px; font-weight: bold; color: #2d3436; } QPushButton:hover { background: #e2e6ea; }"
        
        row1 = QHBoxLayout()
        self.word_input = QLineEdit()
        self.word_input.setPlaceholderText("选择包含所有 Word 文档的文件夹...")
        self.word_input.setStyleSheet(input_style)
        btn1 = QPushButton("📂 来源文件夹")
        btn1.setCursor(Qt.PointingHandCursor)
        btn1.setStyleSheet(btn_style)
        btn1.clicked.connect(self.browse_word_folder)
        row1.addWidget(QLabel("文档来源:"))
        row1.addWidget(self.word_input, 1)
        row1.addWidget(btn1)
        card_layout.addLayout(row1)

        row_hist = QHBoxLayout()
        self.history_input = QLineEdit()
        self.history_input.setPlaceholderText("（可选）选择已整理好的 Excel，程序将跳过里面已有的文件...")
        self.history_input.setStyleSheet(input_style)
        btn_hist = QPushButton("📚 历史记录Excel")
        btn_hist.setCursor(Qt.PointingHandCursor)
        btn_hist.setStyleSheet(btn_style)
        btn_hist.clicked.connect(self.browse_history_file)
        row_hist.addWidget(QLabel("增量对比:"))
        row_hist.addWidget(self.history_input, 1)
        row_hist.addWidget(btn_hist)
        card_layout.addLayout(row_hist)
        
        row2 = QHBoxLayout()
        self.excel_input = QLineEdit()
        self.excel_input.setPlaceholderText("设置新提取内容的保存路径...")
        self.excel_input.setStyleSheet(input_style)
        btn2 = QPushButton("💾 新增结果保存")
        btn2.setCursor(Qt.PointingHandCursor)
        btn2.setStyleSheet(btn_style)
        btn2.clicked.connect(self.browse_excel_path)
        row2.addWidget(QLabel("结果输出:"))
        row2.addWidget(self.excel_input, 1)
        row2.addWidget(btn2)
        card_layout.addLayout(row2)
        
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setStyleSheet("background-color: #f0f2f5; border: none; height: 1px;")
        card_layout.addWidget(line)
        
        row3 = QHBoxLayout()
        row3.addWidget(QLabel("输出格式:"))
        self.fmt_horiz = QRadioButton("横向汇总表 (适合数据分析)")
        self.fmt_vert = QRadioButton("纵向清单表 (适合阅读打印)")
        self.fmt_vert.setChecked(True) # 默认选中纵向
        row3.addSpacing(10); row3.addWidget(self.fmt_horiz)
        row3.addSpacing(20); row3.addWidget(self.fmt_vert)
        row3.addStretch()
        card_layout.addLayout(row3)
        layout.addWidget(card)
        
        btn_layout = QHBoxLayout(); btn_layout.setSpacing(20)
        self.start_btn = QPushButton("🚀 开始增量提取")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.setMinimumHeight(60)
        self.start_btn.setStyleSheet("QPushButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3498db, stop:1 #2980b9); color: white; font-size: 18px; font-weight: bold; border-radius: 30px; border: 1px solid #2980b9; } QPushButton:hover { margin-top: -2px; } QPushButton:pressed { margin-top: 2px; } QPushButton:disabled { background: #bdc3c7; border: 1px solid #bdc3c7; }")
        self.start_btn.clicked.connect(self.start_process)
        
        self.open_folder_btn = QPushButton("📂 打开结果")
        self.open_folder_btn.setCursor(Qt.PointingHandCursor)
        self.open_folder_btn.setMinimumHeight(60)
        self.open_folder_btn.setEnabled(False)
        self.open_folder_btn.setStyleSheet("QPushButton { background: white; color: #2c3e50; font-size: 16px; font-weight: bold; border-radius: 30px; border: 2px solid #bdc3c7; } QPushButton:hover { border-color: #3498db; color: #3498db; } QPushButton:disabled { color: #bdc3c7; border-color: #ecf0f1; }")
        self.open_folder_btn.clicked.connect(self.open_output_folder)
        
        btn_layout.addWidget(self.start_btn, 2); btn_layout.addWidget(self.open_folder_btn, 1)
        layout.addLayout(btn_layout)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("QProgressBar { border: none; background: #e0e0e0; border-radius: 3px; height: 6px; } QProgressBar::chunk { background-color: #2ecc71; border-radius: 3px; }")
        layout.addWidget(self.progress_bar)
        
        self.log_text = QTextEdit(); self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("QTextEdit { background-color: #2c3e50; border-radius: 10px; font-family: 'Menlo', monospace; font-size: 12px; padding: 15px; color: #ecf0f1; }")
        layout.addWidget(self.log_text, 1)

    def browse_word_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择 Word 文件夹")
        if folder:
            self.word_input.setText(folder)
            if not self.excel_input.text(): self.excel_input.setText(os.path.join(folder, "新增附加条款_增量.xlsx"))

    def browse_history_file(self):
        f, _ = QFileDialog.getOpenFileName(self, "选择已整理的 Excel", "", "Excel Files (*.xlsx *.xls)")
        if f: self.history_input.setText(f)

    def browse_excel_path(self):
        f, _ = QFileDialog.getSaveFileName(self, "保存新增结果", "新增附加条款_增量.xlsx", "Excel Files (*.xlsx)")
        if f: self.excel_input.setText(f)

    def append_log(self, message, level="info"):
        colors = {"info": "#bdc3c7", "success": "#2ecc71", "error": "#e74c3c", "warning": "#f1c40f"}
        self.log_text.append(f'<span style="color:{colors.get(level, "#bdc3c7")}">{message}</span>')
        self.log_text.moveCursor(QTextCursor.End)

    def start_process(self):
        wf = self.word_input.text().strip(); ep = self.excel_input.text().strip(); hist = self.history_input.text().strip()
        if not wf or not ep:
            QMessageBox.warning(self, "提示", "请至少选择来源文件夹和输出路径！")
            return
        self.start_btn.setEnabled(False); self.open_folder_btn.setEnabled(False); self.start_btn.setText("⏳ 分析对比中..."); self.progress_bar.setVisible(True); self.progress_bar.setValue(0); self.log_text.clear()
        fmt = 'vertical' if self.fmt_vert.isChecked() else 'horizontal'
        self.worker = ExtractWorker(wf, ep, hist, fmt)
        self.worker.log_signal.connect(self.append_log)
        self.worker.progress_signal.connect(lambda c, t: self.progress_bar.setValue(int(c/t*100)))
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def on_finished(self, success, msg, ok_count, total_count):
        self.start_btn.setEnabled(True); self.start_btn.setText("🚀 开始增量提取"); self.progress_bar.setVisible(False)
        if success:
            if ok_count > 0:
                self.open_folder_btn.setEnabled(True)
                self.open_folder_btn.setStyleSheet("QPushButton { background: white; color: #27ae60; font-size: 16px; font-weight: bold; border-radius: 30px; border: 2px solid #27ae60; } QPushButton:hover { background: #27ae60; color: white; }")
                QMessageBox.information(self, "完成", f"共发现 {total_count} 个新文件\n成功提取: {ok_count} 个")
            else: QMessageBox.information(self, "完成", "没有发现任何新增的有效条款文件。")
        else: QMessageBox.critical(self, "错误", f"失败:\n{msg}")

    def open_output_folder(self):
        path = self.excel_input.text().strip()
        if path and os.path.exists(path): QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))

def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'): QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'): QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    w = WordExtractorGUI()
    w.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
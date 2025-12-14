# -*- coding: utf-8 -*-
"""
Word 条款提取工具 (Word Extractor GUI) - Modern UI Edition

Author: Google Senior Architect
Date: 2025-12-08
"""

import sys
import os
import re
import subprocess
import tempfile
import platform
from pathlib import Path
from typing import List, Dict

# 第三方库导入
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from docx import Document

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QRadioButton, QProgressBar, 
    QTextEdit, QFileDialog, QMessageBox, QStyleFactory, QFrame,
    QGraphicsDropShadowEffect, QSpacerItem, QSizePolicy
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt5.QtGui import QFont, QPalette, QColor, QTextCursor, QIcon

# --------------------------
# 跨平台转换处理器 (逻辑保持不变)
# --------------------------
class PlatformHandler:
    @staticmethod
    def is_windows():
        return platform.system() == "Windows"

    @staticmethod
    def convert_doc_to_docx(doc_path: Path) -> Path:
        temp_dir = tempfile.gettempdir()
        temp_docx_name = f"converted_{doc_path.stem}.docx"
        temp_docx_path = Path(temp_dir) / temp_docx_name

        if temp_docx_path.exists():
            try:
                os.remove(temp_docx_path)
            except:
                pass

        if PlatformHandler.is_windows():
            try:
                import win32com.client
                import pythoncom
                pythoncom.CoInitialize()
                word = win32com.client.Dispatch("Word.Application")
                word.Visible = False
                word.DisplayAlerts = False
                try:
                    doc = word.Documents.Open(str(doc_path))
                    doc.SaveAs2(str(temp_docx_path), FileFormat=16)
                    doc.Close()
                except Exception as e:
                    raise Exception(f"Word 转换内部错误: {e}")
                finally:
                    word.Quit()
                return temp_docx_path
            except ImportError:
                raise Exception("缺少 pywin32 库，请运行 pip install pywin32")
            except Exception as e:
                raise Exception(f"Windows 转换失败 (请确保安装了 Office): {e}")
        else:
            cmd = ['textutil', '-convert', 'docx', str(doc_path), '-output', str(temp_docx_path)]
            try:
                subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                return temp_docx_path
            except subprocess.CalledProcessError as e:
                raise Exception(f"macOS 转换失败: {e}")

class WordExtractorProcessor:
    def extract_clause_info(self, doc_path: Path) -> Dict[str, str]:
        file_name = doc_path.name
        clause_name = doc_path.stem
        result = {'FileName': file_name, 'ClauseName': clause_name, 'RegistrationNo': '', 'Content': '', 'Error': ''}
        temp_file_to_remove = None

        try:
            target_path = doc_path
            if doc_path.suffix.lower() == '.doc':
                try:
                    target_path = PlatformHandler.convert_doc_to_docx(doc_path)
                    temp_file_to_remove = target_path
                except Exception as e:
                    result['Error'] = str(e)
                    return result

            if not os.path.exists(target_path):
                 result['Error'] = "转换后的文件丢失"
                 return result
                 
            doc = Document(str(target_path))
            paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

            if not paragraphs:
                result['Error'] = '文档内容为空'
                return result

            registration_no = ""
            if len(paragraphs) >= 3:
                third_line = paragraphs[2]
                match = re.search(r'[（\(]([^）\)]+)[）\)]', third_line)
                if match:
                    registration_no = match.group(1)
                else:
                    registration_no = third_line.replace('产品注册号:', '').replace('注册号:', '').replace('**', '').strip()
            result['RegistrationNo'] = registration_no

            content_lines = []
            if len(paragraphs) >= 4:
                for para in paragraphs[3:]:
                    clean_para = para.replace('**', '').strip()
                    if clean_para:
                        content_lines.append(clean_para)
            result['Content'] = '\n'.join(content_lines)
            return result
        except Exception as e:
            result['Error'] = f"解析失败: {str(e)}"
            return result
        finally:
            if temp_file_to_remove and temp_file_to_remove.exists():
                try:
                    os.remove(temp_file_to_remove)
                except:
                    pass

    @staticmethod
    def get_word_files(directory: str) -> List[Path]:
        path_obj = Path(directory)
        word_files = []
        for ext in ['*.docx', '*.doc']:
            word_files.extend(path_obj.rglob(ext))
        filtered = []
        for f in word_files:
            if not f.name.startswith('~') and '费率方案' not in f.name:
                filtered.append(f)
        return filtered

    def save_to_excel(self, data_list: List[Dict], output_file: str, format_type: str = 'horizontal') -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '保险条款汇总'

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_wrap = Alignment(wrap_text=True, vertical='top', horizontal='left')
        align_center = Alignment(horizontal='center', vertical='center')
        header_font = Font(bold=True, size=12)
        
        if format_type == 'horizontal':
            headers = ['条款名称', '条款内容', '产品注册号', '原文件名', '错误信息']
            col_widths = [30, 80, 25, 20, 20]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            fill_header = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
            for row_idx, item in enumerate(data_list, 2):
                row_data = [item.get('ClauseName', ''), item.get('Content', ''), item.get('RegistrationNo', ''), item.get('FileName', ''), item.get('Error', '')]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = align_wrap
                    cell.border = thin_border
        else: 
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 30
            fill_title = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            current_row = 1
            for item in data_list:
                if item.get('Error'): continue
                cell_name = ws.cell(row=current_row, column=1, value=item.get('ClauseName', ''))
                cell_name.font = header_font
                cell_name.fill = fill_title
                cell_name.border = thin_border
                cell_reg = ws.cell(row=current_row, column=2, value=item.get('RegistrationNo', ''))
                cell_reg.border = thin_border
                cell_reg.alignment = align_center
                current_row += 1
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                cell_content = ws.cell(row=current_row, column=1, value=item.get('Content', ''))
                cell_content.alignment = align_wrap
                cell_content.border = thin_border
                current_row += 2
        wb.save(output_file)

class WorkerThread(QThread):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    finished_signal = pyqtSignal(bool, str, int, int)
    
    def __init__(self, word_folder: str, excel_path: str, format_type: str):
        super().__init__()
        self.word_folder = word_folder
        self.excel_path = excel_path
        self.format_type = format_type
        self.processor = WordExtractorProcessor()
    
    def run(self):
        try:
            self.log_signal.emit("初始化处理引擎...", "info")
            self.log_signal.emit(f"正在扫描目录: {self.word_folder}", "info")
            files = self.processor.get_word_files(self.word_folder)
            if not files:
                self.finished_signal.emit(False, "未找到有效的 Word 文件 (.docx/.doc)", 0, 0)
                return
            
            total_files = len(files)
            self.log_signal.emit(f"扫描完成，待处理文件数: {total_files}", "success")
            processed_data = []
            success_count = 0
            
            for i, file_path in enumerate(files, 1):
                self.progress_signal.emit(i, total_files)
                self.log_signal.emit(f"[{i}/{total_files}] 处理: {file_path.name}", "info")
                if file_path.suffix.lower() == '.doc':
                     self.log_signal.emit(f"  ➜ 正在自动转换 .doc 格式...", "warning")
                data = self.processor.extract_clause_info(file_path)
                processed_data.append(data)
                if data['Error']:
                    self.log_signal.emit(f"  ✗ 错误: {data['Error']}", "error")
                else:
                    success_count += 1
                    preview = (data['ClauseName'][:30] + '...') if len(data['ClauseName']) > 30 else data['ClauseName']
                    self.log_signal.emit(f"  ✓ 提取成功: {preview}", "success")
            
            self.log_signal.emit("正在写入 Excel 文件...", "info")
            self.processor.save_to_excel(processed_data, self.excel_path, self.format_type)
            self.log_signal.emit(f"全部完成! 成功: {success_count}/{total_files}", "success")
            self.finished_signal.emit(True, self.excel_path, success_count, total_files)
        except Exception as e:
            import traceback
            error_msg = f"发生未预期的错误:\n{traceback.format_exc()}"
            self.log_signal.emit("发生严重错误，进程终止。", "error")
            self.finished_signal.emit(False, str(e), 0, 0)

# --------------------------
# 全新设计的 UI 类
# --------------------------
class WordExtractorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Word 条款提取工具 Pro v3.0 (Modern UI)")
        self.resize(950, 720)
        
        # 强制亮色模式 & Fusion 风格
        QApplication.setStyle(QStyleFactory.create('Fusion'))
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(242, 246, 250)) # 极淡的灰蓝色背景，很护眼
        palette.setColor(QPalette.WindowText, QColor(40, 40, 40))
        palette.setColor(QPalette.Base, QColor(255, 255, 255))
        palette.setColor(QPalette.Text, QColor(40, 40, 40))
        palette.setColor(QPalette.Button, QColor(255, 255, 255))
        palette.setColor(QPalette.ButtonText, QColor(40, 40, 40))
        QApplication.setPalette(palette)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # 主布局
        main_layout = QVBoxLayout(main_widget)
        main_layout.setSpacing(25)
        main_layout.setContentsMargins(40, 40, 40, 40)
        
        # --- 1. 顶部标题区 ---
        title_container = QWidget()
        title_layout = QVBoxLayout(title_container)
        title_layout.setSpacing(5)
        title_layout.setContentsMargins(0, 0, 0, 0)
        
        title_lbl = QLabel("📝 Word 条款智能提取助手")
        title_lbl.setFont(QFont("Microsoft YaHei", 28, QFont.Bold)) 
        title_lbl.setAlignment(Qt.AlignCenter)
        title_lbl.setStyleSheet("color: #2c3e50;")
        
        subtitle_lbl = QLabel("支持 .docx 及 .doc 自动转换 • 智能提取注册号 • 极速导出")
        subtitle_lbl.setAlignment(Qt.AlignCenter)
        subtitle_lbl.setStyleSheet("color: #7f8c8d; font-size: 15px; margin-top: 5px;")
        
        title_layout.addWidget(title_lbl)
        title_layout.addWidget(subtitle_lbl)
        main_layout.addWidget(title_container)

        # --- 2. 核心配置卡片 (Card Design) ---
        # 弃用 QGroupBox，改用 QFrame + 阴影
        card_widget = QFrame()
        card_widget.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border-radius: 15px;
            }
        """)
        # 添加阴影效果
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 20))
        shadow.setOffset(0, 5)
        card_widget.setGraphicsEffect(shadow)
        
        card_layout = QVBoxLayout(card_widget)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(30, 30, 30, 30)
        
        # 样式定义
        label_style = "font-size: 14px; font-weight: bold; color: #34495e;"
        input_style = """
            QLineEdit {
                border: 2px solid #ecf0f1; border-radius: 8px; padding: 10px 15px;
                background-color: #f9fbfd; font-size: 14px; color: #2c3e50;
            }
            QLineEdit:focus { border: 2px solid #3498db; background-color: #fff; }
        """
        btn_browse_style = """
            QPushButton {
                background-color: #f8f9fa; border: 1px solid #dfe6e9; border-radius: 8px;
                padding: 10px 15px; font-size: 13px; color: #2d3436; font-weight: bold;
            }
            QPushButton:hover { background-color: #e2e6ea; border-color: #b2bec3; }
            QPushButton:pressed { background-color: #dae0e5; }
        """
        
        # 2.1 文档输入行
        row1 = QHBoxLayout()
        lbl1 = QLabel("文档来源")
        lbl1.setStyleSheet(label_style)
        self.word_input = QLineEdit()
        self.word_input.setPlaceholderText("请选择包含 Word 文档的文件夹...")
        self.word_input.setStyleSheet(input_style)
        btn1 = QPushButton("📂 选择文件夹")
        btn1.setCursor(Qt.PointingHandCursor)
        btn1.setStyleSheet(btn_browse_style)
        btn1.clicked.connect(self.browse_word_folder)
        
        row1.addWidget(lbl1)
        row1.addWidget(self.word_input, 1) # 1代表拉伸比例
        row1.addWidget(btn1)
        card_layout.addLayout(row1)
        
        # 2.2 Excel 输出行
        row2 = QHBoxLayout()
        lbl2 = QLabel("结果输出")
        lbl2.setStyleSheet(label_style)
        self.excel_input = QLineEdit()
        self.excel_input.setPlaceholderText("请设置 Excel 保存路径...")
        self.excel_input.setStyleSheet(input_style)
        btn2 = QPushButton("💾 保存路径")
        btn2.setCursor(Qt.PointingHandCursor)
        btn2.setStyleSheet(btn_browse_style)
        btn2.clicked.connect(self.browse_excel_path)
        
        row2.addWidget(lbl2)
        row2.addWidget(self.excel_input, 1)
        row2.addWidget(btn2)
        card_layout.addLayout(row2)
        
        # 分割线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("background-color: #f0f2f5; border: none; height: 1px;")
        card_layout.addWidget(line)
        
        # 2.3 格式选择
        row3 = QHBoxLayout()
        lbl3 = QLabel("输出格式")
        lbl3.setStyleSheet(label_style)
        
        radio_style = """
            QRadioButton { font-size: 14px; color: #2c3e50; spacing: 8px; }
            QRadioButton::indicator { width: 18px; height: 18px; border-radius: 9px; border: 2px solid #bdc3c7; }
            QRadioButton::indicator:checked { border: 2px solid #3498db; background-color: #3498db; }
        """
        self.fmt_horiz = QRadioButton("横向汇总表 (适合数据分析)")
        self.fmt_horiz.setChecked(True)
        self.fmt_horiz.setStyleSheet(radio_style)
        self.fmt_horiz.setCursor(Qt.PointingHandCursor)
        
        self.fmt_vert = QRadioButton("纵向清单表 (适合阅读打印)")
        self.fmt_vert.setStyleSheet(radio_style)
        self.fmt_vert.setCursor(Qt.PointingHandCursor)
        
        row3.addWidget(lbl3)
        row3.addSpacing(15)
        row3.addWidget(self.fmt_horiz)
        row3.addSpacing(20)
        row3.addWidget(self.fmt_vert)
        row3.addStretch()
        card_layout.addLayout(row3)
        
        main_layout.addWidget(card_widget)
        
        # --- 3. 启动按钮 ---
        # 悬浮的大按钮设计
        self.start_btn = QPushButton("🚀 开始提取数据")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.setMinimumHeight(60)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3498db, stop:1 #2980b9);
                color: white; font-size: 18px; font-weight: bold; border-radius: 30px;
                border: 1px solid #2980b9;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #5dade2, stop:1 #3498db);
                margin-top: -2px; /* 悬浮上移效果 */
            }
            QPushButton:pressed {
                background: #2573a7; margin-top: 2px;
            }
            QPushButton:disabled {
                background: #bdc3c7; border: 1px solid #bdc3c7; color: #fff;
            }
        """)
        self.start_btn.clicked.connect(self.start_process)
        main_layout.addWidget(self.start_btn)
        
        # 进度条 (紧贴按钮下方)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none; background: #e0e0e0; border-radius: 3px; height: 6px; text-align: center;
            }
            QProgressBar::chunk { background-color: #2ecc71; border-radius: 3px; }
        """)
        main_layout.addWidget(self.progress_bar)
        
        # --- 4. 日志区域 ---
        # 仿终端风格，但用圆角
        log_widget = QWidget()
        log_lay = QVBoxLayout(log_widget)
        log_lay.setContentsMargins(0, 10, 0, 0)
        
        log_lbl = QLabel("运行日志")
        log_lbl.setStyleSheet("font-size: 13px; font-weight: bold; color: #7f8c8d; margin-left: 5px;")
        log_lay.addWidget(log_lbl)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #2c3e50; border-radius: 10px;
                font-family: 'Menlo', 'Monaco', 'Consolas', monospace; 
                font-size: 12px; padding: 15px; color: #ecf0f1;
            }
        """)
        log_lay.addWidget(self.log_text)
        
        main_layout.addWidget(log_widget, 1) # 1 表示占用剩余所有空间

    def browse_word_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择 Word 文件夹")
        if folder:
            self.word_input.setText(folder)
            if not self.excel_input.text():
                default_excel = os.path.join(folder, "保险条款汇总.xlsx")
                self.excel_input.setText(default_excel)

    def browse_excel_path(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "保存 Excel", "保险条款汇总.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            self.excel_input.setText(file_path)

    def append_log(self, message: str, level: str = "info"):
        # 日志颜色微调，适应深色背景
        colors = {
            "info": "#bdc3c7",     # 浅灰
            "success": "#2ecc71",  # 鲜绿
            "error": "#e74c3c",    # 亮红
            "warning": "#f1c40f"   # 金黄
        }
        color_hex = colors.get(level, "#bdc3c7")
        self.log_text.append(f'<span style="color:{color_hex}">{message}</span>')
        self.log_text.moveCursor(QTextCursor.End)

    def start_process(self):
        word_dir = self.word_input.text().strip()
        excel_path = self.excel_input.text().strip()
        
        if not word_dir or not os.path.exists(word_dir):
            QMessageBox.warning(self, "路径错误", "请选择有效的 Word 文件夹！")
            return
        
        if not excel_path:
            QMessageBox.warning(self, "路径错误", "请设置 Excel 保存路径！")
            return

        self.start_btn.setEnabled(False)
        self.start_btn.setText("⏳ 正在处理中...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        
        fmt = 'vertical' if self.fmt_vert.isChecked() else 'horizontal'
        
        self.worker = WorkerThread(word_dir, excel_path, fmt)
        self.worker.log_signal.connect(self.append_log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def update_progress(self, current, total):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)

    def on_finished(self, success, msg, ok_count, total_count):
        self.start_btn.setEnabled(True)
        self.start_btn.setText("🚀 开始提取数据")
        self.progress_bar.setVisible(False)
        
        if success:
            QMessageBox.information(self, "处理完成", f"成功提取: {ok_count}/{total_count}\n文件已保存至:\n{msg}")
        else:
            QMessageBox.critical(self, "处理失败", f"错误信息:\n{msg}")

def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    
    app = QApplication(sys.argv)
    window = WordExtractorGUI()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
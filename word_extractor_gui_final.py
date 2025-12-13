# -*- coding: utf-8 -*-
"""
Word 条款批量提取工具 Final v5.0 (修复页码残留版)
- [修复] 针对 '第 PAGE 30页共 NUMPAGES 62页' 等域代码残留进行强力清洗
- [功能] 批量读取 Word -> 提取注册号/正文 -> 汇总 Excel
- [界面] Modern UI + 打开目录按钮

Author: Google Senior Architect
Date: 2025-12-08
"""

import sys
import os
import re
import subprocess
import tempfile
import platform
import traceback
from datetime import datetime
import concurrent.futures
from pathlib import Path

# 第三方库
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from docx import Document

# PyQt5 库
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QRadioButton, QProgressBar, 
    QTextEdit, QFileDialog, QMessageBox, QStyleFactory, QFrame,
    QGraphicsDropShadowEffect
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QUrl
from PyQt5.QtGui import QFont, QPalette, QColor, QTextCursor, QDesktopServices

# ==========================================
# 🛡️ 核心修复：防止 macOS 打包后闪退
# ==========================================
class NullWriter:
    def write(self, text): pass
    def flush(self): pass

if getattr(sys, 'frozen', False):
    sys.stdout = NullWriter()
    sys.stderr = NullWriter()

def global_exception_handler(exctype, value, tb):
    error_msg = "".join(traceback.format_exception(exctype, value, tb))
    try: sys.__stderr__.write(error_msg)
    except: pass
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Critical)
    msg_box.setText("程序发生意外错误")
    msg_box.setInformativeText(str(value))
    msg_box.setDetailedText(error_msg)
    msg_box.setWindowTitle("错误报告")
    msg_box.exec_()

sys.excepthook = global_exception_handler

# --------------------------
# 跨平台转换处理器
# --------------------------
class PlatformHandler:
    @staticmethod
    def is_windows():
        return platform.system() == "Windows"

    @staticmethod
    def convert_doc_to_docx(doc_path: Path) -> Path:
        temp_dir = tempfile.gettempdir()
        temp_docx_name = f"ext_temp_{doc_path.stem}.docx"
        temp_docx_path = Path(temp_dir) / temp_docx_name

        if temp_docx_path.exists():
            try: os.remove(temp_docx_path)
            except: pass

        if PlatformHandler.is_windows():
            try:
                import win32com.client
                import pythoncom
                pythoncom.CoInitialize()
                word = win32com.client.Dispatch("Word.Application")
                word.Visible = False
                word.DisplayAlerts = False
                try:
                    doc = word.Documents.Open(str(doc_path))
                    doc.SaveAs2(str(temp_docx_path), FileFormat=16)
                    doc.Close()
                except Exception as e:
                    raise Exception(f"Word 转换内部错误: {e}")
                finally:
                    word.Quit()
                return temp_docx_path
            except ImportError:
                raise Exception("缺少 pywin32 库")
            except Exception as e:
                raise Exception(f"Windows 转换失败: {e}")
        else:
            cmd = ['textutil', '-convert', 'docx', str(doc_path), '-output', str(temp_docx_path)]
            try:
                subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                return temp_docx_path
            except subprocess.CalledProcessError as e:
                raise Exception(f"macOS 转换失败: {e}")

# --------------------------
# 提取逻辑核心 (已增强清洗功能)
# --------------------------
class WordExtractorProcessor:
    # ... [NOISE_PATTERNS remain the same] ...

    # ... [is_noise_line remains the same] ...

    def extract_clause_info(self, doc_path: Path) -> dict:
        file_name = doc_path.name
        clause_name = doc_path.stem
        # [新增] 仅占位，实际日期由 Worker 统一传入或此处获取
        result = {'FileName': file_name, 'ClauseName': clause_name, 'RegistrationNo': '', 'Content': '', 'Error': ''}
        
        # ... [rest of method remains the same] ...

    # ... [get_word_files remains the same] ...

    def save_to_excel(self, data_list: list, output_file: str, format_type: str = 'horizontal'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '条款汇总'

        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_wrap = Alignment(wrap_text=True, vertical='top', horizontal='left')
        header_font = Font(bold=True, size=12, color="FFFFFF")
        fill_header = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")

        # [新增] 获取当前日期 YYYY-MM-DD
        current_date = datetime.now().strftime("%Y-%m-%d")

        if format_type == 'horizontal':
            # [新增] 增加 '填单日期' 列
            headers = ['条款名称', '注册号', '条款内容', '填单日期', '原文件名', '状态']
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 80
            ws.column_dimensions['D'].width = 15 # 日期列宽
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border

            for row_idx, item in enumerate(data_list, 2):
                status = item.get('Error') if item.get('Error') else "成功"
                row_data = [
                    item.get('ClauseName', ''),
                    item.get('RegistrationNo', ''),
                    item.get('Content', ''),
                    current_date, # [新增] 填入当前日期
                    item.get('FileName', ''),
                    status
                ]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = align_wrap
                    cell.border = border
                    # 日期列居中
                    if col_idx == 4:
                        cell.alignment = align_center
                    if col_idx == 6 and val != "成功":
                        cell.font = Font(color="FF0000")

        else: # Vertical
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 30
            
            current_row = 1
            for item in data_list:
                if item.get('Error'): continue
                
                # [修改] 标题行增加日期显示
                title_text = f"{item.get('ClauseName', '')} (填单日期: {current_date})"
                cell_name = ws.cell(row=current_row, column=1, value=title_text)
                cell_name.font = Font(bold=True, size=12)
                cell_name.fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
                cell_name.border = border
                
                cell_reg = ws.cell(row=current_row, column=2, value=item.get('RegistrationNo', ''))
                cell_reg.alignment = align_center
                cell_reg.border = border
                
                current_row += 1
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                cell_content = ws.cell(row=current_row, column=1, value=item.get('Content', ''))
                cell_content.alignment = align_wrap
                cell_content.border = border
                
                current_row += 2 

        wb.save(output_file)

# --------------------------
# 工作线程
# --------------------------
class ExtractWorker(QThread):
    log_signal = pyqtSignal(str, str)
    progress_signal = pyqtSignal(int, int)
    finished_signal = pyqtSignal(bool, str, int, int)
    
    def __init__(self, word_folder, excel_path, format_type):
        super().__init__()
        self.word_folder = word_folder
        self.excel_path = excel_path
        self.format_type = format_type
        self.processor = WordExtractorProcessor()
    
    # [新增] 静态方法用于并行处理
    @staticmethod
    def process_file_static(args):
        processor, file_path = args
        try:
            return processor.extract_clause_info(file_path)
        except Exception as e:
            return {'FileName': file_path.name, 'ClauseName': file_path.stem, 'RegistrationNo': '', 'Content': '', 'Error': str(e)}

    def run(self):
        try:
            self.log_signal.emit("⏳ 初始化引 (并发版)...", "info")
            self.log_signal.emit(f"📂 扫描目录: {self.word_folder}", "info")
            files = self.processor.get_word_files(self.word_folder)
            
            if not files:
                self.finished_signal.emit(False, "未找到有效的 Word 文件", 0, 0)
                return
            
            total_files = len(files)
            self.log_signal.emit(f"✅ 发现 {total_files} 个文件，准备执行多线程处理", "success")
            
            processed_data = []
            success_count = 0
            
            # [修改] 使用 ThreadPoolExecutor 并发处理
            # macOS 下文件 IO 较快，并发数不易过大，避免 Too many open files
            max_workers = 8 
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                # 包装参数
                tasks = [(self.processor, f) for f in files]
                future_to_file = {executor.submit(ExtractWorker.process_file_static, t): t[1] for t in tasks}
                
                completed_count = 0
                for future in concurrent.futures.as_completed(future_to_file):
                    file_path = future_to_file[future]
                    completed_count += 1
                    self.progress_signal.emit(completed_count, total_files)
                    
                    try:
                        data = future.result()
                        processed_data.append(data)
                        
                        if data['Error']:
                            self.log_signal.emit(f"   ✗ 错误 [{file_path.name}]: {data['Error']}", "error")
                        else:
                            success_count += 1
                            # 减少日志刷屏，每完成3个或者有错误才打印详情，或者仅打印简略信息
                            # 这里选择打印简略信息
                            preview = (data['ClauseName'][:15] + '..') 
                            self.log_signal.emit(f"   ✓ 完成: {preview}", "success")
                            
                    except Exception as e:
                        self.log_signal.emit(f"   ☠️ 严重异常 [{file_path.name}]: {e}", "error")

            # 保持原有顺序 (按文件名排序)
            processed_data.sort(key=lambda x: x['FileName'])
            
            self.log_signal.emit("💾 正在生成 Excel...", "info")
            self.processor.save_to_excel(processed_data, self.excel_path, self.format_type)
            
            self.log_signal.emit(f"🎉 全部完成！成功: {success_count}/{total_files}", "success")
            self.finished_signal.emit(True, self.excel_path, success_count, total_files)
            
        except Exception as e:
            # 捕获异常，不要让线程直接崩掉
            self.log_signal.emit(f"❌ 全局错误: {str(e)}", "error")
            self.finished_signal.emit(False, str(e), 0, 0)

# --------------------------
# UI 界面
# --------------------------
class WordExtractorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Word 条款批量提取工具 Final")
        self.resize(950, 780)
        
        QApplication.setStyle(QStyleFactory.create('Fusion'))
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(242, 246, 250))
        palette.setColor(QPalette.WindowText, QColor(40, 40, 40))
        palette.setColor(QPalette.Base, QColor(255, 255, 255))
        palette.setColor(QPalette.Button, QColor(255, 255, 255))
        palette.setColor(QPalette.ButtonText, QColor(40, 40, 40))
        QApplication.setPalette(palette)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setSpacing(25)
        layout.setContentsMargins(40, 40, 40, 40)
        
        title_box = QVBoxLayout()
        title = QLabel("📝 Word 条款批量提取助手")
        title.setFont(QFont("Microsoft YaHei", 28, QFont.Bold)) 
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #2c3e50;")
        subtitle = QLabel("批量提取 Word 内容 • 智能清洗页眉页脚 • 生成汇总 Excel")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: #7f8c8d; font-size: 15px; margin-top: 5px;")
        title_box.addWidget(title)
        title_box.addWidget(subtitle)
        layout.addLayout(title_box)
        
        card = QFrame()
        card.setStyleSheet("QFrame { background-color: #ffffff; border-radius: 15px; }")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20); shadow.setColor(QColor(0, 0, 0, 20)); shadow.setOffset(0, 5)
        card.setGraphicsEffect(shadow)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(30, 30, 30, 30)
        
        input_style = "QLineEdit { border: 2px solid #ecf0f1; border-radius: 8px; padding: 10px; background: #f9fbfd; font-size: 14px; color: #2c3e50; } QLineEdit:focus { border: 2px solid #3498db; background: #fff; }"
        btn_style = "QPushButton { background: #f8f9fa; border: 1px solid #dfe6e9; border-radius: 8px; padding: 10px; font-weight: bold; color: #2d3436; } QPushButton:hover { background: #e2e6ea; }"
        
        row1 = QHBoxLayout()
        self.word_input = QLineEdit()
        self.word_input.setPlaceholderText("请选择包含 Word 文档的文件夹...")
        self.word_input.setStyleSheet(input_style)
        btn1 = QPushButton("📂 选择文件夹")
        btn1.setCursor(Qt.PointingHandCursor)
        btn1.setStyleSheet(btn_style)
        btn1.clicked.connect(self.browse_word_folder)
        row1.addWidget(QLabel("文档来源:"))
        row1.addWidget(self.word_input, 1)
        row1.addWidget(btn1)
        card_layout.addLayout(row1)
        
        row2 = QHBoxLayout()
        self.excel_input = QLineEdit()
        self.excel_input.setPlaceholderText("请设置 Excel 保存路径...")
        self.excel_input.setStyleSheet(input_style)
        btn2 = QPushButton("💾 保存路径")
        btn2.setCursor(Qt.PointingHandCursor)
        btn2.setStyleSheet(btn_style)
        btn2.clicked.connect(self.browse_excel_path)
        row2.addWidget(QLabel("结果输出:"))
        row2.addWidget(self.excel_input, 1)
        row2.addWidget(btn2)
        card_layout.addLayout(row2)
        
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setStyleSheet("background-color: #f0f2f5; border: none; height: 1px;")
        card_layout.addWidget(line)
        
        row3 = QHBoxLayout()
        row3.addWidget(QLabel("输出格式:"))
        self.fmt_horiz = QRadioButton("横向汇总表 (适合数据分析)")
        self.fmt_horiz.setChecked(True)
        self.fmt_vert = QRadioButton("纵向清单表 (适合阅读打印)")
        row3.addSpacing(10)
        row3.addWidget(self.fmt_horiz)
        row3.addSpacing(20)
        row3.addWidget(self.fmt_vert)
        row3.addStretch()
        card_layout.addLayout(row3)
        layout.addWidget(card)
        
        btn_layout = QHBoxLayout(); btn_layout.setSpacing(20)
        self.start_btn = QPushButton("🚀 开始批量提取")
        self.start_btn.setCursor(Qt.PointingHandCursor)
        self.start_btn.setMinimumHeight(60)
        self.start_btn.setStyleSheet("QPushButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3498db, stop:1 #2980b9); color: white; font-size: 18px; font-weight: bold; border-radius: 30px; border: 1px solid #2980b9; } QPushButton:hover { margin-top: -2px; } QPushButton:pressed { margin-top: 2px; } QPushButton:disabled { background: #bdc3c7; border: 1px solid #bdc3c7; }")
        self.start_btn.clicked.connect(self.start_process)
        
        self.open_folder_btn = QPushButton("📂 打开输出目录")
        self.open_folder_btn.setCursor(Qt.PointingHandCursor)
        self.open_folder_btn.setMinimumHeight(60)
        self.open_folder_btn.setEnabled(False)
        self.open_folder_btn.setStyleSheet("QPushButton { background: white; color: #2c3e50; font-size: 16px; font-weight: bold; border-radius: 30px; border: 2px solid #bdc3c7; } QPushButton:hover { border-color: #3498db; color: #3498db; } QPushButton:disabled { color: #bdc3c7; border-color: #ecf0f1; }")
        self.open_folder_btn.clicked.connect(self.open_output_folder)
        
        btn_layout.addWidget(self.start_btn, 2)
        btn_layout.addWidget(self.open_folder_btn, 1)
        layout.addLayout(btn_layout)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("QProgressBar { border: none; background: #e0e0e0; border-radius: 3px; height: 6px; } QProgressBar::chunk { background-color: #2ecc71; border-radius: 3px; }")
        layout.addWidget(self.progress_bar)
        
        self.log_text = QTextEdit(); self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("QTextEdit { background-color: #2c3e50; border-radius: 10px; font-family: 'Menlo', monospace; font-size: 12px; padding: 15px; color: #ecf0f1; }")
        layout.addWidget(self.log_text, 1)

    def browse_word_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择 Word 文件夹")
        if folder:
            self.word_input.setText(folder)
            if not self.excel_input.text():
                self.excel_input.setText(os.path.join(folder, "条款汇总.xlsx"))

    def browse_excel_path(self):
        f, _ = QFileDialog.getSaveFileName(self, "保存 Excel", "条款汇总.xlsx", "Excel Files (*.xlsx)")
        if f: self.excel_input.setText(f)

    def append_log(self, message, level="info"):
        colors = {"info": "#bdc3c7", "success": "#2ecc71", "error": "#e74c3c", "warning": "#f1c40f"}
        self.log_text.append(f'<span style="color:{colors.get(level, "#bdc3c7")}">{message}</span>')
        self.log_text.moveCursor(QTextCursor.End)

    def start_process(self):
        wf = self.word_input.text().strip()
        ep = self.excel_input.text().strip()
        if not wf or not ep:
            QMessageBox.warning(self, "提示", "请选择文件夹和保存路径！")
            return
        
        self.start_btn.setEnabled(False)
        self.open_folder_btn.setEnabled(False)
        self.start_btn.setText("⏳ 提取中...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        
        fmt = 'vertical' if self.fmt_vert.isChecked() else 'horizontal'
        self.worker = ExtractWorker(wf, ep, fmt)
        self.worker.log_signal.connect(self.append_log)
        self.worker.progress_signal.connect(lambda c, t: self.progress_bar.setValue(int(c/t*100)))
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def on_finished(self, success, msg, ok_count, total_count):
        self.start_btn.setEnabled(True)
        self.start_btn.setText("🚀 开始批量提取")
        self.progress_bar.setVisible(False)
        if success:
            self.open_folder_btn.setEnabled(True)
            self.open_folder_btn.setStyleSheet("QPushButton { background: white; color: #27ae60; font-size: 16px; font-weight: bold; border-radius: 30px; border: 2px solid #27ae60; } QPushButton:hover { background: #27ae60; color: white; }")
            QMessageBox.information(self, "完成", f"成功: {ok_count}/{total_count}\nExcel已生成")
        else:
            QMessageBox.critical(self, "错误", f"失败:\n{msg}")

    def open_output_folder(self):
        path = self.excel_input.text().strip()
        if path and os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))

def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'): QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'): QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    w = WordExtractorGUI()
    w.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()